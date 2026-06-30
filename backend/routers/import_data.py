import csv
import io
import os
import re
from datetime import datetime, date
from decimal import Decimal, InvalidOperation

from fastapi import APIRouter, Depends, File, Form, UploadFile, Query
from sqlalchemy.orm import Session
from sqlalchemy import func

from database import get_db
from models import (
    DimCountry, DimProduct, DimProductCost, DimFreight, DimTime, DimStore, DimExchangeRate, MonthlySummary, RawTransaction,
    RawAdvertising, RawStorageFee, RawReturns, RawInbound, RawLongTermStorage, RawRemovalFee,
)
from config import UPLOAD_DIR

router = APIRouter()


def _detect_country_from_data(db: Session, header, rows):
    """从数据文件中自动检测国家: 优先查 marketplace/country_code/国家/地区 列，
    找不到时扫描所有列值"""
    mp_idx = None; cc_idx = None
    for i, h in enumerate(header):
        hl = h.lower().strip() if h else ""
        if hl == 'marketplace': mp_idx = i  # 精确匹配，避免 "marketplace withheld tax"
        if hl in ('country_code', 'country', '国家', '国家/地区'): cc_idx = i

    # 已知国家列的匹配模式（精确，可包容缩写）
    _col_patterns = [
        (['mx', 'mex', 'amazon.com.mx', 'mexico'], 'MX'),
        (['ca', 'can', 'amazon.ca', 'canada'], 'CA'),
        (['au', 'aus', 'amazon.com.au', 'australia'], 'AU'),
        (['na', 'north america', 'amazon.com', 'united states', 'us', 'usa'], 'US'),
        (['uk', 'gb', 'gbr', 'amazon.co.uk', 'united kingdom'], 'UK'),
        (['de', 'deu', 'amazon.de', 'germany'], 'DE'),
        (['fr', 'fra', 'amazon.fr', 'france'], 'FR'),
        (['es', 'esp', 'amazon.es', 'spain'], 'ES'),
        (['it', 'ita', 'amazon.it', 'italy'], 'IT'),
        (['nl', 'nld', 'amazon.nl', 'netherlands'], 'NL'),
        (['se', 'swe', 'amazon.se', 'sweden'], 'SE'),
        (['be', 'bel', 'amazon.com.be', 'belgium'], 'BE'),
        (['ie', 'irl', 'amazon.ie', 'ireland'], 'IE'),
        (['ae', 'are', 'amazon.ae', 'uae'], 'AE'),
        (['sa', 'sau', 'amazon.sa', 'saudi'], 'SA'),
    ]
    # 全列扫描的匹配模式（保守，仅匹配长关键词，防止 "us" 误匹配 "USB"、 "de" 误匹配 "code"）
    _full_scan_patterns = [
        (['amazon.com.mx', 'mexico'], 'MX'),
        (['amazon.ca', 'canada'], 'CA'),
        (['amazon.com.au', 'australia'], 'AU'),
        (['amazon.com', 'united states', 'usa'], 'US'),  # 不含 "us" 单字母
        (['amazon.co.uk', 'united kingdom'], 'UK'),
        (['amazon.de', 'germany', 'deu'], 'DE'),  # 不含 "de" 单字母
        (['amazon.fr', 'france', 'fra'], 'FR'),
        (['amazon.es', 'spain', 'esp'], 'ES'),
        (['amazon.it', 'italy', 'ita'], 'IT'),
        (['amazon.nl', 'netherlands', 'nld'], 'NL'),
        (['amazon.se', 'sweden', 'swe'], 'SE'),
        (['amazon.com.be', 'belgium', 'bel'], 'BE'),
        (['amazon.ie', 'ireland', 'irl'], 'IE'),
        (['amazon.ae', 'uae', 'are'], 'AE'),
        (['amazon.sa', 'saudi', 'sau'], 'SA'),
    ]

    def _match(v, patterns):
        v = v.lower().strip()
        for pats, code in patterns:
            if v in pats:
                return code
        return None

    # 第一轮：查已知国家列（marketplace / country_code / 国家/地区）
    for row in rows[:20]:
        vals = []
        if mp_idx is not None and len(row) > mp_idx: vals.append(str(row[mp_idx] or ''))
        if cc_idx is not None and len(row) > cc_idx: vals.append(str(row[cc_idx] or ''))
        for v in vals:
            code = _match(v, _col_patterns)
            if code: return code

    # 第二轮：全列扫描（保守模式，避免误匹配）
    for row in rows[:20]:
        for cell in row:
            if cell is not None:
                code = _match(str(cell), _full_scan_patterns)
                if code: return code

    return None


def _get_or_default_store(db: Session, store: str = None):
    """获取店铺对象，自动去除首尾空格，未指定返回 None"""
    if not store:
        return None
    store = store.strip()
    return db.query(DimStore).filter(DimStore.code == store).first()


def _get_or_create_time(db: Session, year: int, month: int) -> DimTime:
    year_month = f"{year}-{month:02d}"
    time_obj = db.query(DimTime).filter(DimTime.year_month == year_month).first()
    if not time_obj:
        time_obj = DimTime(time_year=year, time_month=month, year_month=year_month)
        db.add(time_obj)
        db.flush()
    return time_obj


def _parse_month_string(month_str: str):
    """将各种月份格式转为 (year, month)，支持格式:
    'May-26', 'Apr-26' → (2026, 5), (2026, 4)
    '2026-05', '2026/5' → (2026, 5)
    '2026-05-01', '2026/5/1' → (2026, 5)
    返回 (year_int, month_int) 或 (None, None)
    """
    import re
    if not month_str or not month_str.strip():
        return None, None
    s = month_str.strip()
    # 格式1a: 'May-26' → 英文缩写月-两位年
    m = re.search(r'^(\w{3})-(\d{2})$', s)
    # 格式1b: '26-May' → 两位年-英文缩写月
    m2 = re.search(r'^(\d{2})-(\w{3})$', s)
    month_map = {"jan":1,"feb":2,"mar":3,"apr":4,"may":5,"jun":6,
                 "jul":7,"aug":8,"sep":9,"oct":10,"nov":11,"dec":12}
    if m:
        abbr = m.group(1).lower()
        if abbr in month_map:
            year = int("20" + m.group(2))
            return year, month_map[abbr]
    if m2:
        abbr = m2.group(2).lower()
        if abbr in month_map:
            year = int("20" + m2.group(1))
            return year, month_map[abbr]
    # 格式2: '2026-05', '2026-05-01', '2026/5', '2026/5/1'
    m = re.search(r'(\d{4})[/\-](\d{1,2})', s)
    if m:
        year = int(m.group(1))
        month = int(m.group(2))
        if 1 <= month <= 12:
            return year, month
    return None, None


def _find_time_by_month_str(db: Session, month_str: str):
    """根据月份字符串找到对应的 DimTime，支持多种格式"""
    year, month = _parse_month_string(month_str)
    if year and month:
        standard = f"{year}-{month:02d}"
        time_obj = db.query(DimTime).filter(DimTime.year_month == standard).first()
        if time_obj:
            return time_obj
        return _get_or_create_time(db, year, month)
    return None


def _get_or_create_product(db: Session, asin: str, sku: str = None, store_id: int = None, year_month: str = None) -> DimProduct:
    """获取或创建产品（严格按店铺+月份隔离，不回退全局）"""
    asin = asin.strip() if asin else ""
    if not asin or asin.startswith("Amazon.") or asin.startswith("amzn.gr."):
        return None
    # 校验ASIN格式：真实Amazon ASIN为10位B0开头，防止sku.split('-')[0]误入
    import re as _re
    if not _re.match(r'^B0[A-Z0-9]{8}$', asin):
        # ASIN不合法，尝试通过SKU在已有产品中查找真实ASIN
        if sku and store_id:
            existing = db.query(DimProduct).filter(
                DimProduct.sku == sku, DimProduct.store_id == store_id,
                DimProduct.asin.op('REGEXP')(r'^B0[A-Z0-9]{8}$')
            ).first()
            if existing:
                return existing
        return None

    product = None
    if store_id:
        # 按 ASIN+店铺查找（同ASIN同店铺不跨月分拆）
        product = db.query(DimProduct).filter(
            DimProduct.asin == asin,
            DimProduct.store_id == store_id,
        ).first()

    if not product:
        product = DimProduct(asin=asin, sku=sku or asin, store_id=store_id, year_month=year_month)
        db.add(product)
        db.flush()
    elif sku:
        product.sku = sku
        db.flush()
    return product


def _extract_real_sku(sku: str):
    """从 amzn.gr.XXX-XXXX-XXXX-xxx-x 中提取真实SKU (XXX-XXXX-XXXX)"""
    if sku and sku.startswith("amzn.gr."):
        import re as _re
        # amzn.gr.WH-O0WE-F4CW-fkUFJmAenNF5IeG1-LN
        # → parts: ['amzn.gr.WH', 'O0WE', 'F4CW', ...]
        # → 真实SKU: WH-O0WE-F4CW (parts[0]最后一段 + parts[1] + parts[2])
        # amzn.gr.MGT-VH280B-48Beige-4ZwgOd8M5v-GD
        # → 真实SKU: MGT-VH280B-48Beige (前缀可以是2-5字符)
        parts = sku.split("-")
        if len(parts) >= 3:
            prefix = parts[0].split(".")[-1]  # 'amzn.gr.MGT' → 'MGT'
            candidate = f"{prefix}-{parts[1]}-{parts[2]}"
            # 放宽匹配：前缀2-5字符，中间段1-10字符，末段1-10字符
            if _re.match(r'^[A-Z0-9]{2,5}-[A-Z0-9]{1,10}-[A-Z0-9]{1,10}$', candidate.upper()):
                return candidate
    return None


def _get_exchange_rate(db: Session, country_obj, import_year=None, import_month=None, store_id=None) -> Decimal:
    """获取汇率：仅店铺专属 → 再默认值"""
    ym = None
    if import_year and import_month:
        ym = f"{import_year}-{import_month:02d}"
    if ym and store_id:
        # 店铺专属汇率（精确月份）
        er = db.query(DimExchangeRate).filter(
            DimExchangeRate.country_id == country_obj.id,
            DimExchangeRate.year_month == ym,
            DimExchangeRate.store_id == store_id,
        ).first()
        if er and er.rate and Decimal(str(er.rate)) != 0:
            return Decimal(str(er.rate))
    if store_id:
        # 店铺专属汇率（任意月份）
        er = db.query(DimExchangeRate).filter(
            DimExchangeRate.country_id == country_obj.id,
            DimExchangeRate.store_id == store_id,
        ).first()
        if er and er.rate and Decimal(str(er.rate)) != 0:
            return Decimal(str(er.rate))
    # 默认值
    defaults = {'US': '6.8', 'UK': '9.0', 'DE': '7.5', 'FR': '7.5', 'ES': '7.5', 'IT': '7.5',
                'NL': '7.5', 'SE': '0.65', 'BE': '7.5', 'IE': '7.5', 'CA': '5.0', 'MX': '0.4',
                'AE': '1.85', 'SA': '1.81'}
    return Decimal(defaults.get(country_obj.code, '6.8'))


def _find_product_by_sku(db: Session, sku: str, store_id: int = None, year_month: str = None) -> DimProduct:
    """通过 SKU 查找产品（按店铺匹配，不回退全局；year_month仅作兼容保留）"""
    if not sku:
        return None

    if store_id:
        product = db.query(DimProduct).filter(
            DimProduct.sku == sku,
            DimProduct.store_id == store_id,
            DimProduct.asin.like("B0%")
        ).first()
        if product:
            return product
        return db.query(DimProduct).filter(
            DimProduct.sku == sku,
            DimProduct.store_id == store_id,
        ).first()

    return None


def _find_product_by_asin(db: Session, asin: str, store_id: int = None, year_month: str = None) -> DimProduct:
    """通过 ASIN 查找产品（按店铺匹配；year_month仅作兼容保留）"""
    if not asin:
        return None
    if store_id:
        return db.query(DimProduct).filter(
            DimProduct.asin == asin,
            DimProduct.store_id == store_id,
        ).first()
    return None


def _get_or_create_monthly_summary(
    db: Session, country_id: int, product_id: int, time_id: int, store_id: int = None
) -> MonthlySummary:
    filters = [
        MonthlySummary.country_id == country_id,
        MonthlySummary.product_id == product_id,
        MonthlySummary.time_id == time_id,
    ]
    if store_id:
        filters.append(MonthlySummary.store_id == store_id)
    else:
        filters.append(MonthlySummary.store_id.is_(None))
    summary = (
        db.query(MonthlySummary)
        .filter(*filters)
        .first()
    )
    if not summary:
        summary = MonthlySummary(
            country_id=country_id,
            product_id=product_id,
            time_id=time_id,
            store_id=store_id,
        )
        db.add(summary)
        db.flush()
    elif store_id and not summary.store_id:
        summary.store_id = store_id
    return summary


def _safe_decimal(value, default=0) -> Decimal:
    """解析数字，支持欧洲格式(106,39)和英语格式(106.39)"""
    if value is None:
        return Decimal(str(default))
    s = str(value).strip()
    if not s or s == '-' or s == '−':
        return Decimal(str(default))
    # Normalize Unicode minus
    s = s.replace('−', '-')
    # Remove currency symbols
    s = s.replace('€', '').replace('$', '').replace('¥', '').replace('£', '').replace('%', '').strip()
    import re
    if re.search(r',\d{1,2}$', s):
        # European format: "106,39" or "1 135,72" or "-10,64"
        s = s.replace(' ', '').replace('.', '').replace(',', '.')
    elif ',' in s and re.search(r',\d{3}', s):
        # US thousands: "1,777.43"
        s = s.replace(',', '')
    else:
        s = s.replace(',', '')
    if s == "" or s == "-":
        return Decimal(str(default))
    try:
        return Decimal(s)
    except InvalidOperation:
        return Decimal(str(default))


def _safe_str(value, max_len=500) -> str:
    """安全字符串提取，自动截断超长值，防止 Data too long for column 错误"""
    if value is None:
        return ""
    s = str(value).strip()
    if len(s) > max_len:
        s = s[:max_len]
    return s


def _safe_int(value, default=0) -> int:
    if value is None:
        return default
    s = str(value).strip().replace(",", "")
    if s == "" or s == "-":
        return default
    try:
        return int(float(s))
    except (ValueError, TypeError):
        return default


def _json_safe(header, row):
    """将 header+row 转为 JSON 安全的 dict（datetime→str, Decimal→str）"""
    result = {}
    for h, v in zip(header, row):
        if isinstance(v, (datetime, date)):
            result[h] = v.isoformat()
        elif isinstance(v, Decimal):
            result[h] = str(v)
        else:
            result[h] = v
    return result


def _detect_date_format(date_str: str):
    """尝试解析多种日期格式（支持英/德/法/西/意/荷/瑞典语月份）"""
    date_str = date_str.strip()
    # 移除时区后缀 (PDT, PST, EST, UTC, GMT-7, GMT+5:30 等)
    date_str_clean = re.sub(r'\s+[A-Z]{2,4}[\d\-+:]*$', '', date_str)
    # 统一 a.m./p.m. → AM/PM（西班牙语格式用 .replace 因为 \b 和 . 冲突）
    date_str_clean = date_str_clean.replace('a.m.', 'AM').replace('p.m.', 'PM')
    date_str_clean = date_str_clean.replace('A.M.', 'AM').replace('P.M.', 'PM')
    # 非英语月份名 → 英语月份名
    _month_map = {
        'janvier': 'January', 'janv.': 'January', 'janv': 'January',
        'février': 'February', 'févr.': 'February', 'févr': 'February',
        'mars': 'March',
        'avril': 'April', 'avr.': 'April', 'avr': 'April',
        'mai': 'May',
        'juin': 'June',
        'juillet': 'July', 'juil.': 'July', 'juil': 'July',
        'août': 'August',
        'septembre': 'September', 'sept.': 'September', 'sept': 'September',
        'octobre': 'October',
        'novembre': 'November',
        'décembre': 'December', 'déc.': 'December', 'déc': 'December',
        'gennaio': 'January', 'febbraio': 'February', 'marzo': 'March',
        'aprile': 'April', 'maggio': 'May', 'giugno': 'June',
        'luglio': 'July', 'agosto': 'August', 'settembre': 'September',
        'ottobre': 'October', 'novembre': 'November', 'dicembre': 'December',
        'januari': 'January', 'februari': 'February', 'maart': 'March',
        'mei': 'May', 'juni': 'June', 'juli': 'July',
        'augustus': 'August', 'oktober': 'October', 'december': 'December',
        'januar': 'January', 'februar': 'February', 'märz': 'March',
        'september': 'September', 'dezember': 'December',
        'enero': 'January', 'febrero': 'February', 'abril': 'April',
        'mayo': 'May', 'junio': 'June', 'julio': 'July',
        'septiembre': 'September', 'octubre': 'October', 'diciembre': 'December',
        # Spanish 3-letter abbreviated (Amazon MX format)
        'ene': 'Jan', 'mar': 'Mar', 'abr': 'Apr',
        'jun': 'Jun', 'jul': 'Jul',
        'sep': 'Sep', 'oct': 'Oct', 'nov': 'Nov',
        'maj': 'May', 'augusti': 'August',
        'gen': 'Jan', 'feb': 'Feb', 'mag': 'May', 'giu': 'Jun',
        'lug': 'Jul', 'ago': 'Aug', 'set': 'Sep', 'ott': 'Oct', 'dic': 'Dec',
        # Swedish months (full + abbreviated)
        'januari': 'January', 'februari': 'February', 'mars': 'March',
        'april': 'April', 'maj': 'May', 'juni': 'June', 'juli': 'July',
        'augusti': 'August', 'oktober': 'October',
        'apr.': 'April', 'apr': 'April',
    }
    lower = date_str_clean.lower()
    for local, eng in _month_map.items():
        if local in lower:
            # 用正则替换完整单词（避免 apr→April 后又把 April 里的 Apr 再替换）
            if local.endswith('.'):
                # 带点号的（如 apr.）用简单替换
                date_str_clean = re.sub(re.escape(local), eng, date_str_clean, flags=re.IGNORECASE)
            else:
                date_str_clean = re.sub(r'\b' + re.escape(local) + r'\b', eng, date_str_clean, flags=re.IGNORECASE)
            break
    formats = [
        "%d %b %Y %I:%M:%S %p",    # 5 may 2026 5:08:09 AM (MX无逗号格式)
        "%d %B %Y %I:%M:%S %p",    # 2 mai 2026 03:04:42 AM (French full month)
        "%d %b %Y %H:%M:%S",       # 9 maj 2026 03:41:35 (Swedish)
        "%d %B %Y %H:%M:%S",       # 9 maj 2026 03:41:35 (Swedish full)
        "%b %d, %Y %I:%M:%S %p",   # May 1, 2026 5:46:45 AM (abbreviated)
        "%B %d, %Y %I:%M:%S %p",   # April 1, 2026 6:50:23 AM (full month, after apr→April)
        "%b %d, %Y %I:%M %p",
        "%B %d, %Y %I:%M %p",
        "%b %d, %Y",
        "%B %d, %Y",
        "%d.%m.%Y %H:%M:%S",       # 01.05.2026 10:05:35 (German)
        "%d.%m.%Y",                 # 01.05.2026 (German date only)
        "%m/%d/%Y %H:%M:%S",
        "%m/%d/%Y %H:%M",
        "%m/%d/%Y",
        "%Y-%m-%dT%H:%M:%S%z",     # ISO 8601
        "%Y-%m-%d %H:%M:%S",
        "%Y-%m-%d %H:%M",
        "%Y-%m-%d",
        "%Y/%m/%d",
    ]
    for fmt in formats:
        try:
            return datetime.strptime(date_str_clean, fmt)
        except ValueError:
            continue
    return None


# ============================================================
# POST /transaction: 导入交易记录 CSV
# ============================================================
@router.post("/transaction")
async def import_transaction(
    file: UploadFile = File(...),
    country: str = Form(..., description="国家代码，如 US"),
    store: str = Form(None, description="店铺代码"),
    import_year: int = Form(None, description="导入年份"),
    import_month: int = Form(None, description="导入月份"),
    db: Session = Depends(get_db),
):
    try:
        country = country.upper()
        country_obj = db.query(DimCountry).filter(DimCountry.code == country).first()
        if not country_obj:
            return {"detail": f"国家 {country} 不存在，请先在 dim_country 中创建"}
        store_obj = _get_or_default_store(db, store)
        if not store_obj:
            return {"detail": f"店铺 {store} 不存在"}

        if not import_year or not import_month:
            raise HTTPException(status_code=400, detail="必须同时提供 year 和 month 参数")

        from sqlalchemy import extract as _extract

        _clear_time = db.query(DimTime).filter(DimTime.time_year == import_year, DimTime.time_month == import_month).first()
        _clear_filters = [RawTransaction.store_id == store_obj.id]
        if _clear_time:
            _clear_filters.append(RawTransaction.time_id == _clear_time.id)
        db.query(RawTransaction).filter(*_clear_filters).delete()

        _time_obj = db.query(DimTime).filter(DimTime.time_year == import_year, DimTime.time_month == import_month).first()
        _clear_ms = [MonthlySummary.store_id == store_obj.id]
        if _time_obj:
            _clear_ms.append(MonthlySummary.time_id == _time_obj.id)
        db.query(MonthlySummary).filter(*_clear_ms).delete()
        db.flush()

        content = await file.read()
        # 尝试多种编码
        for encoding in ["utf-8-sig", "utf-8", "gbk", "latin-1"]:
            try:
                text = content.decode(encoding)
                break
            except UnicodeDecodeError:
                continue
        else:
            return {"detail": "无法解码文件，请确认编码格式"}

        lines = text.splitlines()

        # 前9行是元数据，跳过，第10行是表头
        if len(lines) < 10:
            return {"detail": "CSV 行数不足，至少需要10行（9行元数据+1行表头）"}

        header_line = lines[9]
        data_rows = lines[10:]

        reader = csv.DictReader(io.StringIO(header_line + "\n" + "\n".join(data_rows)))
        headers = reader.fieldnames
        if not headers:
            return {"detail": "无法解析 CSV 表头"}

        # 多语言字段映射
        MULTI_LANG_MAP = {
            # Spanish → English
            'fecha/hora': 'date/time',
            'id. de liquidación': 'settlement id',
            'tipo': 'type',
            'pedido': 'Order', 'reembolso': 'Refund',
            'id. del pedido': 'order id',
            'sku': 'sku',
            'descripción': 'description',
            'cantidad': 'quantity',
            'marketplace': 'marketplace',
            'cumplimiento': 'fulfillment',
            'ventas de productos': 'product sales',
            'impuesto de ventas de productos': 'product sales tax',
            'abonos de envío': 'postage credits',
            'créditos de envío': 'shipping credits',
            'impuesto de abono de envío': 'shipping credits tax',
            'créditos por envoltorio de regalo': 'gift wrap credits',
            'descuentos promocionales': 'promotional rebates',
            'impuesto de reembolsos promocionales': 'promotional rebates tax',
            'impuesto de retenciones en la plataforma': 'marketplace withheld tax',
            'tarifas de venta': 'selling fees',
            'tarifas fba': 'fba fees',
            'tarifas de otra transacción': 'other transaction fees',
            'otro': 'other',
            'total': 'total',
            'estado de la transacción': 'transaction_status',
            'fecha de liberación de la transacción': 'transaction_release_date',
            # French → English
            'numéro de versement': 'settlement id',
            'commande': 'Order', 'remboursement': 'Refund',
            'numéro de la commande': 'order id',
            'quantité': 'quantity',
            'traitement': 'fulfillment',
            'ventes de produits': 'product sales',
            'taxes sur la vente des produits': 'product sales tax',
            "crédits d'expédition": 'postage credits',
            "taxe sur les crédits d'expédition": 'shipping credits tax',
            "crédits sur l'emballage cadeau": 'gift wrap credits',
            'taxes sur les crédits cadeaux': 'giftwrap credits tax',
            'taxes sur les crédits cadeaux': 'gift wrap credits tax',
            'rabais promotionnels': 'promotional rebates',
            'taxes sur les remises promotionnelles': 'promotional rebates tax',
            'taxes retenues sur le site de vente': 'marketplace withheld tax',
            'frais de vente': 'selling fees',
            'frais expédié par amazon': 'fba fees',
            'autres frais de transaction': 'other transaction fees',
            'autre': 'other',
            'statut de la transaction': 'transaction_status',
            'date de sortie de la transaction': 'transaction_release_date',
            'frais de service': 'Service fee',
            'transférer': 'Transfer',
            'frais de transaction expédié par amazon': 'FBA transaction fees',
            # German → English
            'datum/uhrzeit': 'date/time',
            'abrechnungsnummer': 'settlement id',
            'typ': 'type',
            'bestellung': 'Order', 'erstattung': 'Refund',
            'bestellnummer': 'order id',
            'menge': 'quantity',
            'versand': 'fulfillment',
            'umsätze': 'product sales',
            'produktumsatzsteuer': 'product sales tax',
            'gutschrift für versandkosten': 'postage credits',
            'steuer auf versandgutschrift': 'shipping credits tax',
            'gutschrift für geschenkverpackung': 'gift wrap credits',
            'steuer auf geschenkverpackungsgutschriften': 'gift wrap credits tax',
            'rabatte aus werbeaktionen': 'promotional rebates',
            'steuer auf aktionsrabatte': 'promotional rebates tax',
            'einbehaltene steuer auf marketplace': 'marketplace withheld tax',
            'verkaufsgebühren': 'selling fees',
            'gebühren zu versand durch amazon': 'fba fees',
            'andere transaktionsgebühren': 'other transaction fees',
            'andere': 'other',
            'gesamt': 'total',
            'transaktionsstatus': 'transaction_status',
            'freigabedatum der transaktion': 'transaction_release_date',
            'servicegebühr': 'Service fee',
            'übertrag': 'Transfer',
            'transaktionsgebühren für versand durch amazon': 'FBA transaction fees',
            'versand durch amazon lagergebühr': 'FBA Inventory Fee',
            'bestellung_wiedereinzug': 'Order',
            # Dutch → English
            'datum/tijd': 'date/time',
            'schikkings-id': 'settlement id',
            'bestelling': 'Order', 'terugbetaling': 'Refund',
            'bestelnummer': 'order id',
            'aantal': 'quantity',
            'verkoop van producten': 'product sales',
            'verzendtegoeden': 'postage credits',
            'kredietpunten cadeauverpakking': 'gift wrap credits',
            'promotiekortingen': 'promotional rebates',
            'geïnde omzetbelasting': 'marketplace withheld tax',
            'verkoopkosten': 'selling fees',
            'fba-vergoedingen': 'fba fees',
            'overige transactiekosten': 'other transaction fees',
            'overige': 'other',
            'totaal': 'total',
            'transactiestatus': 'transaction_status',
            'publicatiedatum van transactie': 'transaction_release_date',
            'servicekosten': 'Service fee',
            'overboeking': 'Transfer',
            # Swedish → English
            'datum/tid': 'date/time',
            'reglerings-id': 'settlement id',
            'beställning': 'Order', 'återbäring': 'Refund',
            'beställnings-id': 'order id',
            'antal': 'quantity',
            'leverans': 'fulfillment',
            'försäljning av produkter': 'product sales',
            'fraktkrediter': 'postage credits',
            'krediter för presentinslagning': 'gift wrap credits',
            'kampanjrabatter': 'promotional rebates',
            'inkasserad moms': 'marketplace withheld tax',
            'försäljningsavgifter': 'selling fees',
            'fba-avgifter': 'fba fees',
            'övriga transaktionsavgifter': 'other transaction fees',
            'övrigt': 'other',
            'totalt': 'total',
            'transaktionsstatus': 'transaction_status',
            'transaktionens utgivningsdatum': 'transaction_release_date',
            'överföring': 'Transfer',
            # Italian → English
            'data/ora:': 'date/time', 'data/ora': 'date/time',
            'numero pagamento': 'settlement id',
            'ordine': 'Order', 'rimborso': 'Refund',
            'numero ordine': 'order id',
            'quantità': 'quantity',
            'gestione': 'fulfillment',
            'vendite': 'product sales',
            'imposta sulle vendite dei prodotti': 'product sales tax',
            'accrediti per le spedizioni': 'postage credits',
            'imposta accrediti per le spedizioni': 'shipping credits tax',
            'accrediti per confezioni regalo': 'gift wrap credits',
            'imposta sui crediti confezione regalo': 'gift wrap credits tax',
            'sconti promozionali': 'promotional rebates',
            'imposta sugli sconti promozionali': 'promotional rebates tax',
            'trattenuta iva del marketplace': 'marketplace withheld tax',
            'commissioni di vendita': 'selling fees',
            'costi del servizio logistica di amazon': 'fba fees',
            'altri costi relativi alle transazioni': 'other transaction fees',
            'altro': 'other',
            'stato della transazione': 'transaction_status',
            'data di rilascio della transazione': 'transaction_release_date',
            'trasferimento': 'Transfer',
            'modifica': 'Adjustment',
            'commissioni per le transazioni di logistica di amazon': 'FBA transaction fees',
        }
        # 字段映射（去空格、小写化）
        def map_header(h):
            h = h.strip().lower()
            # 先查多语言映射
            if h in MULTI_LANG_MAP:
                h = MULTI_LANG_MAP[h]
            mapping = {
                "date/time": "transaction_date",
                "date / time": "transaction_date",
                "settlement id": "settlement_id",
                "type": "transaction_type",
                "order id": "order_id",
                "sku": "sku",
                "description": "description",
                "quantity": "quantity",
                "marketplace": "marketplace",
                "fulfillment": "fulfillment",
                "order city": "order_city",
                "order state": "order_state",
                "order postal": "order_postal",
                "tax collection model": "tax_collection_model",
                "product sales": "product_sales",
                "product sales tax": "product_sales_tax",
                "postage credits": "postage_credits",
                "postage_credits": "postage_credits",
                "shipping credits": "postage_credits",  # UK叫postage credits, 其他市场叫shipping credits, 统一映射
                "shipping credits tax": "shipping_credits_tax",
                "gift wrap credits": "gift_wrap_credits",
                "giftwrap credits tax": "giftwrap_credits_tax",
                "regulatory fee": "regulatory_fee",
                "tax on regulatory fee": "tax_on_regulatory_fee",
                "promotional rebates": "promotional_rebates",
                "promotional rebates tax": "promotional_rebates_tax",
                "marketplace withheld tax": "marketplace_withheld_tax",
                "selling fees": "selling_fee",
                "fba fees": "fba_fee",
                "fulfilment by amazon fees": "fba_fee",
                "fulfillment by amazon fees": "fba_fee",
                "other transaction fees": "other_transaction_fee",
                "other": "other_amount",
                "total": "total",
                "status": "transaction_status",
                "release date": "transaction_release_date",
            }
            return mapping.get(h, h)

        # 按 SKU 聚合用于 monthly_summary（仅 Order/Refund）
        sku_aggregation = {}  # key: (sku, year, month) -> dict
        adj_aggregation = {}  # key: (sku, year, month) -> {"total": Decimal, "qty": int}
        row_count = 0
        type_counts = {}

        for row in reader:
            if not row:
                continue

            mapped = {}
            for header_name, value in row.items():
                field = map_header(header_name)
                mapped[field] = value.strip() if value else ""

            # 解析日期
            date_str = mapped.get("transaction_date", "")
            txn_date = _detect_date_format(date_str)
            if not txn_date:
                continue  # 跳过无法解析日期的行

            txn_type = mapped.get("transaction_type", "").strip()
            # 多语言类型统一映射为英文
            _type_map = {
                'pedido': 'Order', 'reembolso': 'Refund', 'ajuste': 'Adjustment',
                'order': 'Order', 'refund': 'Refund', 'adjustment': 'Adjustment',
                'bestellung': 'Order', 'erstattung': 'Refund', 'anpassung': 'Adjustment',
                'commande': 'Order', 'remboursement': 'Refund', 'ajustement': 'Adjustment',
                'ordine': 'Order', 'rimborso': 'Refund', 'aggiustamento': 'Adjustment',
                'bestelling': 'Order', 'terugbetaling': 'Refund', 'aanpassing': 'Adjustment',
                'beställning': 'Order', 'återbetalning': 'Refund', 'justering': 'Adjustment',
            }
            txn_type = _type_map.get(txn_type.lower(), txn_type)
            type_counts[txn_type] = type_counts.get(txn_type, 0) + 1

            sku = mapped.get("sku", "").strip()
            asin = sku.split("-")[0] if sku and "-" in sku else sku

            is_replacement = sku.startswith("amzn.gr.") if sku else False
            real_sku = _extract_real_sku(sku) if is_replacement else None
            effective_sku = real_sku if real_sku else sku

            product_sales = _safe_decimal(mapped.get("product_sales"))
            selling_fee = _safe_decimal(mapped.get("selling_fee"))
            fba_fee = _safe_decimal(mapped.get("fba_fee"))
            quantity = _safe_int(mapped.get("quantity"))
            total = _safe_decimal(mapped.get("total"))

            # 费用回填：如果佣金和FBA费都为0但total≠product_sales，从差额推算缺失费用
            if txn_type == "Order" and product_sales > 0 and total > 0:
                other_charges = _safe_decimal(mapped.get("other_transaction_fee")) + \
                               _safe_decimal(mapped.get("promotional_rebates"))
                implied_fees = product_sales - total + other_charges
                captured_fees = abs(selling_fee)
                if fba_fee == 0 and selling_fee == 0 and implied_fees > Decimal("0.5"):
                    est_commission = (product_sales * Decimal("0.15")).quantize(Decimal("0.01"))
                    est_fba = implied_fees - est_commission
                    if est_fba < 0:
                        est_commission = implied_fees
                        est_fba = Decimal("0")
                    selling_fee = -est_commission
                    fba_fee = -est_fba
                elif fba_fee == 0 and selling_fee != 0 and implied_fees > captured_fees + Decimal("0.5"):
                    fba_fee = -(implied_fees - captured_fees)

            # 所有类型都写入 raw_transactions（全量存储）
            # time_id 用导入时选择的月份，忽略数据中的日期
            raw_time_id = None
            if import_year and import_month:
                raw_time_obj = _get_or_create_time(db, import_year, import_month)
                raw_time_id = raw_time_obj.id if raw_time_obj else None

            raw = RawTransaction(
                country_id=country_obj.id,
                store_id=store_obj.id,
                time_id=raw_time_id,
                transaction_date=txn_date,
                settlement_id=_safe_str(mapped.get("settlement_id", ""), 50),
                transaction_type=_safe_str(txn_type, 50),
                order_id=_safe_str(mapped.get("order_id", ""), 50),
                sku=_safe_str(sku, 100),
                description=_safe_str(mapped.get("description", ""), 500),
                quantity=quantity,
                marketplace=_safe_str(mapped.get("marketplace", ""), 20),
                fulfillment=_safe_str(mapped.get("fulfillment", ""), 20),
                order_city=_safe_str(mapped.get("order_city", ""), 100),
                order_state=_safe_str(mapped.get("order_state", ""), 100),
                order_postal=_safe_str(mapped.get("order_postal", ""), 20),
                tax_collection_model=_safe_str(mapped.get("tax_collection_model", ""), 50),
                product_sales=product_sales,
                product_sales_tax=_safe_decimal(mapped.get("product_sales_tax")),
                postage_credits=_safe_decimal(mapped.get("postage_credits")),
                shipping_credits=_safe_decimal(mapped.get("shipping_credits")),
                shipping_credits_tax=_safe_decimal(mapped.get("shipping_credits_tax")),
                gift_wrap_credits=_safe_decimal(mapped.get("gift_wrap_credits")),
                giftwrap_credits_tax=_safe_decimal(mapped.get("giftwrap_credits_tax")),
                regulatory_fee=_safe_decimal(mapped.get("regulatory_fee")),
                tax_on_regulatory_fee=_safe_decimal(mapped.get("tax_on_regulatory_fee")),
                promotional_rebates=_safe_decimal(mapped.get("promotional_rebates")),
                promotional_rebates_tax=_safe_decimal(mapped.get("promotional_rebates_tax")),
                marketplace_withheld_tax=_safe_decimal(mapped.get("marketplace_withheld_tax")),
                selling_fee=selling_fee,
                fba_fee=fba_fee,
                other_transaction_fee=_safe_decimal(mapped.get("other_transaction_fee")),
                other_amount=_safe_decimal(mapped.get("other_amount")),
                total=total,
                transaction_status=_safe_str(mapped.get("transaction_status", ""), 50),
                transaction_release_date=_detect_date_format(mapped.get("transaction_release_date", "")) if mapped.get("transaction_release_date") else None,
            )
            db.add(raw)
            row_count += 1

            # Adjustment 处理：按有/无order_id分别处理
            if txn_type == "Adjustment":
                adj_key = (effective_sku, txn_date.year, txn_date.month)
                if adj_key not in adj_aggregation:
                    adj_aggregation[adj_key] = {
                        "other_with_order": Decimal("0"),  # 有order_id的other_amount
                        "total_no_order_pos": Decimal("0"),  # 无order_id且total>0
                        "total_no_order_neg": Decimal("0"),  # 无order_id且total<0
                        "qty": 0,
                    }
                order_id_val = mapped.get("order_id", "")
                other_amount = _safe_decimal(mapped.get("other_amount"))
                if order_id_val and str(order_id_val).strip():
                    # 有order_id：other字段加到利润
                    adj_aggregation[adj_key]["other_with_order"] += other_amount
                else:
                    # 无order_id
                    if total > 0:
                        adj_aggregation[adj_key]["total_no_order_pos"] += total
                        adj_aggregation[adj_key]["qty"] += abs(quantity)
                    elif total < 0:
                        adj_aggregation[adj_key]["total_no_order_neg"] += total
                        adj_aggregation[adj_key]["qty"] -= abs(quantity)
                continue

            # 仅 Order/Refund 参与 monthly_summary 聚合
            if txn_type not in ("Order", "Refund"):
                continue

            year = txn_date.year
            month = txn_date.month
            key = (effective_sku, year, month)
            if key not in sku_aggregation:
                sku_aggregation[key] = {
                    "product_sales": Decimal("0"),
                    "product_sales_tax": Decimal("0"),
                    "postage_credits": Decimal("0"),
                    "shipping_credits": Decimal("0"),
                    "shipping_credits_tax": Decimal("0"),
                    "gift_wrap_credits": Decimal("0"),
                    "giftwrap_credits_tax": Decimal("0"),
                    "selling_fee": Decimal("0"),
                    "fba_fee": Decimal("0"),
                    "order_count": 0,
                    "quantity": 0,
                    "order_qty": 0,
                    "promo_rebate": Decimal("0"),
                    "promo_rebate_tax": Decimal("0"),
                    "marketplace_withheld_tax": Decimal("0"),
                }
            agg = sku_aggregation[key]
            # 聚合所有亚马逊字段
            agg["product_sales"] += product_sales
            agg["product_sales_tax"] += _safe_decimal(mapped.get("product_sales_tax"))
            agg["postage_credits"] += _safe_decimal(mapped.get("postage_credits"))
            agg["shipping_credits"] += _safe_decimal(mapped.get("shipping_credits"))
            agg["shipping_credits_tax"] += _safe_decimal(mapped.get("shipping_credits_tax"))
            agg["gift_wrap_credits"] += _safe_decimal(mapped.get("gift_wrap_credits"))
            agg["giftwrap_credits_tax"] += _safe_decimal(mapped.get("giftwrap_credits_tax"))
            agg["promo_rebate"] += _safe_decimal(mapped.get("promotional_rebates"))
            agg["promo_rebate_tax"] += _safe_decimal(mapped.get("promotional_rebates_tax"))
            agg["marketplace_withheld_tax"] += _safe_decimal(mapped.get("marketplace_withheld_tax"))
            agg["selling_fee"] += selling_fee
            agg["fba_fee"] += fba_fee
            agg["order_count"] += 1
            # Refund 数量为负（净销量）
            if txn_type == "Refund":
                agg["quantity"] -= abs(quantity)
            else:
                agg["quantity"] += abs(quantity)
                if not is_replacement:
                    agg["order_qty"] += abs(quantity)

        # 按 SKU 聚合写入 monthly_summary
        summary_count = 0

        for (sku, year, month), agg in sku_aggregation.items():
            # 先通过 SKU 查找已有产品（按店铺+月份优先匹配）
            product = _find_product_by_sku(db, sku, store_id=store_obj.id if store_obj else None, year_month=f"{year}-{month:02d}")
            if not product:
                asin = sku.split("-")[0] if sku and "-" in sku else sku
                product = _get_or_create_product(db, asin, sku, store_id=store_obj.id)
            if not product:
                continue
            time_obj = _get_or_create_time(db, year, month)
            exchange_rate = _get_exchange_rate(db, country_obj, year, month, store_id=store_obj.id)

            summary = _get_or_create_monthly_summary(db, country_obj.id, product.id, time_obj.id, store_id=store_obj.id)

            summary.product_sales_usd = agg["product_sales"]
            summary.product_sales_tax = agg["product_sales_tax"]
            summary.postage_credits = agg["postage_credits"]
            summary.shipping_credits = agg["shipping_credits"]
            summary.shipping_credits_tax = agg["shipping_credits_tax"]
            summary.gift_wrap_credits = agg["gift_wrap_credits"]
            summary.giftwrap_credits_tax = agg["giftwrap_credits_tax"]
            summary.commission_usd = agg["selling_fee"]
            summary.fba_fee_usd = agg["fba_fee"]
            summary.promo_rebate_usd = agg["promo_rebate"]
            summary.promo_rebate_tax_usd = agg["promo_rebate_tax"]
            summary.marketplace_withheld_tax_usd = agg["marketplace_withheld_tax"]
            summary.exchange_rate = exchange_rate
            summary.product_sales_rmb = (agg["product_sales"] * exchange_rate).quantize(Decimal("0.01"))
            summary.order_count = agg["quantity"]
            summary.order_qty = agg["order_qty"]

            # 成本和运费
            time_obj = db.query(DimTime).filter(DimTime.id == summary.time_id).first()
            ym_str = time_obj.year_month if time_obj else None
            pc = db.query(DimProductCost).filter(DimProductCost.product_id == product.id, DimProductCost.year_month == ym_str).first()
            if not pc:
                pc = db.query(DimProductCost).filter(DimProductCost.product_id == product.id).first()
            unit_cost = Decimal(str(pc.cost_rmb if pc else 0))
            unit_freight = Decimal(str(pc.freight_per_unit if pc else 0))
            if store_obj:
                df = db.query(DimFreight).filter(DimFreight.product_id == product.id, DimFreight.country_id == country_obj.id, DimFreight.store_id == store_obj.id).first()
            else:
                df = db.query(DimFreight).filter(DimFreight.product_id == product.id, DimFreight.country_id == country_obj.id).first()
            if df:
                unit_freight = Decimal(str(df.freight_rmb))
            summary.product_cost_rmb = (unit_cost * agg["order_qty"]).quantize(Decimal("0.01"))
            summary.freight_cost_rmb = (unit_freight * agg["order_qty"]).quantize(Decimal("0.01"))

            # 统一利润公式
            from services.profit import apply_profit_to_summary
            apply_profit_to_summary(summary, exchange_rate,
                raw_product_sales=agg["product_sales"],
                raw_product_sales_tax=agg["product_sales_tax"],
                raw_postage_credits=agg["postage_credits"],
                raw_shipping_credits_tax=agg["shipping_credits_tax"],
                raw_gift_wrap_credits=agg["gift_wrap_credits"],
                raw_giftwrap_credits_tax=agg["giftwrap_credits_tax"],
                raw_promo_rebate=agg["promo_rebate"],
                raw_promo_rebate_tax=agg["promo_rebate_tax"],
                raw_marketplace_withheld_tax=agg["marketplace_withheld_tax"],
                raw_selling_fee=agg["selling_fee"],
                raw_fba_fee=agg["fba_fee"],
            )

            summary_count += 1

        # 处理 Adjustment（按新算法：有order_id用other_amount，无order_id用total）
        for (sku, adj_year, adj_month), adj_agg in adj_aggregation.items():
            product = _find_product_by_sku(db, sku, store_id=store_obj.id if store_obj else None, year_month=f"{adj_year}-{adj_month:02d}")
            if not product:
                asin = sku.split("-")[0] if sku and "-" in sku else sku
                product = _get_or_create_product(db, asin, sku, store_id=store_obj.id)
            if not product:
                continue
            time_obj = _get_or_create_time(db, adj_year, adj_month)
            exchange_rate = _get_exchange_rate(db, country_obj, adj_year, adj_month, store_id=store_obj.id)
            summary = _get_or_create_monthly_summary(db, country_obj.id, product.id, time_obj.id, store_id=store_obj.id)

            # adjustment_usd = 有order_id的other + 无order_id的total
            adj_total = adj_agg["other_with_order"] + adj_agg["total_no_order_pos"] + adj_agg["total_no_order_neg"]
            summary.adjustment_usd = (summary.adjustment_usd or Decimal("0")) + adj_total

            # 无order_id的Adjustment：调整order_qty和成本
            if adj_agg["qty"] != 0:
                summary.order_qty = (summary.order_qty or 0) + adj_agg["qty"]
                pc = db.query(DimProductCost).filter(DimProductCost.product_id == product.id, DimProductCost.year_month == f"{adj_year}-{adj_month:02d}").first()
                if not pc:
                    pc = db.query(DimProductCost).filter(DimProductCost.product_id == product.id).first()
                if pc:
                    unit_cost = Decimal(str(pc.cost_rmb or 0))
                    unit_freight = Decimal(str(pc.freight_per_unit or 0))
                    if store_obj:
                        df = db.query(DimFreight).filter(DimFreight.product_id == product.id, DimFreight.country_id == country_obj.id, DimFreight.store_id == store_obj.id).first()
                    else:
                        df = db.query(DimFreight).filter(DimFreight.product_id == product.id, DimFreight.country_id == country_obj.id).first()
                    if df:
                        unit_freight = Decimal(str(df.freight_rmb))
                    summary.product_cost_rmb = (unit_cost * summary.order_qty).quantize(Decimal("0.01"))
                    summary.freight_cost_rmb = (unit_freight * summary.order_qty).quantize(Decimal("0.01"))

            # 统一利润公式
            adj_total = adj_agg["other_with_order"] + adj_agg["total_no_order_pos"] + adj_agg["total_no_order_neg"]
            from services.profit import apply_profit_to_summary
            apply_profit_to_summary(summary, exchange_rate, raw_adjustment=adj_total)

        # 补建有成本但无交易的产品summary（按店铺隔离）
        _ensure_all_products_have_summary(db, country_obj, store_id=store_obj.id if store_obj else None, import_year=import_year, import_month=import_month)

        db.commit()

        return {
            "message": "交易记录导入成功",
            "raw_rows": row_count,
            "summary_rows": summary_count,
            "country": country,
            "type_counts": type_counts,
        }

    except Exception as e:
        db.rollback()
        return {"detail": str(e)}


# ============================================================
# POST /product-info: 导入产品信息 XLSX
# ============================================================
@router.post("/product-info")
async def import_product_info(
    file: UploadFile = File(...),
    store: str = Form(None, description="店铺代码"),
    import_year: int = Form(None, description="导入年份"),
    import_month: int = Form(None, description="导入月份"),
    db: Session = Depends(get_db),
):
    try:
        import openpyxl

        store_obj = _get_or_default_store(db, store)
        ym = f"{import_year}-{import_month:02d}" if import_year and import_month else None

        content = await file.read()
        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True)
        ws = wb.active

        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 2:
            return {"detail": "文件数据不足，至少需要表头+1行数据"}

        header = [str(h).strip() if h else "" for h in rows[0]]

        # 找列索引（精确匹配优先，大小写不敏感，避免 "产品" 误匹配 "产品运费/台"）
        col_map = {}
        field_names = ["ASIN", "SKU", "产品", "型号", "颜色", "成本RMB", "产品运费/台", "汇率", "时间",
                       "英国运费", "爱尔兰运费", "其他站点运费", "英国站运费", "英国站点运费", "运费", "爱尔兰站运费", "爱尔兰站点运费"]
        for i, h in enumerate(header):
            hl = h.lower() if h else ""
            for fn in field_names:
                if hl == fn.lower():   # 精确匹配（大小写不敏感）
                    col_map[fn] = i
                    break
            else:
                for fn in field_names:
                    if fn.lower() in hl:  # 模糊匹配（大小写不敏感）
                        col_map[fn] = i
                        break

        if "ASIN" not in col_map:
            return {"detail": "未找到 ASIN 列"}

        # 检测是否有独立国家运费列（英国/爱尔兰）
        _country_freight_map = {
            "英国站运费": "UK", "英国站点运费": "UK", "英国运费": "UK",
            "爱尔兰站运费": "IE", "爱尔兰站点运费": "IE", "爱尔兰运费": "IE",
        }
        has_country_freight = any(k in col_map for k in _country_freight_map)
        # 预查国家 ID
        country_freight_ids = {}  # "UK" -> country_id, "IE" -> country_id
        if has_country_freight:
            for code in set(_country_freight_map.values()):
                c = db.query(DimCountry).filter(DimCountry.code == code).first()
                if c:
                    country_freight_ids[code] = c.id

        count = 0
        for row in rows[1:]:
            if not row or not row[col_map["ASIN"]]:
                continue

            asin = str(row[col_map["ASIN"]]).strip()
            sku = str(row[col_map.get("SKU", 1)]).strip() if "SKU" in col_map and row[col_map["SKU"]] else None
            # 产品名称：优先"产品"列，fallback到"型号"列
            pn_key = "产品" if "产品" in col_map else ("型号" if "型号" in col_map else None)
            product_name = str(row[col_map[pn_key]]).strip() if pn_key and pn_key in col_map and row[col_map[pn_key]] else None
            color = str(row[col_map.get("颜色", 3)]).strip() if "颜色" in col_map and row[col_map["颜色"]] else None
            cost_rmb = _safe_decimal(row[col_map["成本RMB"]]) if "成本RMB" in col_map else Decimal("0")

            # 运费逻辑：
            # - 有独立国家运费列时：用"运费"/"产品运费/台"作为默认运费，国家运费写 dim_freight
            # - 无独立国家运费列时：按优先级取一个值作为通用运费
            if has_country_freight:
                # 默认运费：优先"产品运费/台"，其次"运费"
                freight = Decimal("0")
                for fk in ["产品运费/台", "运费"]:
                    if fk in col_map and row[col_map[fk]]:
                        freight = _safe_decimal(row[col_map[fk]])
                        break
            else:
                freight = Decimal("0")
                for fk in ["产品运费/台", "运费", "其他站点运费", "英国运费", "英国站运费", "英国站点运费"]:
                    if fk in col_map and row[col_map[fk]]:
                        freight = _safe_decimal(row[col_map[fk]])
                        break

            exchange_rate = _safe_decimal(row[col_map["汇率"]]) if "汇率" in col_map else None
            time_val = str(row[col_map["时间"]]).strip() if "时间" in col_map and row[col_map["时间"]] else None

            # 按店铺查找（同ASIN同店铺不跨月分拆）
            product = None
            if store_obj:
                product = db.query(DimProduct).filter(
                    DimProduct.asin == asin,
                    DimProduct.store_id == store_obj.id,
                ).first()

            if not product:
                product = DimProduct(
                    asin=asin, sku=sku or asin, product_name=product_name or "",
                    color=color, store_id=store_obj.id if store_obj else None,
                )
                db.add(product)
                db.flush()
            else:
                if sku: product.sku = sku
                if product_name: product.product_name = product_name
                if color: product.color = color

            # 解析月份并写入 dim_product_cost
            ym = None
            if time_val:
                from datetime import datetime
                try:
                    dt = datetime.strptime(time_val[:10], '%Y-%m-%d')
                    ym = dt.strftime('%Y-%m')
                except: pass
            # fallback: 用导入时选择的年月
            if not ym and import_year and import_month:
                ym = f"{import_year}-{import_month:02d}"
            if ym:
                existing_cost = db.query(DimProductCost).filter(
                    DimProductCost.product_id == product.id,
                    DimProductCost.year_month == ym
                ).first()
                if existing_cost:
                    existing_cost.cost_rmb = cost_rmb
                    existing_cost.freight_per_unit = freight
                else:
                    db.add(DimProductCost(product_id=product.id, year_month=ym, cost_rmb=cost_rmb, freight_per_unit=freight))

            # 独立国家运费写入 dim_freight（upsert）
            if has_country_freight and store_obj:
                for col_name, country_code in _country_freight_map.items():
                    if col_name in col_map and row[col_map[col_name]]:
                        cf_val = _safe_decimal(row[col_map[col_name]])
                        if cf_val > 0 and country_code in country_freight_ids:
                            from sqlalchemy import text as _text
                            db.execute(_text("""
                                INSERT INTO dim_freight (product_id, country_id, store_id, freight_rmb)
                                VALUES (:pid, :cid, :sid, :freight)
                                ON DUPLICATE KEY UPDATE freight_rmb = VALUES(freight_rmb)
                            """), {"pid": product.id, "cid": country_freight_ids[country_code], "sid": store_obj.id, "freight": cf_val})

            count += 1

        db.commit()

        return {"message": "产品信息导入成功", "rows": count}

    except Exception as e:
        db.rollback()
        return {"detail": str(e)}


# ============================================================
# POST /advertising: 导入广告数据 CSV
# ============================================================
@router.post("/advertising")
async def import_advertising(
    file: UploadFile = File(...),
    country: str = Form(..., description="国家代码"),
    store: str = Form(None, description="店铺代码"),
    import_year: int = Form(None, description="导入年份"),
    import_month: int = Form(None, description="导入月份"),
    db: Session = Depends(get_db),
):
    try:
        country = country.upper()
        country_obj = db.query(DimCountry).filter(DimCountry.code == country).first()
        if not country_obj:
            return {"detail": f"国家 {country} 不存在"}
        store_obj = _get_or_default_store(db, store)
        if not store_obj:
            return {"detail": f"店铺 {store} 不存在"}

        if not import_year or not import_month:
            raise HTTPException(status_code=400, detail="必须同时提供 year 和 month 参数")

        # 清除该店铺当月旧广告数据（用 time_id 避免跨月误删）
        _adv_time = db.query(DimTime).filter(DimTime.time_year == import_year, DimTime.time_month == import_month).first()
        _adv_tid = _adv_time.id if _adv_time else None
        if _adv_tid:
            db.query(RawAdvertising).filter(RawAdvertising.store_id == store_obj.id, RawAdvertising.time_id == _adv_tid).delete()
        else:
            db.query(RawAdvertising).filter(RawAdvertising.store_id == store_obj.id).delete()
        db.flush()

        content = await file.read()
        for encoding in ["utf-8-sig", "utf-8", "gbk", "latin-1"]:
            try:
                text = content.decode(encoding)
                break
            except UnicodeDecodeError:
                continue
        else:
            return {"detail": "无法解码文件"}

        reader = csv.DictReader(io.StringIO(text))
        headers = reader.fieldnames
        if not headers:
            return {"detail": "无法解析 CSV 表头"}

        # 中文列名映射
        def get_col(row, *names):
            for name in names:
                if name in row:
                    return row[name]
            return None

        # 按 (ASIN, 年月) 聚合 + 写 raw_advertising
        ad_aggregation = {}  # key: (asin, year, month) -> aggregated values
        row_count = 0
        import re as _re

        for row in reader:
            if not row:
                continue

            # 解析 ASIN 从商品字段（格式 ASIN-SKU）
            product_field = get_col(row, "商品", "ASIN", "asin")
            if not product_field:
                continue

            asin = str(product_field).strip().split("-")[0]

            # 解析 time/date 列确定归属月份
            time_val = get_col(row, "time", "日期", "date")
            ad_year, ad_month = None, None
            if time_val:
                time_str = str(time_val).strip()
                # 支持格式: 2026/5/1, 2026-05-01, 2026-05, May-26, 2026/5
                m = _re.search(r'(\d{4})[/\-](\d{1,2})', time_str)
                if m:
                    ad_year = int(m.group(1))
                    ad_month = int(m.group(2))
                else:
                    m = _re.search(r'(\w+)[\-/](\d{2,4})', time_str)
                    if m:
                        month_map = {"jan":1,"feb":2,"mar":3,"apr":4,"may":5,"jun":6,
                                     "jul":7,"aug":8,"sep":9,"oct":10,"nov":11,"dec":12}
                        month_str = m.group(1).lower()[:3]
                        year_str = m.group(2)
                        if month_str in month_map:
                            ad_month = month_map[month_str]
                            ad_year = int(year_str) if len(year_str) == 4 else int("20" + year_str) if len(year_str) == 2 else None

            ad_spend = _safe_decimal(get_col(row, "花费(USD)", "花费(CAD)", "花费(MX)", "花费(EUR)", "花费(GBP)", "花费(SEK)", "花费(AUD)", "花费", "支出(USD)", "支出(CAD)", "支出(MX)", "支出(EUR)", "支出(GBP)", "支出(SEK)", "支出(AUD)"))
            ad_sales = _safe_decimal(get_col(row, "销售额(USD)", "销售额(CAD)", "销售额(MX)", "销售额"))
            acos_val = _safe_decimal(get_col(row, "ACOS", "acos"))
            roas_val = _safe_decimal(get_col(row, "ROAS", "roas"))
            ctr_val = _safe_decimal(get_col(row, "CTR", "ctr"))
            cpc_val = _safe_decimal(get_col(row, "CPC", "cpc"))
            impressions_val = _safe_int(get_col(row, "展示次数", "impressions"))
            clicks_val = _safe_int(get_col(row, "点击量", "clicks"))
            ad_orders = _safe_int(get_col(row, "订单数量", "ad_orders"))
            conversion = _safe_decimal(get_col(row, "转化率", "conversion_rate"))
            ntb_orders = _safe_int(get_col(row, "NTB 订单数量"))
            ntb_order_pct = _safe_decimal(get_col(row, "NTB 订单数量百分比"))
            ntb_sales = _safe_decimal(get_col(row, "NTB 销售额(USD)"))
            new_to_brand_pct = _safe_decimal(get_col(row, "品牌新客销售额比例"))
            visible_imp = _safe_int(get_col(row, "可见展示量"))

            # 写入 raw_advertising
            raw_adv = RawAdvertising(
                country_id=country_obj.id,
                store_id=store_obj.id,
                product_field=_safe_str(product_field, 200),
                asin=_safe_str(asin, 50),
                status_val=_safe_str(get_col(row, "状态") or "", 50),
                ad_type=_safe_str(get_col(row, "类型") or "", 50),
                eligibility=_safe_str(get_col(row, "商品推广使用资格") or "", 100),
                sales_usd=ad_sales,
                roas=roas_val,
                conversion_rate=conversion,
                impressions=impressions_val,
                clicks=clicks_val,
                ctr=ctr_val,
                spend_usd=ad_spend,
                cpc=cpc_val,
                orders=ad_orders,
                acos=acos_val,
                ntb_orders=ntb_orders,
                ntb_order_pct=ntb_order_pct,
                ntb_sales_usd=ntb_sales,
                new_to_brand_sales_pct=new_to_brand_pct,
                visible_impressions=visible_imp,
                raw_data=dict(row),
                time_id=_adv_tid,
            )
            db.add(raw_adv)

            # 按月聚合: key = (asin, year, month)
            ym_key = (asin, ad_year, ad_month) if ad_year and ad_month else (asin, None, None)
            if ym_key not in ad_aggregation:
                ad_aggregation[ym_key] = {
                    "ad_spend": Decimal("0"),
                    "ad_sales": Decimal("0"),
                    "acos": Decimal("0"),
                    "roas": Decimal("0"),
                    "ctr": Decimal("0"),
                    "cpc": Decimal("0"),
                    "impressions": 0,
                    "clicks": 0,
                    "ad_orders": 0,
                    "conversion_rate": Decimal("0"),
                    "count": 0,
                }
            agg = ad_aggregation[ym_key]
            agg["ad_spend"] += ad_spend
            agg["ad_sales"] += ad_sales
            agg["acos"] += acos_val
            agg["roas"] += roas_val
            agg["ctr"] += ctr_val
            agg["cpc"] += cpc_val
            agg["impressions"] += impressions_val
            agg["clicks"] += clicks_val
            agg["ad_orders"] += ad_orders
            agg["conversion_rate"] += conversion
            agg["count"] += 1
            row_count += 1

        # 更新 monthly_summary
        summary_count = 0

        for (asin, ad_year, ad_month), agg in ad_aggregation.items():
            # 以导入时选择的月份为准，不看数据内日期
            _ym = f"{import_year}-{import_month:02d}" if import_year and import_month else None
            product = _find_product_by_asin(db, asin, store_id=store_obj.id if store_obj else None, year_month=_ym)
            if not product:
                continue

            exchange_rate = _get_exchange_rate(db, country_obj, import_year, import_month, store_id=store_obj.id)

            # 确定目标月份（以导入时选择的月份为准，不看数据内日期）
            if import_year and import_month:
                time_obj = _get_or_create_time(db, import_year, import_month)
                time_id = time_obj.id
                filters = [
                    MonthlySummary.product_id == product.id,
                    MonthlySummary.country_id == country_obj.id,
                    MonthlySummary.time_id == time_id,
                ]
                if store_obj:
                    filters.append(MonthlySummary.store_id == store_obj.id)
                summary = db.query(MonthlySummary).filter(*filters).first()
                if not summary:
                    summary = MonthlySummary(
                        country_id=country_obj.id, product_id=product.id, time_id=time_id,
                        store_id=store_obj.id,
                        order_count=0, order_qty=0,
                        product_sales_usd=Decimal("0"), commission_usd=Decimal("0"), fba_fee_usd=Decimal("0"),
                        ad_spend_usd=Decimal("0"), storage_fee_usd=Decimal("0"), returns_fee_usd=Decimal("0"), inbound_fee_usd=Decimal("0"),
                        removal_fee_usd=Decimal("0"),
                    )
                    db.add(summary)
                target_summaries = [summary]
            else:
                # 无时间信息，更新所有月份（兼容旧逻辑）
                filters = [MonthlySummary.product_id == product.id, MonthlySummary.country_id == country_obj.id]
                if store_obj:
                    filters.append(MonthlySummary.store_id == store_obj.id)
                target_summaries = db.query(MonthlySummary).filter(*filters).all()
                if not target_summaries:
                    continue

            for summary in target_summaries:
                summary.ad_spend_usd = agg["ad_spend"]
                summary.ad_sales_usd = agg["ad_sales"]
                summary.acos = (agg["acos"] / agg["count"]).quantize(Decimal("0.0001")) if agg["count"] else Decimal("0")
                summary.roas = (agg["roas"] / agg["count"]).quantize(Decimal("0.0001")) if agg["count"] else Decimal("0")
                summary.ctr = (agg["ctr"] / agg["count"]).quantize(Decimal("0.0001")) if agg["count"] else Decimal("0")
                summary.cpc = (agg["cpc"] / agg["count"]).quantize(Decimal("0.01")) if agg["count"] else Decimal("0")
                summary.impressions = agg["impressions"]
                summary.clicks = agg["clicks"]
                summary.ad_orders = agg["ad_orders"]
                summary.conversion_rate = (agg["conversion_rate"] / agg["count"]).quantize(Decimal("0.0001")) if agg["count"] else Decimal("0")

                # 统一利润公式
                summary.ad_spend_usd = agg["ad_spend"]
                from services.profit import apply_profit_to_summary
                apply_profit_to_summary(summary, exchange_rate)

                summary_count += 1

        db.commit()

        return {
            "message": "广告数据导入成功",
            "csv_rows": row_count,
            "summary_updated": summary_count,
            "country": country,
        }

    except Exception as e:
        db.rollback()
        return {"detail": str(e)}


# ============================================================
# POST /storage: 导入仓储费 CSV
# ============================================================
@router.post("/storage")
async def import_storage(
    file: UploadFile = File(...),
    country: str = Form(..., description="国家代码"),
    store: str = Form(None, description="店铺代码"),
    db: Session = Depends(get_db),
):
    try:
        country = country.upper()
        country_obj = db.query(DimCountry).filter(DimCountry.code == country).first()
        if not country_obj:
            return {"detail": f"国家 {country} 不存在"}
        store_obj = _get_or_default_store(db, store)

        # 清空该店铺+国家的旧仓储费 raw 数据（防止重复导入累积）
        clear_filters = [
            RawStorageFee.country_id == country_obj.id,
        ]
        if store_obj:
            clear_filters.append(RawStorageFee.store_id == store_obj.id)
        else:
            clear_filters.append(RawStorageFee.store_id.is_(None))
        db.query(RawStorageFee).filter(*clear_filters).delete()

        content = await file.read()
        for encoding in ["utf-8-sig", "utf-8", "gbk", "latin-1"]:
            try:
                text = content.decode(encoding)
                break
            except UnicodeDecodeError:
                continue
        else:
            return {"detail": "无法解码文件"}

        reader = csv.DictReader(io.StringIO(text))
        headers = reader.fieldnames

        def get_col(row, *names):
            for name in names:
                if name in row:
                    return row[name]
            return None

        # 按 (ASIN, 月份) 汇总 + 写 raw_storage_fee
        asin_fees = {}  # key: (asin, month_str)
        row_count = 0
        import re as _re

        for row in reader:
            if not row:
                continue

            asin = None
            for key in ["asin", "ASIN", "fnsku", "FNSKU"]:
                val = get_col(row, key)
                if val:
                    asin = str(val).strip()
                    break

            if not asin:
                continue

            fee = _safe_decimal(get_col(row, "estimated_monthly_storage_fee", "estimated-monthly-storage-fee", "storage_fee", "月度仓储费（预计）"))
            month_str = str(get_col(row, "month_of_charge", "收费月份") or "").strip()

            # 写入 raw_storage_fee
            raw_fee = RawStorageFee(
                country_id=country_obj.id,
                store_id=store_obj.id if store_obj else None,
                asin=asin,
                fnsku=str(get_col(row, "fnsku", "FNSKU") or "").strip(),
                product_name=str(get_col(row, "product_name", "product-name", "product-name") or "").strip(),
                fulfillment_center=str(get_col(row, "fulfillment_center", "亚马逊运营中心") or "").strip(),
                country_code=str(get_col(row, "country_code", "国家/地区代码") or "").strip(),
                longest_side=_safe_decimal(get_col(row, "longest_side", "最长边")),
                median_side=_safe_decimal(get_col(row, "median_side", "次长边")),
                shortest_side=_safe_decimal(get_col(row, "shortest_side", "最短边")),
                measurement_units=str(get_col(row, "measurement_units", "计量单位") or "").strip(),
                weight=_safe_decimal(get_col(row, "weight")),
                weight_units=str(get_col(row, "weight_units", "重量单位") or "").strip(),
                item_volume=_safe_decimal(get_col(row, "item_volume", "商品体积")),
                volume_units=str(get_col(row, "volume_units", "体积单位") or "").strip(),
                product_size_tier=str(get_col(row, "product_size_tier", "商品尺寸分段") or "").strip(),
                average_quantity_on_hand=_safe_decimal(get_col(row, "average_quantity_on_hand", "现货的平均数量")),
                average_quantity_pending_removal=_safe_decimal(get_col(row, "average_quantity_pending_removal", "等待移除的平均数量")),
                estimated_total_item_volume=_safe_decimal(get_col(row, "estimated_total_item_volume", "商品总体积（预计）")),
                month_of_charge=month_str,
                storage_utilization_ratio=_safe_decimal(get_col(row, "storage_utilization_ratio", "storage_utilization_ratio_units")),
                storage_utilization_ratio_units=str(get_col(row, "storage_utilization_ratio_units") or "").strip(),
                base_rate=_safe_decimal(get_col(row, "base_rate", "仓储费费率")),
                utilization_surcharge_rate=_safe_decimal(get_col(row, "utilization_surcharge_rate")),
                avg_qty_for_sus=_safe_decimal(get_col(row, "avg_qty_for_sus")),
                est_vol_for_sus=_safe_decimal(get_col(row, "est_vol_for_sus")),
                est_base_msf=_safe_decimal(get_col(row, "est_base_msf")),
                est_sus=_safe_decimal(get_col(row, "est_sus")),
                currency=str(get_col(row, "currency") or "").strip(),
                estimated_monthly_storage_fee=fee,
                dangerous_goods_storage_type=str(get_col(row, "dangerous_goods_storage_type") or "").strip(),
                eligible_for_inventory_discount=str(get_col(row, "eligible_for_inventory_discount") or "").strip(),
                qualifies_for_inventory_discount=str(get_col(row, "qualifies_for_inventory_discount") or "").strip(),
                total_incentive_fee_amount=_safe_decimal(get_col(row, "total_incentive_fee_amount")),
                breakdown_incentive_fee_amount=_safe_decimal(get_col(row, "breakdown_incentive_fee_amount")),
                average_quantity_customer_orders=_safe_decimal(get_col(row, "average_quantity_customer_orders")),
                raw_data=dict(row),
            )
            db.add(raw_fee)

            key = (asin, month_str)
            if key not in asin_fees:
                asin_fees[key] = Decimal("0")
            asin_fees[key] += fee
            row_count += 1

        summary_count = 0

        for (asin, month_str), total_fee in asin_fees.items():
            # 用导入月份作为year_month（因为费用数据是按月导入的）
            fee_ym = f"{import_year}-{import_month:02d}" if import_year and import_month else None
            product = _find_product_by_asin(db, asin, store_id=store_obj.id if store_obj else None, year_month=fee_ym)
            if not product:
                continue

            # 解析月份并找到对应的 DimTime（支持 Apr-26 / 2026-05 等多种格式）
            time_obj_for_rate = None
            if month_str:
                time_obj_for_rate = _find_time_by_month_str(db, month_str)
                if time_obj_for_rate:
                    filters = [
                        MonthlySummary.product_id == product.id,
                        MonthlySummary.country_id == country_obj.id,
                        MonthlySummary.time_id == time_obj_for_rate.id,
                    ]
                    if store_obj:
                        filters.append(MonthlySummary.store_id == store_obj.id)
                    summary = db.query(MonthlySummary).filter(*filters).first()
                    if not summary:
                        summary = MonthlySummary(
                            country_id=country_obj.id, product_id=product.id, time_id=time_obj_for_rate.id,
                            store_id=store_obj.id if store_obj else None,
                            order_count=0, order_qty=0,
                            product_sales_usd=Decimal("0"), commission_usd=Decimal("0"), fba_fee_usd=Decimal("0"),
                            ad_spend_usd=Decimal("0"), storage_fee_usd=Decimal("0"), returns_fee_usd=Decimal("0"), inbound_fee_usd=Decimal("0"),
                            removal_fee_usd=Decimal("0"),
                        )
                        db.add(summary)
                    target_summaries = [summary]
                else:
                    target_summaries = []
            else:
                filters = [MonthlySummary.product_id == product.id, MonthlySummary.country_id == country_obj.id]
                if store_obj:
                    filters.append(MonthlySummary.store_id == store_obj.id)
                target_summaries = db.query(MonthlySummary).filter(*filters).all()

            # 按月份获取汇率
            ym_parts = None
            if time_obj_for_rate:
                ym_parts = time_obj_for_rate.year_month.split("-")
            exchange_rate = _get_exchange_rate(db, country_obj, int(ym_parts[0]) if ym_parts else None, int(ym_parts[1]) if ym_parts else None, store_id=store_obj.id)

            for summary in target_summaries:
                summary.storage_fee_usd = (summary.storage_fee_usd or Decimal("0")) + total_fee

                # 统一利润公式
                from services.profit import apply_profit_to_summary
                apply_profit_to_summary(summary, exchange_rate)
                summary_count += 1

        db.commit()

        return {
            "message": "仓储费导入成功",
            "csv_rows": row_count,
            "summary_updated": summary_count,
        }

    except Exception as e:
        db.rollback()
        return {"detail": str(e)}


# ============================================================
# POST /returns: 导入退货费 CSV
# ============================================================
@router.post("/returns")
async def import_returns(
    file: UploadFile = File(...),
    country: str = Form(..., description="国家代码"),
    store: str = Form(None, description="店铺代码"),
    db: Session = Depends(get_db),
):
    try:
        country = country.upper()
        country_obj = db.query(DimCountry).filter(DimCountry.code == country).first()
        if not country_obj:
            return {"detail": f"国家 {country} 不存在"}
        store_obj = _get_or_default_store(db, store)

        # 清空该店铺+国家的旧退货费 raw 数据（防止重复导入累积）
        clear_filters = [RawReturns.country_id == country_obj.id]
        if store_obj:
            clear_filters.append(RawReturns.store_id == store_obj.id)
        else:
            clear_filters.append(RawReturns.store_id.is_(None))
        db.query(RawReturns).filter(*clear_filters).delete()

        content = await file.read()
        for encoding in ["utf-8-sig", "utf-8", "gbk", "latin-1"]:
            try:
                text = content.decode(encoding)
                break
            except UnicodeDecodeError:
                continue
        else:
            return {"detail": "无法解码文件"}

        reader = csv.DictReader(io.StringIO(text))
        headers = reader.fieldnames

        def get_col(row, *names):
            for name in names:
                if name in row:
                    return row[name]
            return None

        # 按 ASIN 汇总 + 写 raw_returns
        asin_fees = {}
        row_count = 0

        for row in reader:
            if not row:
                continue

            asin = None
            for key in ["asin", "ASIN", "fnsku", "FNSKU"]:
                val = get_col(row, key)
                if val:
                    asin = str(val).strip()
                    break

            if not asin:
                continue

            fee = _safe_decimal(get_col(row, "sku_returns_fee", "returns_fee", "退货费"))

            # 写入 raw_returns
            raw_ret = RawReturns(
                country_id=country_obj.id,
                store_id=store_obj.id if store_obj else None,
                asin=asin,
                asin_fee_category=str(get_col(row, "asin_fee_category") or "").strip(),
                fnsku=str(get_col(row, "fnsku", "FNSKU") or "").strip(),
                product_name=str(get_col(row, "product_name") or "").strip(),
                longest_side=_safe_decimal(get_col(row, "longest_side")),
                median_side=_safe_decimal(get_col(row, "median_side")),
                shortest_side=_safe_decimal(get_col(row, "shortest_side")),
                measurement_units=str(get_col(row, "measurement-units", "measurement_units") or "").strip(),
                unit_weight=_safe_decimal(get_col(row, "unit_weight")),
                dimensional_weight=_safe_decimal(get_col(row, "dimensional_weight")),
                shipping_weight=_safe_decimal(get_col(row, "shipping_weight")),
                weight_units=str(get_col(row, "weight_units") or "").strip(),
                sku_sizetier=str(get_col(row, "sku_sizetier") or "").strip(),
                month_of_shipment=str(get_col(row, "month_of_shipment") or "").strip(),
                asin_shipped_units=_safe_int(get_col(row, "asin_shipped_units")),
                asin_return_threshold_percent=_safe_decimal(get_col(row, "asin_return_threshold_percent")),
                asin_return_threshold_units=_safe_int(get_col(row, "asin_return_threshold_units")),
                asin_returned_units=_safe_int(get_col(row, "asin_returned_units")),
                sku_returned_units_nsp_exempted=_safe_int(get_col(row, "sku_returned_units_NSP_exempted")),
                sku_returned_units_charged=_safe_int(get_col(row, "sku_returned_units_charged")),
                sku_fee_per_unit=_safe_decimal(get_col(row, "sku_fee_per_unit")),
                sku_returns_fee=fee,
                month_of_charge=str(get_col(row, "month_of_charge") or "").strip(),
                currency=str(get_col(row, "currency") or "").strip(),
                raw_data=dict(row),
            )
            db.add(raw_ret)

            if asin not in asin_fees:
                asin_fees[asin] = Decimal("0")
            asin_fees[asin] += fee
            row_count += 1

        summary_count = 0

        for asin, total_fee in asin_fees.items():
            # 严格按店铺隔离查产品，绝不跨店
            product = db.query(DimProduct).filter(
                DimProduct.asin == asin,
                DimProduct.store_id == store_obj.id if store_obj else None,
            ).first() if store_obj else db.query(DimProduct).filter(DimProduct.asin == asin).first()
            if not product:
                continue

            filters = [MonthlySummary.product_id == product.id, MonthlySummary.country_id == country_obj.id]
            if store_obj:
                filters.append(MonthlySummary.store_id == store_obj.id)
            summaries = (
                db.query(MonthlySummary)
                .filter(*filters)
                .all()
            )

            for summary in summaries:
                summary.returns_fee_usd = total_fee

                # 按月份获取汇率
                time_obj = db.query(DimTime).filter(DimTime.id == summary.time_id).first()
                ym = time_obj.year_month.split("-") if time_obj and time_obj.year_month else None
                exchange_rate = _get_exchange_rate(db, country_obj, int(ym[0]) if ym else None, int(ym[1]) if ym else None, store_id=store_obj.id)

                from services.profit import apply_profit_to_summary
                apply_profit_to_summary(summary, exchange_rate)
                summary_count += 1

        db.commit()

        return {
            "message": "退货费导入成功",
            "csv_rows": row_count,
            "summary_updated": summary_count,
        }

    except Exception as e:
        db.rollback()
        return {"detail": str(e)}


# ============================================================
# POST /inbound: 导入入库费 CSV
# ============================================================
@router.post("/inbound")
async def import_inbound(
    file: UploadFile = File(...),
    country: str = Form(..., description="国家代码"),
    store: str = Form(None, description="店铺代码"),
    db: Session = Depends(get_db),
):
    try:
        country = country.upper()
        country_obj = db.query(DimCountry).filter(DimCountry.code == country).first()
        if not country_obj:
            return {"detail": f"国家 {country} 不存在"}
        store_obj = _get_or_default_store(db, store)

        # 清空该店铺+国家的旧入库费 raw 数据（防止重复导入累积）
        clear_filters = [RawInbound.country_id == country_obj.id]
        if store_obj:
            clear_filters.append(RawInbound.store_id == store_obj.id)
        else:
            clear_filters.append(RawInbound.store_id.is_(None))
        db.query(RawInbound).filter(*clear_filters).delete()

        content = await file.read()
        for encoding in ["utf-8-sig", "utf-8", "gbk", "latin-1"]:
            try:
                text = content.decode(encoding)
                break
            except UnicodeDecodeError:
                continue
        else:
            return {"detail": "无法解码文件"}

        reader = csv.DictReader(io.StringIO(text))
        headers = reader.fieldnames

        def get_col(row, *names):
            for name in names:
                if name in row:
                    return row[name]
            return None

        # 按 ASIN 汇总 + 写 raw_inbound
        asin_fees = {}
        row_count = 0

        for row in reader:
            if not row:
                continue

            asin = None
            for key in ["asin", "ASIN", "fnsku", "FNSKU"]:
                val = get_col(row, key)
                if val:
                    asin = str(val).strip()
                    break

            if not asin:
                continue

            fee = _safe_decimal(
                get_col(
                    row,
                    "亚马逊物流入库配置服务费用总计",
                    "inbound_fee",
                    "入库费",
                    "fba_inbound_placement_fee",
                )
            )

            # 写入 raw_inbound
            txn_date_str = str(get_col(row, "交易日期") or "").strip()
            txn_date_val = _detect_date_format(txn_date_str) if txn_date_str else None

            raw_inb = RawInbound(
                country_id=country_obj.id,
                store_id=store_obj.id if store_obj else None,
                transaction_date=txn_date_val,
                inbound_plan_id=str(get_col(row, "入库计划编号") or "").strip(),
                fba_shipment_id=str(get_col(row, "亚马逊物流货件编号") or "").strip(),
                country_region=str(get_col(row, "国家/地区") or "").strip(),
                fnsku=str(get_col(row, "FNSKU", "fnsku") or "").strip(),
                asin=asin,
                planned_inbound_service=str(get_col(row, "计划的亚马逊物流入库配置服务") or "").strip(),
                planned_shipment_qty=_safe_int(get_col(row, "计划货件数量")),
                eligible_shipment_qty=_safe_int(get_col(row, "符合要求的货件数量")),
                inbound_defect_type=str(get_col(row, "入库缺陷类型") or "").strip(),
                actual_fee_segment=str(get_col(row, "实际费用分段") or "").strip(),
                planned_inbound_region=str(get_col(row, "计划入库区域") or "").strip(),
                actual_inbound_region=str(get_col(row, "实际入库区域") or "").strip(),
                actual_received_qty=_safe_int(get_col(row, "实际接收数量")),
                product_size_segment=str(get_col(row, "商品尺寸分段") or "").strip(),
                shipping_weight=_safe_decimal(get_col(row, "发货重量")),
                weight_unit=str(get_col(row, "重量单位") or "").strip(),
                inbound_placement_fee_rate=_safe_decimal(get_col(row, "亚马逊物流入库配置服务费率（按商品）")),
                eligible_actual_incentive=_safe_decimal(get_col(row, "符合条件的实际奖励额")),
                currency=str(get_col(row, "货币") or "").strip(),
                inbound_placement_fee_total=fee,
                total_fee=_safe_decimal(get_col(row, "总费用")),
                raw_data=dict(row),
            )
            db.add(raw_inb)

            if asin not in asin_fees:
                asin_fees[asin] = Decimal("0")
            asin_fees[asin] += fee
            row_count += 1

        summary_count = 0

        for asin, total_fee in asin_fees.items():
            # 严格按店铺隔离查产品，绝不跨店
            product = db.query(DimProduct).filter(
                DimProduct.asin == asin,
                DimProduct.store_id == store_obj.id if store_obj else None,
            ).first() if store_obj else db.query(DimProduct).filter(DimProduct.asin == asin).first()
            if not product:
                continue

            filters = [MonthlySummary.product_id == product.id, MonthlySummary.country_id == country_obj.id]
            if store_obj:
                filters.append(MonthlySummary.store_id == store_obj.id)
            summaries = (
                db.query(MonthlySummary)
                .filter(*filters)
                .all()
            )

            for summary in summaries:
                summary.inbound_fee_usd = total_fee

                # 按月份获取汇率
                time_obj = db.query(DimTime).filter(DimTime.id == summary.time_id).first()
                ym = time_obj.year_month.split("-") if time_obj and time_obj.year_month else None
                exchange_rate = _get_exchange_rate(db, country_obj, int(ym[0]) if ym else None, int(ym[1]) if ym else None, store_id=store_obj.id)

                from services.profit import apply_profit_to_summary
                apply_profit_to_summary(summary, exchange_rate)
                summary_count += 1

        db.commit()

        return {
            "message": "入库费导入成功",
            "csv_rows": row_count,
            "summary_updated": summary_count,
        }

    except Exception as e:
        db.rollback()
        return {"detail": str(e)}


@router.post("/workbook")
async def import_workbook(
    file: UploadFile = File(...),
    country: str = Form("US", description="国家代码"),
    store: str = Form(None, description="店铺代码"),
    import_year: int = Form(None, description="导入年份"),
    import_month: int = Form(None, description="导入月份"),
    db: Session = Depends(get_db),
):
    """上传合并后的工作簿（如 美国站全部数据.xlsx），自动识别每个 sheet 并导入"""
    try:
        import openpyxl

        store_obj = _get_or_default_store(db, store)
        if not store_obj:
            return {"detail": f"店铺 {store} 不存在，请先在系统管理创建"}

        if not import_year or not import_month:
            raise HTTPException(status_code=400, detail="必须同时提供 year 和 month 参数")

        from sqlalchemy import extract, text as _text, func as _func
        _clear_time = db.query(DimTime).filter(DimTime.time_year == import_year, DimTime.time_month == import_month).first()
        _clear_filters = [RawTransaction.store_id == store_obj.id]
        if _clear_time:
            _clear_filters.append(RawTransaction.time_id == _clear_time.id)
        db.query(RawTransaction).filter(*_clear_filters).delete()

        # RawAdvertising：按店铺+月份清空
        if _clear_time:
            db.query(RawAdvertising).filter(RawAdvertising.store_id == store_obj.id, RawAdvertising.time_id == _clear_time.id).delete()
        else:
            db.query(RawAdvertising).filter(RawAdvertising.store_id == store_obj.id).delete()

        # 仓储费/退货/入库：按店铺+月份清空
        _fee_month_prefix = f"{import_year}-{import_month:02d}"
        db.query(RawStorageFee).filter(RawStorageFee.store_id == store_obj.id, RawStorageFee.month_of_charge.like(f"{_fee_month_prefix}%")).delete()
        db.query(RawReturns).filter(RawReturns.store_id == store_obj.id, RawReturns.month_of_charge.like(f"{_fee_month_prefix}%")).delete()
        db.query(RawInbound).filter(RawInbound.store_id == store_obj.id, _func.date_format(RawInbound.transaction_date, '%Y-%m') == _fee_month_prefix).delete()
        db.query(RawLongTermStorage).filter(RawLongTermStorage.store_id == store_obj.id).delete()

        # MonthlySummary + 汇率：按店铺+月份清空
        _time_obj = db.query(DimTime).filter(DimTime.time_year == import_year, DimTime.time_month == import_month).first()
        if _time_obj:
            db.query(MonthlySummary).filter(MonthlySummary.store_id == store_obj.id, MonthlySummary.time_id == _time_obj.id).delete()
            db.query(DimExchangeRate).filter(DimExchangeRate.store_id == store_obj.id, DimExchangeRate.year_month == f"{import_year}-{import_month:02d}").delete()
        db.flush()

        content = await file.read()
        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True)

        results = {}

        # ===== 识别所有 sheet 类型 =====
        _txn_date_headers = ('date/time', 'date / time', 'fecha/hora', 'fecha y hora', 'datum/uhrzeit',
                             'date/heure', 'data/ora', 'data/ora:', 'datum/tijd', 'datum/tid')
        def identify_sheet(header, rows):
            """返回 (sheet_type, header, rows)"""
            # 1. Row 1 直接匹配交易表头
            if any(h in _txn_date_headers for h in header):
                return "transaction", header, rows[1:]

            # 2. 扫描前15行，找交易表头行：必须同时包含 date/time 和 sku
            _txn_keywords = {
                'date/time', 'date / time', 'fecha/hora', 'fecha y hora', 'datum/uhrzeit',
                'date/heure', 'data/ora', 'data/ora:', 'datum/tijd', 'datum/tid',
                'sku', 'order id', 'bestellnummer', 'beställnings-id', 'bestelnummer',
                'numero ordine', 'numéro de la commande', 'número de pedido',
                'product sales', 'selling fees', 'fba fees', 'total',
                'quantity', 'menge', 'cantidad', 'quantité', 'quantità', 'antal', 'aantal',
                'type', 'typ', 'tipo',
            }
            best_score = 0
            best_idx = -1
            for i in range(1, min(15, len(rows))):
                row_cells = [str(v).strip().lower() if v else "" for v in rows[i]]
                has_date = any(c in _txn_date_headers for c in row_cells)
                has_sku = 'sku' in row_cells
                if has_date and has_sku:
                    score = sum(1 for c in row_cells if c in _txn_keywords)
                    if score > best_score:
                        best_score = score
                        best_idx = i
            if best_idx >= 0:
                row_h = [str(v).strip().lower() if v else "" for v in rows[best_idx]]
                return "transaction", row_h, rows[best_idx + 1:]

            # 3. 其他类型识别
            header_set = set(h for h in header if h)
            data_rows = rows[1:]
            if "asin" in header_set and any("成本" in h or "cost" in h for h in header):
                return "product_info", header, data_rows
            if any(kw in h for kw in ("商品", "campaign", "asin", "product", "advertised") for h in header) \
                    and any(kw in h for kw in ("花费", "支出", "spend", "cost", "acos", "roas") for h in header):
                return "advertising", header, data_rows
            if any("returns_fee" in h or "returned_units" in h for h in header):
                return "returns", header, data_rows
            if any("入库" in h or "inbound" in h or "shipped_units" in h for h in header):
                return "inbound", header, data_rows
            if any("estimated_monthly_storage_fee" in h or "estimated_total_item_volume" in h or "月度仓储费" in h or "亚马逊运营中心" in h for h in header):
                return "storage", header, data_rows
            if any("amount-charged" in h or "amount_charged" in h for h in header):
                if any("surcharge" in h or "snapshot" in h for h in header):
                    return "long_term_storage", header, data_rows
                return "storage", header, data_rows
            # 汇率表：含"国家"和"汇率"列
            if "汇率" in header and ("国家" in header or "country" in header):
                return "exchange_rate", header, data_rows
            # 移除费：含 request-date 和 removal-fee
            if "request-date" in header and "removal-fee" in header:
                return "removal_fee", header, data_rows
            return None, header, data_rows

        # 解析所有 sheet
        sheets = {}  # name -> (type, header, rows)
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            all_rows = list(ws.iter_rows(values_only=True))
            if not all_rows or len(all_rows) < 2:
                results[sheet_name] = {"status": "skipped", "reason": "数据不足"}
                continue
            header = [str(h).strip().lower() if h else "" for h in all_rows[0]]
            stype, h, r = identify_sheet(header, all_rows)
            if not stype:
                results[sheet_name] = {"status": "skipped", "reason": "无法识别类型", "headers": header[:5]}
                continue
            sheets[sheet_name] = (stype, h, r)

        # ===== 自动检测国家（每个sheet独立检测，支持多国家工作簿）=====
        # 多国家 sheet 类型：从数据行的 country_code/country 列读取国家，不按 sheet 分配
        _multi_country_types = {"storage", "long_term_storage", "returns", "inbound", "exchange_rate", "removal_fee"}
        _sheet_country_kw = {
            '英国': 'UK', 'UK': 'UK',
            '德国': 'DE', 'DE': 'DE',
            '法国': 'FR', 'FR': 'FR', 'FA': 'FR',  # FA = France abbreviation in some files
            '西班牙': 'ES', 'ES': 'ES',
            '意大利': 'IT', 'IT': 'IT',
            '荷兰': 'NL', 'NL': 'NL',
            '瑞典': 'SE', 'SE': 'SE',
            '比利时': 'BE', 'BE': 'BE',
            '爱尔兰': 'IE', 'IE': 'IE',
            '阿联酋': 'AE', 'AE': 'AE',
            '沙特': 'SA', 'SA': 'SA',
            '澳大利亚': 'AU', '澳洲': 'AU', 'AU': 'AU',
        }
        sheet_countries = {}  # sheet_name -> country_code (仅单国家sheet)
        for sheet_name, (stype, header, rows) in sheets.items():
            if country and country.upper() != 'AUTO':
                sheet_countries[sheet_name] = country.upper()
            elif stype in _multi_country_types:
                # 多国家sheet：不分配默认国家，由处理函数从数据行读取
                sheet_countries[sheet_name] = None
            else:
                detected = _detect_country_from_data(db, header, rows)
                if not detected:
                    sn = sheet_name
                    sn_upper = sn.upper()
                    for kw, cc in _sheet_country_kw.items():
                        if kw in sn or kw in sn_upper:
                            detected = cc; break
                sheet_countries[sheet_name] = detected  # 可能为None，后续跳过

        # 加载所有国家对象（多国家sheet需要从数据行读取国家）
        country_objs = {}
        for co in db.query(DimCountry).all():
            country_objs[co.code] = co

        # ===== 第零轮：汇率表（优先导入，后续计算利润需要）=====
        for sheet_name, (stype, header, rows) in sheets.items():
            if stype != "exchange_rate":
                continue
            try:
                # 解析列索引
                country_idx = next(i for i, h in enumerate(header) if h in ("国家", "country"))
                rate_idx = next(i for i, h in enumerate(header) if h == "汇率")
                # 确定月份：优先用传入参数，否则取当前月
                ym = None
                if import_year and import_month:
                    ym = f"{import_year}-{import_month:02d}"
                else:
                    from datetime import date
                    today = date.today()
                    ym = f"{today.year}-{today.month:02d}"
                imported = 0
                for row in rows:
                    cc = str(row[country_idx] or "").strip().upper()
                    rate = row[rate_idx]
                    if not cc or rate is None:
                        continue
                    co = db.query(DimCountry).filter(DimCountry.code == cc).first()
                    if not co:
                        continue
                    from decimal import Decimal as _D
                    rate_val = _D(str(rate))
                    # upsert: 查找已有记录（按店铺+国家+月份）
                    existing = db.query(DimExchangeRate).filter(
                        DimExchangeRate.country_id == co.id,
                        DimExchangeRate.year_month == ym,
                        DimExchangeRate.store_id == store_obj.id,
                    ).first()
                    if existing:
                        existing.rate = rate_val
                    else:
                        db.add(DimExchangeRate(country_id=co.id, year_month=ym, rate=rate_val, store_id=store_obj.id))
                    imported += 1
                db.flush()
                results[sheet_name] = {"status": "success", "type": "exchange_rate", "month": ym, "imported": imported}
            except Exception as e:
                results[sheet_name] = {"status": "error", "type": "exchange_rate", "detail": str(e)}

        # ===== 第一轮：先产品信息，再交易记录（确保SKU匹配）=====
        product_info_sheets = [(n, h, r) for n, (t, h, r) in sheets.items() if t == "product_info"]
        transaction_sheets = [(n, h, r) for n, (t, h, r) in sheets.items() if t == "transaction"]

        for sheet_name, header, rows in product_info_sheets:
            try:
                result = _process_product_info_sheet(db, header, rows, import_year=import_year, import_month=import_month, store_id=store_obj.id if store_obj else None)
                results[sheet_name] = {"status": "success", "type": "product_info", **result}
            except Exception as e:
                results[sheet_name] = {"status": "error", "type": "product_info", "detail": str(e)}

        for sheet_name, header, rows in transaction_sheets:
            cc = sheet_countries.get(sheet_name)
            co = country_objs.get(cc) if cc else None
            # 交易表如果数据中检测不到国家，从sheet名推断
            if not co:
                sn = sheet_name
                sn_upper = sn.upper()
                for kw, c in _sheet_country_kw.items():
                    if kw in sn or kw in sn_upper:
                        cc = c; break
                co = country_objs.get(cc)
            if not co:
                results[sheet_name] = {"status": "error", "type": "transaction", "detail": f"无法检测国家: {sheet_name}"}
                continue
            try:
                result = _process_transaction_sheet(db, co, header, rows, store_id=store_obj.id, import_year=import_year, import_month=import_month)
                results[sheet_name] = {"status": "success", "type": "transaction", "country": cc, **result}
            except Exception as e:
                results[sheet_name] = {"status": "error", "type": "transaction", "country": cc, "detail": str(e)}

        # ===== 补建所有产品的 summary（确保广告/仓储等数据有记录可更新）=====
        for cc, co in country_objs.items():
            _ensure_all_products_have_summary(db, co, store_id=store_obj.id, import_year=import_year, import_month=import_month)
        db.flush()

        # ===== 第二轮：广告/退货/入库/仓储（更新已有的 summary）=====
        for sheet_name, (stype, header, rows) in sheets.items():
            if stype in ("product_info", "transaction"):
                continue
            cc = sheet_countries.get(sheet_name)
            co = country_objs.get(cc) if cc else None
            if not co and stype in _multi_country_types:
                co = None
            if not co and stype not in _multi_country_types:
                for _sn, _c in sheet_countries.items():
                    if _c and _c in country_objs:
                        co = country_objs[_c]
                        break
                if not co:
                    _top = db.query(RawTransaction.country_id, func.count(RawTransaction.id).label('cnt'))\
                        .filter(RawTransaction.store_id == store_obj.id)\
                        .group_by(RawTransaction.country_id).order_by(func.count(RawTransaction.id).desc()).first()
                    if _top:
                        _top_co = db.query(DimCountry).filter(DimCountry.id == _top[0]).first()
                        if _top_co:
                            co = _top_co
                if not co:
                    results[sheet_name] = {"status": "error", "type": stype, "detail": f"国家 {cc} 不存在，无法自动检测"}
                    continue
            try:
                if stype == "advertising":
                    result = _process_advertising_sheet(db, co, header, rows, store_id=store_obj.id, import_year=import_year, import_month=import_month)
                elif stype == "returns":
                    result = _process_fee_sheet(db, co, header, rows, "returns", store_id=store_obj.id, import_year=import_year, import_month=import_month)
                elif stype == "inbound":
                    result = _process_fee_sheet(db, co, header, rows, "inbound", store_id=store_obj.id, import_year=import_year, import_month=import_month)
                elif stype == "storage":
                    result = _process_fee_sheet(db, co, header, rows, "storage", store_id=store_obj.id, import_year=import_year, import_month=import_month)
                elif stype == "long_term_storage":
                    result = _process_fee_sheet(db, co, header, rows, "long_term_storage", store_id=store_obj.id, import_year=import_year, import_month=import_month)
                elif stype == "removal_fee":
                    result = _process_removal_fee_sheet(db, co, header, rows, store_id=store_obj.id, import_year=import_year, import_month=import_month)
                results[sheet_name] = {"status": "success", "type": stype, "country": cc or "multi", **result}
            except Exception as e:
                results[sheet_name] = {"status": "error", "type": stype, "country": cc or "multi", "detail": str(e)}

        # ===== 重算利润——只重算当前店铺的已导入月份 =====
        _time_obj_for_rec = None
        if import_year and import_month:
            _time_obj_for_rec = db.query(DimTime).filter(DimTime.time_year == import_year, DimTime.time_month == import_month).first()
        for cc, co in country_objs.items():
            _recalculate_all_profit(db, co, store_id=store_obj.id, time_id=_time_obj_for_rec.id if _time_obj_for_rec else None)
        db.commit()

        # ===== 导入后验证：按国家汇总数据 =====
        country_summary = {}
        for cc, co in country_objs.items():
            stats = db.query(
                func.sum(MonthlySummary.order_count),
                func.sum(MonthlySummary.order_qty),
                func.sum(MonthlySummary.product_sales_usd),
                func.sum(MonthlySummary.ad_spend_usd),
                func.sum(MonthlySummary.storage_fee_usd),
                func.sum(MonthlySummary.net_profit_rmb),
            ).filter(MonthlySummary.country_id == co.id).first()
            raw_orders = db.query(func.count()).filter(
                RawTransaction.country_id == co.id,
                RawTransaction.transaction_type.in_(["Order"]),
            ).scalar()
            raw_refunds = db.query(func.count()).filter(
                RawTransaction.country_id == co.id,
                RawTransaction.transaction_type.in_(["Refund"]),
            ).scalar()
            country_summary[cc] = {
                "order_count": int(stats[0] or 0),
                "order_qty": int(stats[1] or 0),
                "sales_usd": round(float(stats[2] or 0), 2),
                "ad_spend_usd": round(float(stats[3] or 0), 2),
                "storage_fee_usd": round(float(stats[4] or 0), 2),
                "net_profit_rmb": round(float(stats[5] or 0), 2),
                "raw_orders": raw_orders,
                "raw_refunds": raw_refunds,
            }

        return {"message": "工作簿导入完成", "countries": list(set(sheet_countries.values()) - {None}), "country_summary": country_summary, "sheets": results}

    except Exception as e:
        db.rollback()
        return {"detail": str(e)}


@router.post("/folder")
async def import_folder(
    files: list[UploadFile] = File(..., description="选择文件夹中的所有文件"),
    import_year: int = Form(2026, description="导入年份"),
    db: Session = Depends(get_db),
):
    """文件夹批量导入：自动从文件名解析店铺+月份，如 LMG-EU_04.xlsx → 店铺LMG-EU，月份4"""
    try:
        import openpyxl
        import re as _re

        # 加载所有店铺
        all_stores = db.query(DimStore).all()
        store_map = {}  # code -> DimStore
        for s in all_stores:
            store_map[s.code.upper()] = s

        # 解析文件名：{store_code}_{month}.xlsx
        # webkitdirectory 上传的文件名带文件夹前缀，如 "04/LMG-EU_04.xlsx"，需提取纯文件名
        file_plan = []  # [(upload_file, store_obj, month_int, filename)]
        skipped = []
        for f in files:
            fname = f.filename.split('/')[-1].split('\\')[-1]  # 取纯文件名
            if not (fname.endswith('.xlsx') or fname.endswith('.csv')):
                skipped.append({"file": fname, "reason": "非xlsx/csv格式"})
                continue
            # 匹配: LMG-EU_04.xlsx → store=LMG-EU, month=04
            name_no_ext = fname.rsplit('.', 1)[0]
            m = _re.match(r'^(.+?)_(\d{1,2})$', name_no_ext)
            if not m:
                skipped.append({"file": fname, "reason": "文件名不符合 {店铺}_{月份} 格式"})
                continue
            store_code = m.group(1).strip()
            month_int = int(m.group(2))
            store_obj = store_map.get(store_code.upper())
            if not store_obj:
                # 尝试模糊匹配
                for sc, so in store_map.items():
                    if sc in store_code.upper() or store_code.upper() in sc:
                        store_obj = so
                        break
            if not store_obj:
                skipped.append({"file": fname, "reason": f"店铺 '{store_code}' 不存在"})
                continue
            if not (1 <= month_int <= 12):
                skipped.append({"file": fname, "reason": f"月份 {month_int} 无效"})
                continue
            file_plan.append((f, store_obj, month_int, fname))

        if not file_plan:
            return {"detail": "没有可导入的文件", "skipped": skipped}

        results = {}
        _txn_date_headers = ('date/time', 'date / time', 'fecha/hora', 'fecha y hora', 'datum/uhrzeit',
                             'date/heure', 'data/ora', 'data/ora:', 'datum/tijd', 'datum/tid')

        # 按店铺+月份分组，每组清空一次数据
        from sqlalchemy import extract, text as _text, func as _func
        processed_keys = set()  # (store_id, year, month)

        for upload_file, store_obj, month_int, fname in file_plan:
            import_month = month_int
            import_year_val = import_year
            key = (store_obj.id, import_year_val, import_month)

            # 首次遇到该店铺+月份时清空旧数据（用 time_id 匹配）
            if key not in processed_keys:
                processed_keys.add(key)
                _clear_filters = [RawTransaction.store_id == store_obj.id]
                _clear_time = db.query(DimTime).filter(DimTime.time_year == import_year_val, DimTime.time_month == import_month).first()
                if _clear_time:
                    _clear_filters.append(RawTransaction.time_id == _clear_time.id)
                db.query(RawTransaction).filter(*_clear_filters).delete()
                # RawAdvertising：按店铺+月份清空
                if _clear_time:
                    db.query(RawAdvertising).filter(RawAdvertising.store_id == store_obj.id, RawAdvertising.time_id == _clear_time.id).delete()
                else:
                    db.query(RawAdvertising).filter(RawAdvertising.store_id == store_obj.id).delete()
                # 仓储/退货/入库/长期仓储费：按店铺+月份清空
                _fee_month_prefix = f"{import_year_val}-{import_month:02d}"
                db.query(RawStorageFee).filter(RawStorageFee.store_id == store_obj.id, RawStorageFee.month_of_charge.like(f"{_fee_month_prefix}%")).delete()
                db.query(RawReturns).filter(RawReturns.store_id == store_obj.id, RawReturns.month_of_charge.like(f"{_fee_month_prefix}%")).delete()
                db.query(RawInbound).filter(RawInbound.store_id == store_obj.id, _func.date_format(RawInbound.transaction_date, '%Y-%m') == _fee_month_prefix).delete()
                db.query(RawLongTermStorage).filter(RawLongTermStorage.store_id == store_obj.id).delete()

                _time_obj = db.query(DimTime).filter(DimTime.time_year == import_year_val, DimTime.time_month == import_month).first()
                if _time_obj:
                    db.query(MonthlySummary).filter(MonthlySummary.store_id == store_obj.id, MonthlySummary.time_id == _time_obj.id).delete()
                    db.query(DimExchangeRate).filter(DimExchangeRate.store_id == store_obj.id, DimExchangeRate.year_month == f"{import_year_val}-{import_month:02d}").delete()
                db.flush()

            try:
                content = await upload_file.read()

                if fname.endswith('.xlsx'):
                    wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True)
                    for sheet_name in wb.sheetnames:
                        ws = wb[sheet_name]
                        all_rows = list(ws.iter_rows(values_only=True))
                        if not all_rows or len(all_rows) < 2:
                            continue
                        header = [str(h).strip().lower() if h else "" for h in all_rows[0]]
                        stype = None
                        if any(h in _txn_date_headers for h in header):
                            stype = "transaction"
                            data_rows = all_rows[1:]
                        else:
                            for i in range(1, min(15, len(all_rows))):
                                row_cells = [str(v).strip().lower() if v else "" for v in all_rows[i]]
                                if any(c in _txn_date_headers for c in row_cells) and 'sku' in row_cells:
                                    stype = "transaction"
                                    header = row_cells
                                    data_rows = all_rows[i+1:]
                                    break
                            if not stype:
                                data_rows = all_rows[1:]
                                if "asin" in set(h for h in header if h) and any("成本" in h or "cost" in h for h in header):
                                    stype = "product_info"
                                elif any(kw in h for kw in ("商品", "campaign", "asin", "product", "advertised") for h in header) \
                                        and any(kw in h for kw in ("花费", "支出", "spend", "cost", "acos", "roas") for h in header):
                                    stype = "advertising"
                                elif any("returns_fee" in h or "returned_units" in h for h in header):
                                    stype = "returns"
                                elif any("入库" in h or "inbound" in h for h in header):
                                    stype = "inbound"
                                elif any("estimated_monthly_storage_fee" in h or "月度仓储费" in h or "亚马逊运营中心" in h for h in header):
                                    stype = "storage"
                                elif any("amount-charged" in h or "amount_charged" in h for h in header):
                                    stype = "storage" if not any("surcharge" in h or "snapshot" in h for h in header) else "long_term_storage"
                                elif "汇率" in header and ("国家" in header or "country" in header):
                                    stype = "exchange_rate"
                                elif "request-date" in header and "removal-fee" in header:
                                    stype = "removal_fee"

                        if not stype:
                            results[f"{fname}/{sheet_name}"] = {"status": "skipped", "reason": "无法识别"}
                            continue

                        # 自动检测国家
                        country_obj = None
                        detected = _detect_country_from_data(db, header, data_rows)
                        if detected:
                            country_obj = db.query(DimCountry).filter(DimCountry.code == detected).first()
                        if not country_obj:
                            # 从店铺名推断
                            sc = store_obj.code.upper()
                            if 'NA' in sc:
                                country_obj = db.query(DimCountry).filter(DimCountry.code == 'US').first()
                            elif 'EU' in sc:
                                country_obj = db.query(DimCountry).filter(DimCountry.code == 'UK').first()
                            else:
                                country_obj = db.query(DimCountry).filter(DimCountry.code == 'US').first()

                        key_name = f"{fname}/{sheet_name}"
                        if stype == "transaction":
                            r = _process_transaction_sheet(db, country_obj, header, data_rows, store_id=store_obj.id, import_year=import_year_val, import_month=import_month)
                        elif stype == "product_info":
                            r = _process_product_info_sheet(db, header, data_rows, import_year=import_year_val, import_month=import_month, store_id=store_obj.id)
                        elif stype == "advertising":
                            r = _process_advertising_sheet(db, country_obj, header, data_rows, store_id=store_obj.id, import_year=import_year_val, import_month=import_month)
                        elif stype == "exchange_rate":
                            ci = next((i for i, h in enumerate(header) if h in ("国家", "country")), None)
                            ri = next((i for i, h in enumerate(header) if h == "汇率"), None)
                            imported = 0
                            if ci is not None and ri is not None:
                                ym = f"{import_year_val}-{import_month:02d}"
                                for row in data_rows:
                                    cc = str(row[ci] or "").strip().upper()
                                    rate = row[ri]
                                    if not cc or rate is None:
                                        continue
                                    co = db.query(DimCountry).filter(DimCountry.code == cc).first()
                                    if co:
                                        existing = db.query(DimExchangeRate).filter(DimExchangeRate.country_id == co.id, DimExchangeRate.year_month == ym, DimExchangeRate.store_id == store_obj.id).first()
                                        if existing:
                                            existing.rate = Decimal(str(rate))
                                        else:
                                            db.add(DimExchangeRate(country_id=co.id, year_month=ym, rate=Decimal(str(rate)), store_id=store_obj.id))
                                        imported += 1
                            r = {"imported": imported}
                        elif stype in ("storage", "returns", "inbound", "long_term_storage"):
                            r = _process_fee_sheet(db, country_obj, header, data_rows, stype, store_id=store_obj.id, import_year=import_year_val, import_month=import_month)
                        elif stype == "removal_fee":
                            r = _process_removal_fee_sheet(db, country_obj, header, data_rows, store_id=store_obj.id, import_year=import_year_val, import_month=import_month)
                        else:
                            r = {}
                        results[key_name] = {"status": "success", "type": stype, "store": store_obj.code, "month": import_month, **r}
                    wb.close()

            except Exception as e:
                results[fname] = {"status": "error", "detail": str(e)}

        # 补建所有国家+店铺的产品 summary（防止费用数据无行可写）
        for (store_id, yr, mo) in processed_keys:
            for co in db.query(DimCountry).all():
                _ensure_all_products_have_summary(db, co, store_id=store_id, import_year=yr, import_month=mo)

        # 重算利润——只算本次导入的店铺+月份
        for (store_id, yr, mo) in processed_keys:
            _time_obj = db.query(DimTime).filter(DimTime.time_year == yr, DimTime.time_month == mo).first()
            if _time_obj:
                for co in db.query(DimCountry).all():
                    _recalculate_all_profit(db, co, store_id=store_id, time_id=_time_obj.id)
        db.commit()

        # 汇总结果
        summary_by_store = {}
        for (store_id, yr, mo) in processed_keys:
            so = db.query(DimStore).filter(DimStore.id == store_id).first()
            if not so:
                continue
            stats = db.query(
                func.sum(MonthlySummary.order_count), func.sum(MonthlySummary.order_qty),
                func.sum(MonthlySummary.product_sales_usd), func.sum(MonthlySummary.net_profit_rmb),
            ).filter(MonthlySummary.store_id == store_id).first()
            summary_by_store[so.code] = {
                "order_count": int(stats[0] or 0), "order_qty": int(stats[1] or 0),
                "sales_usd": round(float(stats[2] or 0), 2),
                "net_profit_rmb": round(float(stats[3] or 0), 2),
            }

        return {
            "message": f"文件夹导入完成，处理 {len(file_plan)} 个文件",
            "files_processed": len(file_plan),
            "skipped": skipped,
            "summary_by_store": summary_by_store,
            "sheets": results,
        }
    except Exception as e:
        db.rollback()
        return {"detail": str(e)}


def _find_col(header, *names):
    """在表头中查找列索引（精确匹配优先，大小写不敏感，避免 "产品" 误匹配 "产品运费/台"）"""
    # 第一轮：精确匹配（大小写不敏感）
    for i, h in enumerate(header):
        hl = h.lower() if h else ""
        for name in names:
            if hl == name.lower():
                return i
    # 第二轮：模糊匹配（大小写不敏感）
    for i, h in enumerate(header):
        hl = h.lower() if h else ""
        for name in names:
            if name.lower() in hl:
                return i
    return None


def _parse_eu_number(value):
    """解析欧洲格式数字：30,24→30.24 / 1 135,72→1135.72 / 1,777.43→1777.43 / 30.24→30.24"""
    if value is None:
        return Decimal("0")
    s = str(value).strip()
    if not s or s == '-' or s == '−':
        return Decimal("0")
    # Normalize Unicode minus (U+2212) to regular hyphen-minus
    s = s.replace('−', '-')
    # Remove currency symbols
    s = s.replace('€', '').replace('$', '').replace('£', '').replace('%', '').strip()
    # Detect format: if has comma followed by 1-2 digits at end → European decimal
    import re
    if re.search(r',\d{1,2}$', s):
        # European format: "30,24" or "1 135,72" or "-28,22"
        s = s.replace(' ', '').replace('.', '').replace(',', '.')
    elif ',' in s and re.search(r',\d{3}', s):
        # US thousands: "1,777.43"
        s = s.replace(',', '')
    else:
        # Already dot format or no separator
        s = s.replace(',', '')
    try:
        return Decimal(s)
    except:
        return Decimal("0")


def _process_transaction_sheet(db, country_obj, header, rows, store_id=None, import_year=None, import_month=None):
    """处理交易记录 sheet，import_year/month 覆盖文件中时间"""
    # Normalize multi-language headers to English
    _header_normalize = {
        'datum/uhrzeit': 'date/time', 'date/heure': 'date/time', 'data/ora': 'date/time',
        'data/ora:': 'date/time', 'datum/tijd': 'date/time', 'datum/tid': 'date/time',
        'fecha/hora': 'date/time', 'fecha y hora': 'date/time',
        'typ': 'type', 'tipo': 'type',
        'abrechnungsnummer': 'settlement id', 'numéro de versement': 'settlement id',
        'identifiant du paiement': 'settlement id',
        'numero pagamento': 'settlement id', 'schikkings-id': 'settlement id', 'reglerings-id': 'settlement id',
        'id. de liquidación': 'settlement id', 'identificador de pago': 'settlement id',
        'bestellnummer': 'order id', 'numéro de la commande': 'order id', 'numero ordine': 'order id',
        'bestelnummer': 'order id', 'beställnings-id': 'order id', 'id. del pedido': 'order id',
        'número de pedido': 'order id',
        'beschreibung': 'description', 'descripción': 'description', 'descrizione': 'description',
        'beschrijving': 'description', 'beskrivning': 'description',
        'menge': 'quantity', 'cantidad': 'quantity', 'quantité': 'quantity', 'quantità': 'quantity', 'aantal': 'quantity', 'antal': 'quantity',
        'versand': 'fulfillment', 'cumplimiento': 'fulfillment', 'traitement': 'fulfillment',
        'expédition': 'fulfillment', 'gestión logística': 'fulfillment',
        'gestione': 'fulfillment', 'leverans': 'fulfillment',
        'ort der bestellung': 'order city', "ville d'où provient la commande": 'order city',
        "città di provenienza dell'ordine": 'order city', 'bestelling stad': 'order city',
        'stad för beställning': 'order city', 'ciudad del pedido': 'order city',
        'ville de la commande': 'order city', 'ciudad de procedencia del pedido': 'order city',
        'marknadsplats': 'marketplace', 'web de amazon': 'marketplace', 'site de vente': 'marketplace',
        'formulario de recaudación de impuestos': 'tax collection model',
        'bundesland': 'order state', "région d'où provient la commande": 'order state',
        'état de la commande': 'order state', 'comunidad autónoma de procedencia del pedido': 'order state',
        "provincia di provenienza dell'ordine": 'order state', 'status bestelling': 'order state',
        'delstat för beställning': 'order state', 'estado del pedido': 'order state',
        'postleitzahl': 'order postal', 'code postal de la commande': 'order postal',
        'commande postale': 'order postal', 'código postal de procedencia del pedido': 'order postal',
        "cap dell'ordine": 'order postal', 'bestelling per post': 'order postal',
        'postadress för beställning': 'order postal', 'código postal del pedido': 'order postal',
        'umsätze': 'product sales', 'ventes de produits': 'product sales', 'ventas de productos': 'product sales',
        'vendite': 'product sales', 'verkoop van producten': 'product sales',
        'försäljning av produkter': 'product sales',
        'produktumsatzsteuer': 'product sales tax', 'impuesto de ventas de productos': 'product sales tax',
        'taxes sur la vente des produits': 'product sales tax', 'imposta sulle vendite dei prodotti': 'product sales tax',
        'gutschrift für versandkosten': 'postage credits', "crédits d'expédition": 'postage credits',
        "crédits d'expédition ": 'postage credits',
        'abonos de envío': 'postage credits', 'accrediti per le spedizioni': 'postage credits',
        'verzendtegoeden': 'postage credits', 'fraktkrediter': 'postage credits',
        'steuer auf versandgutschrift': 'shipping credits tax', 'impuesto de abono de envío': 'shipping credits tax',
        'impuestos por abonos de envío': 'shipping credits tax',
        'imposta accrediti per le spedizioni': 'shipping credits tax',
        'gutschrift für geschenkverpackung': 'gift wrap credits', "crédits sur l'emballage cadeau": 'gift wrap credits',
        "crédits d'emballage-cadeau": 'gift wrap credits',
        'créditos por envoltorio de regalo': 'gift wrap credits', 'abonos de envoltorio para regalo': 'gift wrap credits',
        'accrediti per confezioni regalo': 'gift wrap credits',
        'kredietpunten cadeauverpakking': 'gift wrap credits', 'krediter för presentinslagning': 'gift wrap credits',
        'steuer auf geschenkverpackungsgutschriften': 'giftwrap credits tax',
        'taxes sur les crédits cadeaux': 'giftwrap credits tax', 'imposta sui crediti confezione regalo': 'giftwrap credits tax',
        'impuesto de créditos de envoltura': 'giftwrap credits tax',
        'impuestos por abonos de envoltorio para regalo': 'giftwrap credits tax',
        'rabatte aus werbeaktionen': 'promotional rebates', 'rabais promotionnels': 'promotional rebates',
        'total des réductions': 'promotional rebates',
        'descuentos promocionales': 'promotional rebates', 'devoluciones promocionales': 'promotional rebates',
        'sconti promozionali': 'promotional rebates',
        'promotiekortingen': 'promotional rebates', 'kampanjrabatter': 'promotional rebates',
        'steuer auf aktionsrabatte': 'promotional rebates tax', 'taxes sur les remises promotionnelles': 'promotional rebates tax',
        'impuesto de reembolsos promocionales': 'promotional rebates tax',
        'impuestos de descuentos por promociones': 'promotional rebates tax',
        'imposta sugli sconti promozionali': 'promotional rebates tax',
        'einbehaltene steuer auf marketplace': 'marketplace withheld tax',
        'taxes retenues sur le site de vente': 'marketplace withheld tax',
        'impuesto de retenciones en la plataforma': 'marketplace withheld tax',
        'impuesto retenido en el sitio web': 'marketplace withheld tax',
        'trattenuta iva del marketplace': 'marketplace withheld tax',
        'verkaufsgebühren': 'selling fees', 'frais de vente': 'selling fees', 'tarifas de venta': 'selling fees',
        'commissioni di vendita': 'selling fees', 'verkoopkosten': 'selling fees', 'försäljningsavgifter': 'selling fees',
        'gebühren zu versand durch amazon': 'fba fees', 'frais expédié par amazon': 'fba fees',
        'frais pour le service expédié par amazon': 'fba fees',
        'tarifas fba': 'fba fees', 'tarifas de logística de amazon': 'fba fees',
        'costi del servizio logistica di amazon': 'fba fees',
        'fba-vergoedingen': 'fba fees', 'fba-avgifter': 'fba fees',
        'fulfilment by amazon fees': 'fba fees', 'fulfillment by amazon fees': 'fba fees',
        'andere transaktionsgebühren': 'other transaction fees', 'autres frais de transaction': 'other transaction fees',
        'tarifas de otra transacción': 'other transaction fees', 'tarifas de otras transacciones': 'other transaction fees',
        'altri costi relativi alle transazioni': 'other transaction fees',
        'overige transactiekosten': 'other transaction fees', 'övriga transaktionsavgifter': 'other transaction fees',
        'andere': 'other', 'autre': 'other', 'autres': 'other', 'otro': 'other', 'altro': 'other', 'overige': 'other', 'övrigt': 'other',
        'gesamt': 'total', 'totaal': 'total', 'totalt': 'total',
        'transaktionsstatus': 'transaction status', 'statut de la transaction': 'transaction status',
        'estado de la transacción': 'transaction status', 'stato della transazione': 'transaction status',
        'transactiestatus': 'transaction status', 'transaktionens utgivningsdatum': 'transaction release date',
        'freigabedatum der transaktion': 'transaction release date', 'date de sortie de la transaction': 'transaction release date',
        'date de délivrance de la transaction': 'transaction release date',
        'fecha de liberación de la transacción': 'transaction release date',
        'data di rilascio della transazione': 'transaction release date',
        'publicatiedatum van transactie': 'transaction release date',
    }
    header = [_header_normalize.get(h, h) for h in header]
    col_date = _find_col(header, "date/time", "date / time")
    col_type = _find_col(header, "type")
    col_order = _find_col(header, "order id")
    col_sku = _find_col(header, "sku")
    col_desc = _find_col(header, "description")
    col_qty = _find_col(header, "quantity")
    col_ps = _find_col(header, "product sales")
    col_postage = _find_col(header, "postage credits")
    col_sf = _find_col(header, "selling fees")
    col_fba = _find_col(header, "fba fees")
    col_total = _find_col(header, "total")
    col_marketplace = _find_col(header, "marketplace")
    col_fulfillment = _find_col(header, "fulfillment")

    if col_date is None or col_ps is None:
        return {"raw_rows": 0, "summary_rows": 0, "error": "缺少必要列"}

    sku_aggregation = {}
    adj_aggregation3 = {}  # Adjustment aggregation for Path 3
    raw_count = 0
    exchange_rate = _get_exchange_rate(db, country_obj, import_year, import_month, store_id=store_id)

    # 查找额外列的索引
    col_settlement = _find_col(header, "settlement id")
    col_city = _find_col(header, "order city")
    col_state = _find_col(header, "order state")
    col_postal = _find_col(header, "order postal")
    col_tax_model = _find_col(header, "tax collection model")
    col_ps_tax = _find_col(header, "product sales tax")
    col_postage = _find_col(header, "postage credits", "shipping credits")  # UK=postcode, 其他=shipping
    col_ship_credit_tax = _find_col(header, "shipping credits tax")
    col_gift = _find_col(header, "gift wrap credits")
    col_gift_tax = _find_col(header, "giftwrap credits tax")
    col_reg_fee = _find_col(header, "regulatory fee")
    col_reg_tax = _find_col(header, "tax on regulatory fee")
    col_promo = _find_col(header, "promotional rebates")
    col_promo_tax = _find_col(header, "promotional rebates tax")
    col_mkt_tax = _find_col(header, "marketplace withheld tax")
    col_other_fee = _find_col(header, "other transaction fees")
    col_other = _find_col(header, "other")
    col_status = _find_col(header, "status", "transaction status")
    col_release = _find_col(header, "release date", "transaction release date")

    for row in rows:
        if not row or len(row) <= col_date:
            continue

        date_val = row[col_date]
        if not date_val:
            continue

        # 解析日期
        if isinstance(date_val, datetime):
            txn_date = date_val
        else:
            txn_date = _detect_date_format(str(date_val))
        if not txn_date:
            continue

        txn_type = str(row[col_type]).strip() if col_type is not None and row[col_type] else ""
        # 多语言类型映射
        _type_map = {
            'pedido': 'Order', 'reembolso': 'Refund',
            'order': 'Order', 'refund': 'Refund',
            'ajuste': 'Adjustment', 'adjustment': 'Adjustment',
            'bestellung': 'Order', 'erstattung': 'Refund', 'anpassung': 'Adjustment',
            'commande': 'Order', 'remboursement': 'Refund', 'ajustement': 'Adjustment',
            'ordine': 'Order', 'rimborso': 'Refund', 'aggiustamento': 'Adjustment',
            'bestelling': 'Order', 'terugbetaling': 'Refund', 'aanpassing': 'Adjustment',
            'beställning': 'Order', 'återbetalning': 'Refund', 'justering': 'Adjustment',
        }
        txn_type = _type_map.get(txn_type.lower(), txn_type)
        sku = str(row[col_sku]).strip() if col_sku is not None and row[col_sku] else ""
        asin = sku.split("-")[0] if sku and "-" in sku else sku

        # 识别 amzn.gr 替换件，提取真实SKU
        is_replacement = sku.startswith("amzn.gr.") if sku else False
        real_sku = _extract_real_sku(sku) if is_replacement else None
        effective_sku = real_sku if real_sku else sku

        product_sales = _parse_eu_number(row[col_ps] if col_ps is not None else 0)
        selling_fee = _parse_eu_number(row[col_sf] if col_sf is not None else 0)
        fba_fee = _parse_eu_number(row[col_fba] if col_fba is not None else 0)
        quantity = _safe_int(row[col_qty] if col_qty is not None else 0)
        total = _parse_eu_number(row[col_total] if col_total is not None else 0)

        # 费用回填：如果佣金和FBA费都为0但total≠product_sales，从差额推算缺失费用
        if txn_type == "Order" and product_sales > 0 and total > 0:
            other_charges = _parse_eu_number(row[col_other_fee] if col_other_fee is not None else 0) + \
                           _parse_eu_number(row[col_promo] if col_promo is not None else 0)
            implied_fees = product_sales - total + other_charges  # 总费用（正值）
            captured_fees = abs(selling_fee)  # 已捕获的佣金（取绝对值）
            if fba_fee == 0 and selling_fee == 0 and implied_fees > Decimal("0.5"):
                # 佣金和FBA都没解析到，total已扣费→按亚马逊典型佣金率15%拆分
                est_commission = (product_sales * Decimal("0.15")).quantize(Decimal("0.01"))
                est_fba = implied_fees - est_commission
                if est_fba < 0:
                    est_commission = implied_fees
                    est_fba = Decimal("0")
                selling_fee = -est_commission
                fba_fee = -est_fba
            elif fba_fee == 0 and selling_fee != 0 and implied_fees > captured_fees + Decimal("0.5"):
                # 佣金有值但FBA为0，从差额推算FBA
                fba_fee = -(implied_fees - captured_fees)

        # 所有类型都写 raw_transactions
        raw_time_id = None
        if import_year and import_month:
            raw_time_obj = _get_or_create_time(db, import_year, import_month)
            raw_time_id = raw_time_obj.id if raw_time_obj else None

        raw = RawTransaction(
            country_id=country_obj.id,
            store_id=store_id,
            time_id=raw_time_id,
            transaction_date=txn_date,
            settlement_id=_safe_str(row[col_settlement], 50) if col_settlement is not None and row[col_settlement] else "",
            transaction_type=_safe_str(txn_type, 50),
            order_id=_safe_str(row[col_order], 50) if col_order is not None and row[col_order] else "",
            sku=_safe_str(sku, 100),
            description=_safe_str(row[col_desc], 500) if col_desc is not None and row[col_desc] else "",
            quantity=quantity,
            marketplace=_safe_str(row[col_marketplace], 20) if col_marketplace is not None and row[col_marketplace] else "",
            fulfillment=_safe_str(row[col_fulfillment], 20) if col_fulfillment is not None and row[col_fulfillment] else "",
            order_city=_safe_str(row[col_city], 100) if col_city is not None and row[col_city] else "",
            order_state=_safe_str(row[col_state], 100) if col_state is not None and row[col_state] else "",
            order_postal=_safe_str(row[col_postal], 20) if col_postal is not None and row[col_postal] else "",
            tax_collection_model=_safe_str(row[col_tax_model], 50) if col_tax_model is not None and row[col_tax_model] else "",
            product_sales=product_sales,
            product_sales_tax=_parse_eu_number(row[col_ps_tax] if col_ps_tax is not None else 0),
            postage_credits=_parse_eu_number(row[col_postage] if col_postage is not None else 0),
            shipping_credits_tax=_parse_eu_number(row[col_ship_credit_tax] if col_ship_credit_tax is not None else 0),
            gift_wrap_credits=_parse_eu_number(row[col_gift] if col_gift is not None else 0),
            giftwrap_credits_tax=_parse_eu_number(row[col_gift_tax] if col_gift_tax is not None else 0),
            regulatory_fee=_parse_eu_number(row[col_reg_fee] if col_reg_fee is not None else 0),
            tax_on_regulatory_fee=_parse_eu_number(row[col_reg_tax] if col_reg_tax is not None else 0),
            promotional_rebates=_parse_eu_number(row[col_promo] if col_promo is not None else 0),
            promotional_rebates_tax=_parse_eu_number(row[col_promo_tax] if col_promo_tax is not None else 0),
            marketplace_withheld_tax=_parse_eu_number(row[col_mkt_tax] if col_mkt_tax is not None else 0),
            selling_fee=selling_fee,
            fba_fee=fba_fee,
            other_transaction_fee=_parse_eu_number(row[col_other_fee] if col_other_fee is not None else 0),
            other_amount=_parse_eu_number(row[col_other] if col_other is not None else 0),
            total=total,
            transaction_status=_safe_str(row[col_status], 50) if col_status is not None and row[col_status] else "",
            transaction_release_date=_detect_date_format(str(row[col_release])) if col_release is not None and row[col_release] else None,
        )
        db.add(raw)
        raw_count += 1

        # Adjustment 处理（按有/无order_id分别处理）
        if txn_type == "Adjustment":
            year = import_year if import_year else txn_date.year
            month = import_month if import_month else txn_date.month
            adj_key = (effective_sku, year, month)
            if adj_key not in adj_aggregation3:
                adj_aggregation3[adj_key] = {
                    "other_with_order": Decimal("0"),
                    "total_no_order_pos": Decimal("0"),
                    "total_no_order_neg": Decimal("0"),
                    "qty": 0,
                }
            order_id_val = row[col_order] if col_order is not None else ""
            other_amount = _parse_eu_number(row[col_other] if col_other is not None else 0)
            if order_id_val and str(order_id_val).strip():
                adj_aggregation3[adj_key]["other_with_order"] += other_amount
            else:
                if total > 0:
                    adj_aggregation3[adj_key]["total_no_order_pos"] += total
                    adj_aggregation3[adj_key]["qty"] += quantity
                elif total < 0:
                    adj_aggregation3[adj_key]["total_no_order_neg"] += total
                    adj_aggregation3[adj_key]["qty"] -= abs(quantity)
            continue

        # 仅 Order/Refund 参与 monthly_summary 聚合
        if txn_type not in ("Order", "Refund"):
            continue

        year = import_year if import_year else txn_date.year
        month = import_month if import_month else txn_date.month
        key = (effective_sku, year, month)
        if key not in sku_aggregation:
            sku_aggregation[key] = {
                "product_sales": Decimal("0"), "product_sales_tax": Decimal("0"),
                "postage_credits": Decimal("0"), "shipping_credits": Decimal("0"),
                "shipping_credits_tax": Decimal("0"),
                "gift_wrap_credits": Decimal("0"), "giftwrap_credits_tax": Decimal("0"),
                "selling_fee": Decimal("0"), "fba_fee": Decimal("0"),
                "quantity": 0, "order_qty": 0,
                "promo_rebate": Decimal("0"), "promo_rebate_tax": Decimal("0"),
                "marketplace_withheld_tax": Decimal("0"),
            }
        agg = sku_aggregation[key]
        agg["product_sales"] += product_sales
        agg["product_sales_tax"] += _parse_eu_number(row[col_ps_tax] if col_ps_tax is not None else 0)
        agg["postage_credits"] += _parse_eu_number(row[col_postage] if col_postage is not None else 0)
        agg["postage_credits"] += _parse_eu_number(row[col_postage] if col_postage is not None else 0)
        agg["shipping_credits_tax"] += _parse_eu_number(row[col_ship_credit_tax] if col_ship_credit_tax is not None else 0)
        agg["gift_wrap_credits"] += _parse_eu_number(row[col_gift] if col_gift is not None else 0)
        agg["giftwrap_credits_tax"] += _parse_eu_number(row[col_gift_tax] if col_gift_tax is not None else 0)
        agg["promo_rebate"] += _parse_eu_number(row[col_promo] if col_promo is not None else 0)
        agg["promo_rebate_tax"] += _parse_eu_number(row[col_promo_tax] if col_promo_tax is not None else 0)
        agg["marketplace_withheld_tax"] += _parse_eu_number(row[col_mkt_tax] if col_mkt_tax is not None else 0)
        agg["selling_fee"] += selling_fee
        agg["fba_fee"] += fba_fee
        # Refund 数量为负（净销量）
        if txn_type == "Refund":
            agg["quantity"] -= abs(quantity)
        else:
            agg["quantity"] += abs(quantity)
            # amzn.gr 替换件不计入 order_qty（不产生采购成本）
            if not is_replacement:
                agg["order_qty"] += abs(quantity)

    # 写 monthly_summary（新利润算法）
    summary_count = 0
    for (sku, year, month), agg in sku_aggregation.items():
        product = _find_product_by_sku(db, sku, store_id=store_id, year_month=f"{year}-{month:02d}")
        if not product:
            asin = sku.split("-")[0] if sku and "-" in sku else sku
            product = _get_or_create_product(db, asin, sku, store_id=store_id)
        if not product:
            continue
        time_obj = _get_or_create_time(db, year, month)
        summary = _get_or_create_monthly_summary(db, country_obj.id, product.id, time_obj.id, store_id=store_id)

        summary.product_sales_usd = agg["product_sales"]
        summary.commission_usd = agg["selling_fee"]
        summary.fba_fee_usd = agg["fba_fee"]
        summary.promo_rebate_usd = agg["promo_rebate"]
        summary.promo_rebate_tax_usd = agg["promo_rebate_tax"]
        summary.marketplace_withheld_tax_usd = agg["marketplace_withheld_tax"]
        summary.exchange_rate = exchange_rate
        summary.product_sales_rmb = (agg["product_sales"] * exchange_rate).quantize(Decimal("0.01"))
        summary.order_count = agg["quantity"]
        summary.order_qty = agg["order_qty"]

        # 成本和运费
        time_obj = db.query(DimTime).filter(DimTime.id == summary.time_id).first()
        ym_str = time_obj.year_month if time_obj else None
        pc = db.query(DimProductCost).filter(DimProductCost.product_id == product.id, DimProductCost.year_month == ym_str).first()
        if not pc:
            pc = db.query(DimProductCost).filter(DimProductCost.product_id == product.id).first()
        unit_cost = Decimal(str(pc.cost_rmb if pc else 0))
        unit_freight = Decimal(str(pc.freight_per_unit if pc else 0))
        if store_id:
            df = db.query(DimFreight).filter(DimFreight.product_id == product.id, DimFreight.country_id == country_obj.id, DimFreight.store_id == store_id).first()
        else:
            df = db.query(DimFreight).filter(DimFreight.product_id == product.id, DimFreight.country_id == country_obj.id).first()
        if df:
            unit_freight = Decimal(str(df.freight_rmb))
        cost_rmb = (unit_cost * agg["order_qty"]).quantize(Decimal("0.01"))
        freight_rmb = (unit_freight * agg["order_qty"]).quantize(Decimal("0.01"))
        summary.product_cost_rmb = cost_rmb
        summary.freight_cost_rmb = freight_rmb

        # 写入新字段
        summary.product_sales_tax = agg["product_sales_tax"]
        summary.postage_credits = agg["postage_credits"]
        summary.shipping_credits = agg["shipping_credits"]
        summary.shipping_credits_tax = agg["shipping_credits_tax"]
        summary.gift_wrap_credits = agg["gift_wrap_credits"]
        summary.giftwrap_credits_tax = agg["giftwrap_credits_tax"]

        # 统一利润公式
        from services.profit import apply_profit_to_summary
        apply_profit_to_summary(summary, exchange_rate,
            raw_product_sales=agg["product_sales"],
            raw_product_sales_tax=agg["product_sales_tax"],
            raw_postage_credits=agg["postage_credits"],
            raw_shipping_credits_tax=agg["shipping_credits_tax"],
            raw_gift_wrap_credits=agg["gift_wrap_credits"],
            raw_giftwrap_credits_tax=agg["giftwrap_credits_tax"],
            raw_promo_rebate=agg["promo_rebate"],
            raw_promo_rebate_tax=agg["promo_rebate_tax"],
            raw_marketplace_withheld_tax=agg["marketplace_withheld_tax"],
            raw_selling_fee=agg["selling_fee"],
            raw_fba_fee=agg["fba_fee"],
        )
        summary_count += 1

    # 处理 Adjustment
    for (adj_sku, adj_year, adj_month), adj_agg in adj_aggregation3.items():
        product = _find_product_by_sku(db, adj_sku, store_id=store_id, year_month=f"{adj_year}-{adj_month:02d}")
        if not product:
            asin = adj_sku.split("-")[0] if adj_sku and "-" in adj_sku else adj_sku
            product = _get_or_create_product(db, asin, adj_sku, store_id=store_id)
        if not product:
            continue
        time_obj = _get_or_create_time(db, adj_year, adj_month)
        summary = _get_or_create_monthly_summary(db, country_obj.id, product.id, time_obj.id, store_id=store_id)

        adj_total = adj_agg["other_with_order"] + adj_agg["total_no_order_pos"] + adj_agg["total_no_order_neg"]
        summary.adjustment_usd = (summary.adjustment_usd or Decimal("0")) + adj_total

        if adj_agg["qty"] != 0:
            summary.order_qty = (summary.order_qty or 0) + adj_agg["qty"]
            pc = db.query(DimProductCost).filter(DimProductCost.product_id == product.id, DimProductCost.year_month == f"{adj_year}-{adj_month:02d}").first()
            if not pc:
                pc = db.query(DimProductCost).filter(DimProductCost.product_id == product.id).first()
            if pc:
                unit_cost = Decimal(str(pc.cost_rmb or 0))
                unit_freight = Decimal(str(pc.freight_per_unit or 0))
                if store_id:
                    df = db.query(DimFreight).filter(DimFreight.product_id == product.id, DimFreight.country_id == country_obj.id, DimFreight.store_id == store_id).first()
                else:
                    df = db.query(DimFreight).filter(DimFreight.product_id == product.id, DimFreight.country_id == country_obj.id).first()
                if df:
                    unit_freight = Decimal(str(df.freight_rmb))
                summary.product_cost_rmb = (unit_cost * summary.order_qty).quantize(Decimal("0.01"))
                summary.freight_cost_rmb = (unit_freight * summary.order_qty).quantize(Decimal("0.01"))

        # 统一利润公式
        adj_total = adj_agg.get("other_with_order", Decimal("0")) + adj_agg.get("total_no_order_pos", Decimal("0")) + adj_agg.get("total_no_order_neg", Decimal("0"))
        from services.profit import apply_profit_to_summary
        apply_profit_to_summary(summary, exchange_rate, raw_adjustment=adj_total)
        summary_count += 1

    return {"raw_rows": raw_count, "summary_rows": summary_count}


def _process_product_info_sheet(db, header, rows, import_year=None, import_month=None, store_id=None):
    """处理产品信息 sheet"""
    col_asin = _find_col(header, "asin")
    col_sku = _find_col(header, "sku")
    col_name = _find_col(header, "产品", "型号", "product")
    col_color = _find_col(header, "颜色", "color")
    col_cost = _find_col(header, "成本RMB", "成本", "cost")
    col_freight = _find_col(header, "产品运费/台", "运费", "其他站点运费", "英国运费", "英国站运费", "英国站点运费", "freight")
    col_time = _find_col(header, "时间", "time")

    # 检测独立国家运费列
    _country_freight_cols = {
        _find_col(header, "英国站运费", "英国站点运费", "英国运费"): "UK",
        _find_col(header, "爱尔兰站运费", "爱尔兰站点运费", "爱尔兰运费"): "IE",
    }
    _country_freight_cols = {k: v for k, v in _country_freight_cols.items() if k is not None}
    has_country_freight = len(_country_freight_cols) > 0
    country_freight_ids = {}
    if has_country_freight:
        for code in set(_country_freight_cols.values()):
            c = db.query(DimCountry).filter(DimCountry.code == code).first()
            if c:
                country_freight_ids[code] = c.id

    if col_asin is None:
        return {"rows": 0, "error": "缺少 ASIN 列"}

    count = 0
    for row in rows:
        if not row or not row[col_asin]:
            continue

        asin = str(row[col_asin]).strip()
        sku = str(row[col_sku]).strip() if col_sku is not None and row[col_sku] else None
        name = str(row[col_name]).strip() if col_name is not None and row[col_name] else None
        color = str(row[col_color]).strip() if col_color is not None and row[col_color] else None
        cost = _safe_decimal(row[col_cost]) if col_cost is not None else Decimal("0")

        # 运费：有独立国家运费列时，用通用运费作为默认；否则按优先级取值
        if has_country_freight:
            freight = Decimal("0")
            for fk in ["产品运费/台", "运费"]:
                idx = _find_col(header, fk)
                if idx is not None and row[idx]:
                    freight = _safe_decimal(row[idx])
                    break
        else:
            freight = _safe_decimal(row[col_freight]) if col_freight is not None else Decimal("0")

        # 按店铺+月份严格查找（不回退全局）
        ym_key = f"{import_year}-{import_month:02d}" if import_year and import_month else None
        product = None
        if store_id and ym_key:
            product = db.query(DimProduct).filter(
                DimProduct.asin == asin,
                DimProduct.store_id == store_id,
                DimProduct.year_month == ym_key,
            ).first()

        if not product:
            # 不存在则新建（每个店铺每个月份独立产品）
            product = DimProduct(
                asin=asin, sku=sku or asin, product_name=name or "",
                color=color, store_id=store_id, year_month=ym_key,
            )
            db.add(product)
            db.flush()
        else:
            # 存在则更新
            if sku: product.sku = sku
            if name: product.product_name = name
            if color: product.color = color
        # 解析时间写 dim_product_cost
        ym = None
        if col_time is not None and row[col_time]:
            import re as _re
            s = str(row[col_time]).strip()[:10]
            if _re.match(r'\d{4}-\d{2}', s):
                ym = s[:7]
        if not ym and import_year and import_month:
            ym = f"{import_year}-{import_month:02d}"
        if ym:
            pc = db.query(DimProductCost).filter(DimProductCost.product_id == product.id, DimProductCost.year_month == ym).first()
            if pc:
                pc.cost_rmb = cost
                pc.freight_per_unit = freight
            else:
                db.add(DimProductCost(product_id=product.id, year_month=ym, cost_rmb=cost, freight_per_unit=freight))

        # 独立国家运费写入 dim_freight（upsert）
        if has_country_freight and store_id:
            for col_idx, country_code in _country_freight_cols.items():
                if row[col_idx]:
                    cf_val = _safe_decimal(row[col_idx])
                    if cf_val > 0 and country_code in country_freight_ids:
                        from sqlalchemy import text as _text
                        db.execute(_text("""
                            INSERT INTO dim_freight (product_id, country_id, store_id, freight_rmb)
                            VALUES (:pid, :cid, :sid, :freight)
                            ON DUPLICATE KEY UPDATE freight_rmb = VALUES(freight_rmb)
                        """), {"pid": product.id, "cid": country_freight_ids[country_code], "sid": store_id, "freight": cf_val})

        count += 1

    return {"rows": count}


def _process_advertising_sheet(db, country_obj, header, rows, time_id=None, store_id=None, import_year=None, import_month=None):
    """处理广告 sheet，使用导入时选择的年月"""
    col_product = _find_col(header, "商品", "asin", "product", "campaign")
    col_spend = _find_col(header, "花费(usd)", "花费(cad)", "花费(mx)", "花费(eur)", "花费(gbp)", "花费(sek)", "花费(aud)", "花费", "支出(usd)", "支出(cad)", "支出(mx)", "支出(eur)", "支出(gbp)", "支出(sek)", "支出(aud)", "spend", "cost")
    col_sales = _find_col(header, "销售额(usd)", "销售额(cad)", "销售额(mx)", "销售额", "sales")
    col_time = _find_col(header, "time", "日期", "date")
    col_acos = _find_col(header, "acos")
    col_roas = _find_col(header, "roas")
    col_ctr = _find_col(header, "ctr")
    col_cpc = _find_col(header, "cpc")
    col_imp = _find_col(header, "展示", "impression")
    col_clicks = _find_col(header, "点击", "click")
    col_orders = _find_col(header, "订单", "order")
    col_conv = _find_col(header, "转化", "conversion")
    col_status = _find_col(header, "状态", "status")
    col_type = _find_col(header, "类型", "type")
    col_elig = _find_col(header, "资格")
    col_ntb_orders = _find_col(header, "ntb 订单数量", "ntb orders")
    col_ntb_pct = _find_col(header, "ntb 订单数量百分比", "ntb order")
    col_ntb_sales = _find_col(header, "ntb 销售额", "ntb sales")
    col_new_brand = _find_col(header, "品牌新客", "new to brand")
    col_vis_imp = _find_col(header, "可见展示", "visible imp")

    if col_product is None:
        return {"csv_rows": 0, "summary_updated": 0, "error": "缺少商品列"}

    ad_agg = {}  # key: (asin, year, month)
    row_count = 0

    for row in rows:
        if not row or not row[col_product]:
            continue

        product_field = str(row[col_product]).strip()
        asin = product_field.split("-")[0]

        # 使用导入时选择的年月，不从数据中解析
        ad_year, ad_month = import_year, import_month

        ad_spend = _safe_decimal(row[col_spend]) if col_spend is not None else Decimal("0")
        ad_sales = _safe_decimal(row[col_sales]) if col_sales is not None else Decimal("0")

        # 写入 raw_advertising
        raw_adv = RawAdvertising(
            country_id=country_obj.id,
            store_id=store_id,
            product_field=_safe_str(product_field, 200),
            asin=_safe_str(asin, 50),
            status_val=_safe_str(row[col_status], 50) if col_status is not None and row[col_status] else "",
            ad_type=_safe_str(row[col_type], 50) if col_type is not None and row[col_type] else "",
            eligibility=_safe_str(row[col_elig], 100) if col_elig is not None and row[col_elig] else "",
            sales_usd=ad_sales,
            roas=_safe_decimal(row[col_roas]) if col_roas is not None else Decimal("0"),
            conversion_rate=_safe_decimal(row[col_conv]) if col_conv is not None else Decimal("0"),
            impressions=_safe_int(row[col_imp]) if col_imp is not None else 0,
            clicks=_safe_int(row[col_clicks]) if col_clicks is not None else 0,
            ctr=_safe_decimal(row[col_ctr]) if col_ctr is not None else Decimal("0"),
            spend_usd=ad_spend,
            cpc=_safe_decimal(row[col_cpc]) if col_cpc is not None else Decimal("0"),
            orders=_safe_int(row[col_orders]) if col_orders is not None else 0,
            acos=_safe_decimal(row[col_acos]) if col_acos is not None else Decimal("0"),
            ntb_orders=_safe_int(row[col_ntb_orders]) if col_ntb_orders is not None else 0,
            ntb_order_pct=_safe_decimal(row[col_ntb_pct]) if col_ntb_pct is not None else Decimal("0"),
            ntb_sales_usd=_safe_decimal(row[col_ntb_sales]) if col_ntb_sales is not None else Decimal("0"),
            new_to_brand_sales_pct=_safe_decimal(row[col_new_brand]) if col_new_brand is not None else Decimal("0"),
            visible_impressions=_safe_int(row[col_vis_imp]) if col_vis_imp is not None else 0,
            raw_data=_json_safe(header, row),
            time_id=time_id,
        )
        db.add(raw_adv)

        # 按月聚合
        ym_key = (asin, ad_year, ad_month) if ad_year and ad_month else (asin, None, None)
        if ym_key not in ad_agg:
            ad_agg[ym_key] = {"ad_spend": Decimal("0"), "ad_sales": Decimal("0"), "acos": [], "roas": [], "ctr": [], "cpc": [], "imp": 0, "clicks": 0, "orders": 0, "conv": []}
        agg = ad_agg[ym_key]
        agg["ad_spend"] += ad_spend
        agg["ad_sales"] += ad_sales
        if col_acos is not None and row[col_acos]: agg["acos"].append(_safe_decimal(row[col_acos]))
        if col_roas is not None and row[col_roas]: agg["roas"].append(_safe_decimal(row[col_roas]))
        if col_ctr is not None and row[col_ctr]: agg["ctr"].append(_safe_decimal(row[col_ctr]))
        if col_cpc is not None and row[col_cpc]: agg["cpc"].append(_safe_decimal(row[col_cpc]))
        if col_imp is not None and row[col_imp]: agg["imp"] += _safe_int(row[col_imp])
        if col_clicks is not None and row[col_clicks]: agg["clicks"] += _safe_int(row[col_clicks])
        if col_orders is not None and row[col_orders]: agg["orders"] += _safe_int(row[col_orders])
        if col_conv is not None and row[col_conv]: agg["conv"].append(_safe_decimal(row[col_conv]))
        row_count += 1

    summary_count = 0

    for (asin, ad_year, ad_month), agg in ad_agg.items():
        # 严格按店铺+月份隔离查产品
        product = _find_product_by_asin(db, asin, store_id=store_id, year_month=f"{ad_year}-{ad_month:02d}" if ad_year and ad_month else None)
        if not product:
            continue

        exchange_rate = _get_exchange_rate(db, country_obj, import_year, import_month, store_id=store_id)

        # 确定目标月份（以导入时选择的月份为准，不看数据内日期）
        if import_year and import_month:
            time_obj = _get_or_create_time(db, import_year, import_month)
            time_id = time_obj.id
            summary = db.query(MonthlySummary).filter(
                MonthlySummary.product_id == product.id,
                MonthlySummary.country_id == country_obj.id,
                MonthlySummary.time_id == time_id,
                MonthlySummary.store_id == store_id,
            ).first() if store_id else db.query(MonthlySummary).filter(
                MonthlySummary.product_id == product.id,
                MonthlySummary.country_id == country_obj.id,
                MonthlySummary.time_id == time_id,
            ).first()
            if not summary:
                # 仅有广告数据的产品：补建 summary
                summary = MonthlySummary(
                    country_id=country_obj.id, product_id=product.id, time_id=time_id,
                    store_id=store_id, order_count=0, order_qty=0,
                    product_sales_usd=Decimal("0"), commission_usd=Decimal("0"), fba_fee_usd=Decimal("0"),
                    ad_spend_usd=Decimal("0"), storage_fee_usd=Decimal("0"), returns_fee_usd=Decimal("0"), inbound_fee_usd=Decimal("0"),
                    removal_fee_usd=Decimal("0"),
                )
                db.add(summary)
                db.flush()
            target_summaries = [summary]
        else:
            # 无时间信息，更新该国家所有月份
            target_summaries = db.query(MonthlySummary).filter(
                MonthlySummary.product_id == product.id,
                MonthlySummary.country_id == country_obj.id,
            ).all()

        for summary in target_summaries:
            summary.ad_spend_usd = agg["ad_spend"]
            summary.ad_sales_usd = agg["ad_sales"]
            summary.acos = (sum(agg["acos"]) / len(agg["acos"])).quantize(Decimal("0.0001")) if agg["acos"] else Decimal("0")
            summary.roas = (sum(agg["roas"]) / len(agg["roas"])).quantize(Decimal("0.0001")) if agg["roas"] else Decimal("0")
            summary.ctr = (sum(agg["ctr"]) / len(agg["ctr"])).quantize(Decimal("0.0001")) if agg["ctr"] else Decimal("0")
            summary.cpc = (sum(agg["cpc"]) / len(agg["cpc"])).quantize(Decimal("0.01")) if agg["cpc"] else Decimal("0")
            summary.impressions = agg["imp"]
            summary.clicks = agg["clicks"]
            summary.ad_orders = agg["orders"]
            summary.conversion_rate = (sum(agg["conv"]) / len(agg["conv"])).quantize(Decimal("0.0001")) if agg["conv"] else Decimal("0")
            summary_count += 1

    return {"csv_rows": row_count, "summary_updated": summary_count}


def _process_fee_sheet(db, country_obj, header, rows, fee_type, store_id=None, import_year=None, import_month=None):
    """处理费用类 sheet（仓储/退货/入库/长期仓储），使用导入时选择的年月"""
    col_asin = _find_col(header, "asin", "ASIN")
    col_fee = None

    if fee_type == "storage":
        col_fee = _find_col(header, "estimated_monthly_storage_fee", "amount-charged", "amount_charged", "storage_fee", "月度仓储费（预计）")
    elif fee_type == "returns":
        col_fee = _find_col(header, "sku_returns_fee", "returns_fee")
    elif fee_type == "inbound":
        col_fee = _find_col(header, "入库配置服务费用总计", "inbound_fee", "总费用")
    elif fee_type == "long_term_storage":
        col_fee = _find_col(header, "amount-charged", "amount_charged")

    if col_asin is None or col_fee is None:
        return {"csv_rows": 0, "summary_updated": 0, "error": f"缺少必要列 (asin={col_asin}, fee={col_fee})"}

    # 按行识别国家的列
    col_row_country = _find_col(header, "country_code", "country", "国家/地区")
    # 国家缓存：country_code -> country_obj
    _country_cache = {}
    def _get_row_country(row):
        """从行中读取国家代码，返回 country_obj；无法识别返回 None（不 fallback，避免错配到错误国家）"""
        if col_row_country is not None and row[col_row_country]:
            cc = str(row[col_row_country]).strip().upper()
            if cc in _country_cache:
                return _country_cache[cc]
            # 映射：US/USA -> US, CA/CAN -> CA, MX/MEX -> MX
            cc_map = {'US': 'US', 'USA': 'US', 'CA': 'CA', 'CAN': 'CA', 'MX': 'MX', 'MEX': 'MX',
                      'UK': 'UK', 'GB': 'UK', 'DE': 'DE', 'DEU': 'DE', 'AU': 'AU', 'AUS': 'AU',
                      'FR': 'FR', 'FRA': 'FR', 'ES': 'ES', 'ESP': 'ES', 'IT': 'IT', 'ITA': 'IT',
                      'NL': 'NL', 'NLD': 'NL', 'BE': 'BE', 'BEL': 'BE', 'IE': 'IE', 'IRL': 'IE',
                      'SE': 'SE', 'SWE': 'SE', 'AE': 'AE', 'ARE': 'AE', 'SA': 'SA', 'SAU': 'SA'}
            code = cc_map.get(cc, cc)
            co = db.query(DimCountry).filter(DimCountry.code == code).first()
            _country_cache[cc] = co
            if co:
                return co
        # 无法识别：返回 None，由调用方决定（单国家sheet用 country_obj，多国家sheet跳过该行）
        return None

    asin_fees = {}  # (country_id, asin, month_str) -> Decimal
    asin_names = {}  # asin -> product_name（从原始数据中收集，用于回填空名称）
    row_count = 0
    # 使用导入时选择的年月，不从数据中解析
    _import_ym = f"{import_year}-{import_month:02d}" if import_year and import_month else None
    col_moc = _find_col(header, "month_of_charge", "交易日期", "snapshot-date", "收费月份")

    for row in rows:
        if not row or not row[col_asin]:
            continue
        asin = str(row[col_asin]).strip()
        if not asin or asin.startswith("Amazon."):
            continue
        fee = _safe_decimal(row[col_fee]) if row[col_fee] else Decimal("0")

        # 按行确定国家
        row_country = _get_row_country(row)
        # 无法从行识别国家且无有效 fallback → 跳过该行（避免错配到错误国家）
        if row_country is None:
            if country_obj is not None:
                row_country = country_obj  # 单国家sheet：用传入的国家
            else:
                continue  # 多国家sheet：无法识别国家，跳过

        # 确定月份：使用导入时选择的年月
        month_str = _import_ym or ""

        # 写入对应的 raw 表
        if fee_type == "storage":
            col_fnsku = _find_col(header, "fnsku")
            col_pname = _find_col(header, "product_name", "product-name")
            col_fc = _find_col(header, "fulfillment_center", "亚马逊运营中心")
            col_cc = _find_col(header, "country_code", "国家/地区代码")
            col_tier = _find_col(header, "product_size_tier", "商品尺寸分段")
            col_moc = _find_col(header, "month_of_charge", "收费月份")
            col_currency = _find_col(header, "currency")

            raw = RawStorageFee(
                country_id=row_country.id,
                store_id=store_id,
                asin=_safe_str(asin, 50),
                fnsku=_safe_str(row[col_fnsku], 50) if col_fnsku is not None and row[col_fnsku] else "",
                product_name=_safe_str(row[col_pname], 500) if col_pname is not None and row[col_pname] else "",
                fulfillment_center=_safe_str(row[col_fc], 100) if col_fc is not None and row[col_fc] else "",
                country_code=_safe_str(row[col_cc], 20) if col_cc is not None and row[col_cc] else "",
                product_size_tier=_safe_str(row[col_tier], 50) if col_tier is not None and row[col_tier] else "",
                month_of_charge=_safe_str(row[col_moc], 30) if col_moc is not None and row[col_moc] else "",
                currency=_safe_str(row[col_currency], 50) if col_currency is not None and row[col_currency] else "",
                estimated_monthly_storage_fee=fee,
                raw_data=_json_safe(header, row),
            )
            db.add(raw)
            # 收集产品名称用于回填
            if col_pname is not None and row[col_pname]:
                pname = _safe_str(row[col_pname], 500)
                if pname and asin not in asin_names:
                    asin_names[asin] = pname

        elif fee_type == "returns":
            col_fnsku = _find_col(header, "fnsku")
            col_pname = _find_col(header, "product_name")
            col_cat = _find_col(header, "asin_fee_category")
            col_mos = _find_col(header, "month_of_shipment")
            col_moc = _find_col(header, "month_of_charge")
            col_currency = _find_col(header, "currency")
            col_shipped = _find_col(header, "asin_shipped_units")
            col_ret_units = _find_col(header, "asin_returned_units")
            col_fee_per = _find_col(header, "sku_fee_per_unit")

            raw = RawReturns(
                country_id=row_country.id,
                store_id=store_id,
                asin=_safe_str(asin, 50),
                asin_fee_category=_safe_str(row[col_cat], 50) if col_cat is not None and row[col_cat] else "",
                fnsku=_safe_str(row[col_fnsku], 50) if col_fnsku is not None and row[col_fnsku] else "",
                product_name=_safe_str(row[col_pname], 500) if col_pname is not None and row[col_pname] else "",
                month_of_shipment=_safe_str(row[col_mos], 30) if col_mos is not None and row[col_mos] else "",
                asin_shipped_units=_safe_int(row[col_shipped]) if col_shipped is not None else 0,
                asin_returned_units=_safe_int(row[col_ret_units]) if col_ret_units is not None else 0,
                sku_fee_per_unit=_safe_decimal(row[col_fee_per]) if col_fee_per is not None else Decimal("0"),
                sku_returns_fee=fee,
                month_of_charge=_safe_str(row[col_moc], 30) if col_moc is not None and row[col_moc] else "",
                currency=_safe_str(row[col_currency], 50) if col_currency is not None and row[col_currency] else "",
                raw_data=_json_safe(header, row),
            )
            db.add(raw)
            if col_pname is not None and row[col_pname]:
                pname = _safe_str(row[col_pname], 500)
                if pname and asin not in asin_names:
                    asin_names[asin] = pname

        elif fee_type == "inbound":
            col_fnsku = _find_col(header, "fnsku", "FNSKU")
            col_date = _find_col(header, "交易日期")
            col_plan = _find_col(header, "入库计划编号")
            col_shipment = _find_col(header, "亚马逊物流货件编号")
            col_country = _find_col(header, "国家/地区")
            col_currency = _find_col(header, "货币")
            col_total = _find_col(header, "总费用")

            txn_date_str = str(row[col_date]).strip() if col_date is not None and row[col_date] else ""
            txn_date_val = _detect_date_format(txn_date_str) if txn_date_str else None

            raw = RawInbound(
                country_id=row_country.id,
                store_id=store_id,
                transaction_date=txn_date_val,
                inbound_plan_id=_safe_str(row[col_plan], 50) if col_plan is not None and row[col_plan] else "",
                fba_shipment_id=_safe_str(row[col_shipment], 50) if col_shipment is not None and row[col_shipment] else "",
                country_region=_safe_str(row[col_country], 50) if col_country is not None and row[col_country] else "",
                fnsku=_safe_str(row[col_fnsku], 50) if col_fnsku is not None and row[col_fnsku] else "",
                asin=_safe_str(asin, 50),
                inbound_placement_fee_total=fee,
                currency=_safe_str(row[col_currency], 50) if col_currency is not None and row[col_currency] else "",
                total_fee=_safe_decimal(row[col_total]) if col_total is not None else Decimal("0"),
                raw_data=_json_safe(header, row),
            )
            db.add(raw)

        elif fee_type == "long_term_storage":
            # 用表头映射代替列索引，防止空单元格导致列错位
            row_dict = {}
            for i, h in enumerate(header):
                if h and i < len(row) and row[i] is not None:
                    v = row[i]
                    if isinstance(v, (datetime, date)):
                        v = v.isoformat()
                    elif isinstance(v, Decimal):
                        v = str(v)
                    row_dict[h] = v

            raw = RawLongTermStorage(
                country_id=row_country.id,
                store_id=store_id,
                snapshot_date=_safe_str(row_dict.get("snapshot-date", row_dict.get("snapshot_date", "")), 30),
                sku=_safe_str(row_dict.get("sku", ""), 100),
                fnsku=_safe_str(row_dict.get("fnsku", ""), 50),
                asin=_safe_str(asin, 50),
                product_name=_safe_str(row_dict.get("product-name", row_dict.get("product_name", "")), 500),
                condition_val=_safe_str(row_dict.get("condition", ""), 50),
                per_unit_volume=_safe_decimal(row_dict.get("per-unit-volume", row_dict.get("per_unit_volume", 0))),
                currency=_safe_str(row_dict.get("currency", ""), 50),
                volume_unit=_safe_str(row_dict.get("volume-unit", row_dict.get("volume_unit", "")), 50),
                country=_safe_str(row_dict.get("country", ""), 50),
                qty_charged=_safe_int(row_dict.get("qty-charged", row_dict.get("qty_charged", 0))),
                amount_charged=fee,
                surcharge_age_tier=_safe_str(row_dict.get("surcharge-age-tier", row_dict.get("surcharge_age_tier", "")), 50),
                rate_surcharge=_safe_decimal(row_dict.get("rate-surcharge", row_dict.get("rate_surcharge", 0))),
                raw_data=row_dict,  # row_dict 本身就是 header→value 的映射
            )
            db.add(raw)
            pname_lts = _safe_str(row_dict.get("product-name", row_dict.get("product_name", "")), 500)
            if pname_lts and asin not in asin_names:
                asin_names[asin] = pname_lts

        key = (row_country.id, asin, month_str)
        if key not in asin_fees:
            asin_fees[key] = Decimal("0")
        asin_fees[key] += fee
        row_count += 1

    # 回填空产品名称（严格按店铺隔离）
    if asin_names:
        for asin, pname in asin_names.items():
            product = db.query(DimProduct).filter(
                DimProduct.asin == asin,
                DimProduct.store_id == store_id,
            ).first()
            if product and not product.product_name:
                product.product_name = pname

    summary_count = 0

    for (fee_country_id, asin, month_str), total_fee in asin_fees.items():
        fee_ym = f"{import_year}-{import_month:02d}" if import_year and import_month else None
        product = _find_product_by_asin(db, asin, store_id=store_id, year_month=fee_ym)
        if not product:
            continue

        # 查找对应月份的 summary（支持 Apr-26 / 2026-05 等多种格式）
        time_obj = None
        if month_str:
            time_obj = _find_time_by_month_str(db, month_str)

        if time_obj:
            summary = db.query(MonthlySummary).filter(
                MonthlySummary.product_id == product.id,
                MonthlySummary.country_id == fee_country_id,
                MonthlySummary.time_id == time_obj.id,
                MonthlySummary.store_id == store_id,
            ).first() if store_id else db.query(MonthlySummary).filter(
                MonthlySummary.product_id == product.id,
                MonthlySummary.country_id == fee_country_id,
                MonthlySummary.time_id == time_obj.id,
            ).first()
            if not summary:
                # 仅有费用数据的产品：补建 summary
                summary = MonthlySummary(
                    country_id=fee_country_id, product_id=product.id, time_id=time_obj.id,
                    store_id=store_id, order_count=0, order_qty=0,
                    product_sales_usd=Decimal("0"), commission_usd=Decimal("0"), fba_fee_usd=Decimal("0"),
                    ad_spend_usd=Decimal("0"), storage_fee_usd=Decimal("0"), returns_fee_usd=Decimal("0"), inbound_fee_usd=Decimal("0"),
                    removal_fee_usd=Decimal("0"),
                )
                db.add(summary)
                db.flush()
            target_summaries = [summary]
        else:
            # 没有月份信息，更新该国家所有月份
            target_summaries = db.query(MonthlySummary).filter(
                MonthlySummary.product_id == product.id,
                MonthlySummary.country_id == fee_country_id,
            ).all()

        for summary in target_summaries:
            if fee_type == "storage" or fee_type == "long_term_storage":
                summary.storage_fee_usd = (summary.storage_fee_usd or Decimal("0")) + total_fee
            elif fee_type == "returns":
                summary.returns_fee_usd = total_fee
            elif fee_type == "inbound":
                summary.inbound_fee_usd = total_fee
            summary_count += 1

    return {"csv_rows": row_count, "summary_updated": summary_count}


def _ensure_all_products_have_summary(db, country_obj, store_id=None, import_year=None, import_month=None):
    """为有实际原始数据的产品补建指定月份的 monthly_summary 记录（批量查询优化）"""
    if not import_year or not import_month:
        return
    time_obj = _get_or_create_time(db, import_year, import_month)

    # 批量查出有交易数据的 SKU 集合
    txn_q = db.query(RawTransaction.sku).filter(
        RawTransaction.country_id == country_obj.id,
        RawTransaction.sku.isnot(None),
        RawTransaction.sku != "",
    )
    if store_id:
        txn_q = txn_q.filter(RawTransaction.store_id == store_id)
    txn_skus = set(r[0] for r in txn_q.distinct().all())

    # 批量查出有广告数据的 ASIN 集合
    adv_q = db.query(RawAdvertising.asin).filter(
        RawAdvertising.country_id == country_obj.id,
        RawAdvertising.asin.isnot(None),
        RawAdvertising.asin != "",
    )
    if store_id:
        adv_q = adv_q.filter(RawAdvertising.store_id == store_id)
    adv_asins = set(r[0] for r in adv_q.distinct().all())

    # 批量查出有仓储/退货/入库/移除费数据的 ASIN 集合（这些费用也要算成本）
    cost_asins = set()
    for raw_model in [RawStorageFee, RawReturns, RawInbound]:
        q = db.query(raw_model.asin).filter(
            raw_model.country_id == country_obj.id,
            raw_model.asin.isnot(None),
            raw_model.asin != "",
        )
        if store_id:
            q = q.filter(raw_model.store_id == store_id)
        for r in q.distinct().all():
            cost_asins.add(r[0])

    # 移除费按 SKU 匹配
    cost_skus = set()
    rm_q = db.query(RawRemovalFee.sku).filter(
        RawRemovalFee.country_id == country_obj.id,
        RawRemovalFee.sku.isnot(None),
        RawRemovalFee.sku != "",
    )
    if store_id:
        rm_q = rm_q.filter(RawRemovalFee.store_id == store_id)
    for r in rm_q.distinct().all():
        cost_skus.add(r[0])

    if not txn_skus and not adv_asins and not cost_asins and not cost_skus:
        return

    # 批量查出已有 summary 的 product_id 集合
    existing_q = db.query(MonthlySummary.product_id).filter(
        MonthlySummary.country_id == country_obj.id,
        MonthlySummary.time_id == time_obj.id,
    )
    if store_id:
        existing_q = existing_q.filter(MonthlySummary.store_id == store_id)
    else:
        existing_q = existing_q.filter(MonthlySummary.store_id.is_(None))
    existing_pids = set(r[0] for r in existing_q.distinct().all())

    # 只处理有数据且无 summary 的产品（严格按店铺隔离，绝不跨店；同 SKU 只建一行）
    _pq = db.query(DimProduct).filter(DimProduct.asin.notlike("Amazon.%"))
    if store_id:
        _pq = _pq.filter(DimProduct.store_id == store_id)
    all_products = _pq.all()
    added = 0
    created_skus = set()  # 同 SKU 去重（一个 SKU 可能有多个 ASIN 变体）
    for product in all_products:
        if product.id in existing_pids:
            continue
        if product.sku and product.sku in created_skus:
            continue  # 同 SKU 已有行，跳过
        has_data = (product.sku in txn_skus) or (product.asin in adv_asins) or (product.asin in cost_asins) or (product.sku in cost_skus)
        if not has_data:
            continue
        if product.sku:
            created_skus.add(product.sku)
        db.add(MonthlySummary(
            country_id=country_obj.id, product_id=product.id, time_id=time_obj.id,
            store_id=store_id, order_count=0, order_qty=0,
            product_sales_usd=Decimal("0"), commission_usd=Decimal("0"), fba_fee_usd=Decimal("0"),
            ad_spend_usd=Decimal("0"), storage_fee_usd=Decimal("0"), returns_fee_usd=Decimal("0"),
            inbound_fee_usd=Decimal("0"), removal_fee_usd=Decimal("0"),
        ))
        added += 1
    if added:
        db.flush()


def _process_removal_fee_sheet(db, country_obj, header, rows, store_id=None, import_year=None, import_month=None):
    """处理移除费 sheet：写入 raw_removal_fee，按 SKU 匹配产品，汇总到 monthly_summary"""
    from models import RawRemovalFee, DimProduct, DimTime, MonthlySummary

    # 查找列索引
    col_request_date = _find_col(header, "request-date", "request_date")
    col_order_id = _find_col(header, "order-id", "order_id")
    col_order_source = _find_col(header, "order-source", "order_source")
    col_order_type = _find_col(header, "order-type", "order_type")
    col_service_speed = _find_col(header, "service-speed", "service_speed")
    col_order_status = _find_col(header, "order-status", "order_status")
    col_last_updated = _find_col(header, "last-updated-date", "last_updated_date")
    col_sku = _find_col(header, "sku")
    col_fnsku = _find_col(header, "fnsku")
    col_disposition = _find_col(header, "disposition")
    col_requested_qty = _find_col(header, "requested-quantity", "requested_quantity")
    col_cancelled_qty = _find_col(header, "cancelled-quantity", "cancelled_quantity")
    col_disposed_qty = _find_col(header, "disposed-quantity", "disposed_quantity")
    col_shipped_qty = _find_col(header, "shipped-quantity", "shipped_quantity")
    col_in_process_qty = _find_col(header, "in-process-quantity", "in_process_quantity")
    col_removal_fee = _find_col(header, "removal-fee", "removal_fee")
    col_currency = _find_col(header, "currency")
    col_country = _find_col(header, "国家", "country")

    if col_sku is None or col_removal_fee is None:
        return {"csv_rows": 0, "summary_updated": 0, "error": "缺少必要列 (sku/removal-fee)"}

    # 导入月份
    _import_ym = f"{import_year}-{import_month:02d}" if import_year and import_month else None

    # 国家缓存
    _country_cache = {}
    def _get_row_country(row):
        if col_country is not None and row[col_country]:
            cc = str(row[col_country]).strip().upper()
            if cc in _country_cache:
                return _country_cache[cc]
            cc_map = {'US': 'US', 'USA': 'US', 'CA': 'CA', 'CAN': 'CA', 'MX': 'MX', 'MEX': 'MX',
                      'UK': 'UK', 'GB': 'UK', 'DE': 'DE', 'DEU': 'DE', 'AU': 'AU', 'AUS': 'AU',
                      'FR': 'FR', 'FRA': 'FR', 'ES': 'ES', 'ESP': 'ES', 'IT': 'IT', 'ITA': 'IT',
                      'NL': 'NL', 'NLD': 'NL', 'BE': 'BE', 'BEL': 'BE', 'IE': 'IE', 'IRL': 'IE',
                      'SE': 'SE', 'SWE': 'SE', 'AE': 'AE', 'ARE': 'AE', 'SA': 'SA', 'SAU': 'SA'}
            code = cc_map.get(cc, cc)
            co = db.query(DimCountry).filter(DimCountry.code == code).first()
            _country_cache[cc] = co
            return co
        return None

    # 按 (country_id, sku, month) 汇总移除费
    sku_fees = {}  # (country_id, sku, month_str) -> Decimal
    row_count = 0

    for row in rows:
        if not row or not row[col_sku]:
            continue
        sku = str(row[col_sku]).strip()
        if not sku:
            continue
        fee = _safe_decimal(row[col_removal_fee]) if row[col_removal_fee] else Decimal("0")
        if fee == 0:
            continue

        # 按行确定国家
        row_country = _get_row_country(row)
        if row_country is None:
            if country_obj is not None:
                row_country = country_obj
            else:
                continue

        # 解析日期确定月份
        request_date_str = str(row[col_request_date]).strip() if col_request_date is not None and row[col_request_date] else ""
        request_date_val = _detect_date_format(request_date_str) if request_date_str else None
        month_str = _import_ym
        if not month_str and request_date_val:
            month_str = f"{request_date_val.year}-{request_date_val.month:02d}"

        # 写入 raw_removal_fee
        raw = RawRemovalFee(
            country_id=row_country.id,
            store_id=store_id,
            request_date=request_date_val,
            order_id=_safe_str(row[col_order_id], 50) if col_order_id is not None and row[col_order_id] else "",
            order_source=_safe_str(row[col_order_source], 200) if col_order_source is not None and row[col_order_source] else "",
            order_type=_safe_str(row[col_order_type], 50) if col_order_type is not None and row[col_order_type] else "",
            service_speed=_safe_str(row[col_service_speed], 50) if col_service_speed is not None and row[col_service_speed] else "",
            order_status=_safe_str(row[col_order_status], 50) if col_order_status is not None and row[col_order_status] else "",
            last_updated_date=_detect_date_format(str(row[col_last_updated]).strip()) if col_last_updated is not None and row[col_last_updated] else None,
            sku=_safe_str(sku, 100),
            fnsku=_safe_str(row[col_fnsku], 50) if col_fnsku is not None and row[col_fnsku] else "",
            disposition=_safe_str(row[col_disposition], 50) if col_disposition is not None and row[col_disposition] else "",
            requested_quantity=_safe_int(row[col_requested_qty]) if col_requested_qty is not None else 0,
            cancelled_quantity=_safe_int(row[col_cancelled_qty]) if col_cancelled_qty is not None else 0,
            disposed_quantity=_safe_int(row[col_disposed_qty]) if col_disposed_qty is not None else 0,
            shipped_quantity=_safe_int(row[col_shipped_qty]) if col_shipped_qty is not None else 0,
            in_process_quantity=_safe_int(row[col_in_process_qty]) if col_in_process_qty is not None else 0,
            removal_fee=fee,
            currency=_safe_str(row[col_currency], 10) if col_currency is not None and row[col_currency] else "",
            raw_data=_json_safe(header, row),
        )
        db.add(raw)

        # 汇总到 sku_fees
        key = (row_country.id, sku, month_str or "")
        sku_fees[key] = sku_fees.get(key, Decimal("0")) + fee
        row_count += 1

    db.flush()

    # 按 SKU 匹配产品，更新 monthly_summary 的 removal_fee_usd
    summary_updated = 0
    for (cid, sku, ym), total_fee in sku_fees.items():
        # SKU 匹配产品（严格按店铺+月份隔离，绝不跨店）
        product = db.query(DimProduct).filter(
            DimProduct.sku == sku,
            DimProduct.store_id == store_id,
            DimProduct.year_month == _import_ym,
        ).first()
        if not product:
            continue

        # 查找或创建 monthly_summary
        time_obj = db.query(DimTime).filter(DimTime.year_month == ym).first() if ym else None
        if not time_obj:
            continue

        summary = db.query(MonthlySummary).filter(
            MonthlySummary.country_id == cid,
            MonthlySummary.product_id == product.id,
            MonthlySummary.time_id == time_obj.id,
            MonthlySummary.store_id == store_id,
        ).first()
        if not summary:
            # 不存在则创建
            summary = MonthlySummary(
                country_id=cid,
                store_id=store_id,
                product_id=product.id,
                time_id=time_obj.id,
                removal_fee_usd=total_fee,
            )
            db.add(summary)
        else:
            summary.removal_fee_usd = (summary.removal_fee_usd or Decimal("0")) + total_fee
        summary_updated += 1

    db.flush()
    return {"csv_rows": row_count, "summary_updated": summary_updated}


def _recalculate_all_profit(db, country_obj, store_id=None, time_id=None):
    """重新计算 monthly_summary 的净利润（可按店铺+月份限制范围，批量预加载消除N+1查询）"""
    from sqlalchemy import text

    ms_filter = [MonthlySummary.country_id == country_obj.id]
    rt_filter = ["rt.country_id = :cid"]
    rt_params = {"cid": country_obj.id}
    if store_id:
        ms_filter.append(MonthlySummary.store_id == store_id)
        rt_filter.append("rt.store_id = :sid")
        rt_params["sid"] = store_id
    if time_id:
        ms_filter.append(MonthlySummary.time_id == time_id)
        rt_filter.append("rt.time_id = :tid")
        rt_params["tid"] = time_id

    # Step 1: SQL 补充 order_qty
    _step1_extra = ""
    _step1_params = {"country_id": country_obj.id}
    if store_id:
        _step1_extra += " AND ms.store_id = :store_id"
        _step1_params["store_id"] = store_id
    if time_id:
        _step1_extra += " AND ms.time_id = :time_id"
        _step1_params["time_id"] = time_id
    db.execute(text(f"""
        UPDATE monthly_summary ms
        JOIN dim_product dp ON dp.id = ms.product_id
        SET ms.order_qty = (
            SELECT COALESCE(SUM(
                CASE 
                    WHEN rt.transaction_type = 'Order' THEN ABS(rt.quantity)
                    WHEN rt.transaction_type = 'Adjustment' AND rt.total > 0 AND (rt.order_id IS NULL OR rt.order_id = '') THEN ABS(rt.quantity)
                    WHEN rt.transaction_type = 'Adjustment' AND rt.total < 0 AND (rt.order_id IS NULL OR rt.order_id = '') THEN -ABS(rt.quantity)
                    ELSE 0
                END
            ), 0)
            FROM raw_transactions rt
            WHERE rt.transaction_type IN ('Order', 'Adjustment')
              AND rt.country_id = :country_id
              AND rt.sku = dp.sku
              AND rt.time_id = ms.time_id
              AND (rt.store_id = ms.store_id OR (rt.store_id IS NULL AND ms.store_id IS NULL))
        )
        WHERE ms.country_id = :country_id{_step1_extra}
    """), _step1_params)

    # Step 2: 批量预加载 raw_transactions 数据
    raw_agg = {}
    from sqlalchemy import text as _text
    _raw_extra = ""
    _raw_params = {"cid": country_obj.id}
    if store_id:
        _raw_extra += " AND rt.store_id = :sid"
        _raw_params["sid"] = store_id
    if time_id:
        _raw_extra += " AND rt.time_id = :tid"
        _raw_params["tid"] = time_id
    raw_rows = db.execute(_text(f"""
        SELECT rt.sku, rt.time_id, rt.store_id,
               SUM(CASE WHEN rt.transaction_type IN ('Order', 'Refund') THEN rt.product_sales ELSE 0 END) as product_sales,
               SUM(CASE WHEN rt.transaction_type IN ('Order', 'Refund') THEN rt.product_sales_tax ELSE 0 END) as product_sales_tax,
               SUM(CASE WHEN rt.transaction_type IN ('Order', 'Refund') THEN rt.postage_credits ELSE 0 END) as postage_credits,
               SUM(CASE WHEN rt.transaction_type IN ('Order', 'Refund') THEN rt.shipping_credits_tax ELSE 0 END) as shipping_credits_tax,
               SUM(CASE WHEN rt.transaction_type IN ('Order', 'Refund') THEN rt.gift_wrap_credits ELSE 0 END) as gift_wrap_credits,
               SUM(CASE WHEN rt.transaction_type IN ('Order', 'Refund') THEN rt.giftwrap_credits_tax ELSE 0 END) as giftwrap_credits_tax,
               SUM(CASE WHEN rt.transaction_type IN ('Order', 'Refund') THEN rt.promotional_rebates ELSE 0 END) as promo_rebate,
               SUM(CASE WHEN rt.transaction_type IN ('Order', 'Refund') THEN rt.promotional_rebates_tax ELSE 0 END) as promo_rebate_tax,
               SUM(CASE WHEN rt.transaction_type IN ('Order', 'Refund') THEN rt.marketplace_withheld_tax ELSE 0 END) as marketplace_withheld_tax,
               SUM(CASE WHEN rt.transaction_type IN ('Order', 'Refund') THEN rt.selling_fee ELSE 0 END) as selling_fee,
               SUM(CASE WHEN rt.transaction_type IN ('Order', 'Refund') THEN rt.fba_fee ELSE 0 END) as fba_fee,
               SUM(CASE WHEN rt.transaction_type = 'Adjustment' THEN rt.total ELSE 0 END) as adj_total,
               COUNT(CASE WHEN rt.transaction_type = 'Order' AND rt.sku NOT LIKE 'amzn.gr.%' THEN 1 END) as order_cnt,
               SUM(CASE WHEN rt.transaction_type = 'Order' AND rt.sku NOT LIKE 'amzn.gr.%' THEN ABS(rt.quantity) ELSE 0 END) as order_qty_raw
        FROM raw_transactions rt
        WHERE rt.country_id = :cid{_raw_extra}
        GROUP BY rt.sku, rt.time_id, rt.store_id
    """), _raw_params).fetchall()

    for row in raw_rows:
        sku = row[0] or ""
        tid = row[1]
        sid = row[2]
        # amzn.gr.* 替换件映射到真实SKU
        real_sku = _extract_real_sku(sku)
        effective_sku = real_sku if real_sku else sku
        key = (effective_sku, tid, sid)
        if key not in raw_agg:
            raw_agg[key] = {
                "product_sales": Decimal("0"), "product_sales_tax": Decimal("0"),
                "postage_credits": Decimal("0"),
                "shipping_credits_tax": Decimal("0"),
                "gift_wrap_credits": Decimal("0"), "giftwrap_credits_tax": Decimal("0"),
                "promo_rebate": Decimal("0"), "promo_rebate_tax": Decimal("0"),
                "marketplace_withheld_tax": Decimal("0"),
                "selling_fee": Decimal("0"), "fba_fee": Decimal("0"),
                "adj_no_order": Decimal("0"),
                "order_cnt": 0,
                "order_qty_raw": 0,
            }
        agg = raw_agg[key]
        agg["product_sales"] += Decimal(str(row[3] or 0))
        agg["product_sales_tax"] += Decimal(str(row[4] or 0))
        agg["postage_credits"] += Decimal(str(row[5] or 0))
        agg["shipping_credits_tax"] += Decimal(str(row[6] or 0))
        agg["gift_wrap_credits"] += Decimal(str(row[7] or 0))
        agg["giftwrap_credits_tax"] += Decimal(str(row[8] or 0))
        agg["promo_rebate"] += Decimal(str(row[9] or 0))
        agg["promo_rebate_tax"] += Decimal(str(row[10] or 0))
        agg["marketplace_withheld_tax"] += Decimal(str(row[11] or 0))
        agg["selling_fee"] += Decimal(str(row[12] or 0))
        agg["fba_fee"] += Decimal(str(row[13] or 0))
        agg["adj_no_order"] += Decimal(str(row[14] or 0))
        agg["order_cnt"] += int(row[15] or 0)
        agg["order_qty_raw"] += int(row[16] or 0)

    # 产品表: id -> (asin, sku)
    product_map = {}
    for p in db.query(DimProduct).all():
        product_map[p.id] = p

    # 时间表: id -> year_month
    time_map = {}
    for t in db.query(DimTime).all():
        time_map[t.id] = t.year_month

    # 成本表: (product_id, year_month) -> (cost, freight) + product_id -> (cost, freight) fallback
    cost_by_ym = {}
    cost_first = {}
    for pc in db.query(DimProductCost).all():
        cost_by_ym[(pc.product_id, pc.year_month)] = (Decimal(str(pc.cost_rmb or 0)), Decimal(str(pc.freight_per_unit or 0)))
        if pc.product_id not in cost_first:
            cost_first[pc.product_id] = (Decimal(str(pc.cost_rmb or 0)), Decimal(str(pc.freight_per_unit or 0)))

    # 国家独立运费: (product_id, country_id, store_id) -> freight
    freight_map = {}
    for df in db.query(DimFreight).filter(DimFreight.country_id == country_obj.id).all():
        freight_map[(df.product_id, df.store_id)] = Decimal(str(df.freight_rmb))

    # 汇率表: (store_id, year_month) -> rate
    rate_map = {}
    for er in db.query(DimExchangeRate).filter(DimExchangeRate.country_id == country_obj.id).all():
        rate_map[(er.store_id, er.year_month)] = Decimal(str(er.rate))
    default_rate = Decimal(_get_exchange_rate(db, country_obj).__str__())

    # Step 3: 重算 summary（限定范围）
    _summary_q = db.query(MonthlySummary).filter(MonthlySummary.country_id == country_obj.id)
    if store_id:
        _summary_q = _summary_q.filter(MonthlySummary.store_id == store_id)
    if time_id:
        _summary_q = _summary_q.filter(MonthlySummary.time_id == time_id)
    summaries = _summary_q.all()

    processed_sku_keys = set()  # (sku, time_id, store_id) 去重：同 SKU 只算一次
    for summary in summaries:
        product = product_map.get(summary.product_id)
        if not product:
            continue

        # 同 SKU 去重：同一 (sku, time, store) 只处理第一行
        sku_key = (product.sku, summary.time_id, summary.store_id)
        if sku_key in processed_sku_keys:
            continue
        processed_sku_keys.add(sku_key)

        ym = time_map.get(summary.time_id)
        # 先清零所有亚马逊字段（防止无raw_agg数据的summary保留旧值）
        summary.product_sales_usd = Decimal("0")
        summary.product_sales_tax = Decimal("0")
        summary.postage_credits = Decimal("0")
        summary.shipping_credits_tax = Decimal("0")
        summary.gift_wrap_credits = Decimal("0")
        summary.giftwrap_credits_tax = Decimal("0")
        summary.commission_usd = Decimal("0")
        summary.fba_fee_usd = Decimal("0")
        summary.promo_rebate_usd = Decimal("0")
        summary.promo_rebate_tax_usd = Decimal("0")
        summary.marketplace_withheld_tax_usd = Decimal("0")
        summary.adjustment_usd = Decimal("0")

        # 提前获取 ra（order_qty_raw 和后续用）
        product = product_map.get(summary.product_id)
        sku = product.sku if product else None
        raw_key = (sku, summary.time_id, summary.store_id) if sku else None
        ra = raw_agg.get(raw_key, {}) if raw_key else {}

        # order_qty 用纯Order件数（Order里已包含后来退货的）
        if ra and "order_qty_raw" in ra:
            order_qty = ra["order_qty_raw"]
            summary.order_qty = order_qty
        else:
            order_qty = summary.order_qty or 0
            summary.order_qty = order_qty

        # 没有净订单时成本归零
        if (summary.order_count or 0) <= 0:
            summary.product_cost_rmb = Decimal("0")
            summary.freight_cost_rmb = Decimal("0")
            order_qty = 0
            summary.order_qty = 0

        # 成本查找（内存查找，无DB查询）
        cost_per_unit, freight_per_unit = Decimal("0"), Decimal("0")
        if ym:
            pc = cost_by_ym.get((summary.product_id, ym))
            if pc:
                cost_per_unit, freight_per_unit = pc
        if cost_per_unit == 0:
            pc = cost_first.get(summary.product_id)
            if pc:
                cost_per_unit, freight_per_unit = pc

        # 国家独立运费覆盖
        ff = freight_map.get((summary.product_id, summary.store_id))
        if ff is not None:
            freight_per_unit = ff

        # 汇率（内存查找）
        er = rate_map.get((summary.store_id, ym))
        if not er:
            er = default_rate

        summary.product_cost_rmb = (cost_per_unit * order_qty).quantize(Decimal("0.01"))
        summary.freight_cost_rmb = (freight_per_unit * order_qty).quantize(Decimal("0.01"))
        summary.exchange_rate = er

        # 从 raw 数据更新 order_count
        if ra:
            summary.order_count = ra.get("order_cnt", 0)

        raw_ps = ra.get("product_sales", Decimal(str(summary.product_sales_usd or 0)))
        summary.product_sales_usd = raw_ps
        summary.product_sales_tax = ra.get("product_sales_tax", Decimal(str(summary.product_sales_tax or 0)))
        summary.postage_credits = ra.get("postage_credits", Decimal(str(getattr(summary, 'postage_credits', 0) or 0)))
        summary.shipping_credits_tax = ra.get("shipping_credits_tax", Decimal(str(getattr(summary, 'shipping_credits_tax', 0) or 0)))
        summary.gift_wrap_credits = ra.get("gift_wrap_credits", Decimal(str(getattr(summary, 'gift_wrap_credits', 0) or 0)))
        summary.giftwrap_credits_tax = ra.get("giftwrap_credits_tax", Decimal(str(getattr(summary, 'giftwrap_credits_tax', 0) or 0)))
        summary.commission_usd = ra.get("selling_fee", Decimal(str(summary.commission_usd or 0)))
        summary.fba_fee_usd = ra.get("fba_fee", Decimal(str(summary.fba_fee_usd or 0)))
        summary.promo_rebate_usd = ra.get("promo_rebate", Decimal(str(summary.promo_rebate_usd or 0)))
        summary.promo_rebate_tax_usd = ra.get("promo_rebate_tax", Decimal(str(summary.promo_rebate_tax_usd or 0)))
        summary.marketplace_withheld_tax_usd = ra.get("marketplace_withheld_tax", Decimal(str(summary.marketplace_withheld_tax_usd or 0)))

        summary.product_sales_rmb = (raw_ps * er).quantize(Decimal("0.01"))
        summary.amazon_payout_usd = (
            raw_ps + summary.commission_usd + summary.fba_fee_usd
        ).quantize(Decimal("0.01"))
        # Adjustment 在清零后需重新赋值
        adj_val = ra.get("adj_no_order", Decimal("0"))
        summary.adjustment_usd = adj_val

        # 统一利润公式
        from services.profit import apply_profit_to_summary
        apply_profit_to_summary(summary, er,
            raw_product_sales=ra.get("product_sales", Decimal("0")),
            raw_product_sales_tax=ra.get("product_sales_tax", Decimal("0")),
            raw_postage_credits=ra.get("postage_credits", Decimal("0")),
            raw_shipping_credits_tax=ra.get("shipping_credits_tax", Decimal("0")),
            raw_gift_wrap_credits=ra.get("gift_wrap_credits", Decimal("0")),
            raw_giftwrap_credits_tax=ra.get("giftwrap_credits_tax", Decimal("0")),
            raw_promo_rebate=ra.get("promo_rebate", Decimal("0")),
            raw_promo_rebate_tax=ra.get("promo_rebate_tax", Decimal("0")),
            raw_marketplace_withheld_tax=ra.get("marketplace_withheld_tax", Decimal("0")),
            raw_selling_fee=ra.get("selling_fee", Decimal("0")),
            raw_fba_fee=ra.get("fba_fee", Decimal("0")),
            raw_adjustment=ra.get("adj_no_order", Decimal("0")),
        )


# ============================================================
# GET /supported: 返回支持的导入类型
# ============================================================
@router.get("/supported")
def get_supported_imports():
    return {
        "supported_types": [
            {
                "type": "transaction",
                "name": "交易记录",
                "endpoint": "/api/import/transaction",
                "method": "POST",
                "file_type": "CSV",
                "description": "Amazon Transaction Report，前9行元数据跳过，第10行表头",
                "params": ["country"],
            },
            {
                "type": "product-info",
                "name": "产品信息",
                "endpoint": "/api/import/product-info",
                "method": "POST",
                "file_type": "XLSX",
                "description": "产品信息表，列：ASIN, SKU, 产品, 颜色, 成本RMB, 产品运费/台, 汇率",
                "params": [],
            },
            {
                "type": "advertising",
                "name": "广告数据",
                "endpoint": "/api/import/advertising",
                "method": "POST",
                "file_type": "CSV",
                "description": "广告报告，列：商品, 花费(USD), 销售额(USD), ROAS, CTR, CPC, ACOS 等",
                "params": ["country"],
            },
            {
                "type": "storage",
                "name": "仓储费",
                "endpoint": "/api/import/storage",
                "method": "POST",
                "file_type": "CSV",
                "description": "FBA Inventory Storage 报告，按 ASIN 汇总 estimated_monthly_storage_fee",
                "params": ["country"],
            },
            {
                "type": "returns",
                "name": "退货费",
                "endpoint": "/api/import/returns",
                "method": "POST",
                "file_type": "CSV",
                "description": "退货报告，按 ASIN 汇总 sku_returns_fee",
                "params": ["country"],
            },
            {
                "type": "inbound",
                "name": "入库费",
                "endpoint": "/api/import/inbound",
                "method": "POST",
                "file_type": "CSV",
                "description": "入库配置费报告，按 ASIN 汇总入库配置服务费用总计",
                "params": ["country"],
            },
        ]
    }


# ============================================================
# POST /recalculate: 重算所有国家的净利润
# ============================================================
@router.post("/recalculate")
def recalculate_profit(
    country: str = Query(None, description="国家代码，空=全部国家"),
    db: Session = Depends(get_db),
):
    """重新计算 monthly_summary 的净利润（修复汇率/成本后刷新数据）"""
    try:
        if country:
            co = db.query(DimCountry).filter(DimCountry.code == country.upper()).first()
            if not co:
                return {"detail": f"国家 {country} 不存在"}
            countries = [co]
        else:
            countries = db.query(DimCountry).all()

        results = {}
        for co in countries:
            _recalculate_all_profit(db, co)
            # 统计该国家的汇总数据
            stats = db.query(
                func.sum(MonthlySummary.product_sales_rmb),
                func.sum(MonthlySummary.net_profit_rmb),
                func.count(),
            ).filter(MonthlySummary.country_id == co.id).first()
            results[co.code] = {
                "sales_rmb": round(float(stats[0] or 0), 2),
                "net_profit_rmb": round(float(stats[1] or 0), 2),
                "summary_count": int(stats[2] or 0),
            }
        db.commit()
        return {"message": "重算完成", "results": results}
    except Exception as e:
        db.rollback()
        return {"detail": str(e)}
