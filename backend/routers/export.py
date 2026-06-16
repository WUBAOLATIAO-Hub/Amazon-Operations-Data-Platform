"""数据导出 — 生成Excel文件供前端下载"""
from io import BytesIO
from urllib.parse import quote
from fastapi import APIRouter, Depends, Query as FastAPIQuery, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import func, and_
from database import get_db
from models import MonthlySummary, DimCountry, DimProduct, DimTime, DimProductCost, DimStore

router = APIRouter()


def _build_base_query(db, store_id=None, country_id=None, year=None, month=None):
    """复用 query.py 的筛选逻辑"""
    time_query = db.query(DimTime.id)
    if year:
        time_query = time_query.filter(DimTime.time_year == year)
    if month:
        time_query = time_query.filter(DimTime.time_month == month)
    time_ids = [r[0] for r in time_query.all()]

    filters = []
    if country_id:
        filters.append(MonthlySummary.country_id == country_id)
    if year or month:
        if not time_ids:
            filters.append(MonthlySummary.time_id == -1)
        else:
            filters.append(MonthlySummary.time_id.in_(time_ids))
    if store_id:
        filters.append(MonthlySummary.store_id == store_id)

    return time_ids, filters


@router.get("/monthly-summary")
def export_monthly_summary(
    country: str = FastAPIQuery("", description="国家代码"),
    year: int = FastAPIQuery(None),
    month: int = FastAPIQuery(None),
    store: str = FastAPIQuery(None),
    keyword: str = FastAPIQuery(None),
    db: Session = Depends(get_db),
):
    """导出产品明细为Excel"""
    try:
        country_id = None
        if country:
            country_obj = db.query(DimCountry).filter(DimCountry.code == country.upper()).first()
            if country_obj:
                country_id = country_obj.id
            else:
                raise HTTPException(status_code=400, detail="国家不存在")

        store_id = None
        if store:
            store_obj = db.query(DimStore).filter(DimStore.code == store).first()
            if store_obj:
                store_id = store_obj.id
            else:
                raise HTTPException(status_code=400, detail="店铺不存在")

        time_ids, filters = _build_base_query(db, store_id=store_id, country_id=country_id, year=year, month=month)

        ms_on = [MonthlySummary.product_id == DimProduct.id]
        if country_id:
            ms_on.append(MonthlySummary.country_id == country_id)
        if year or month:
            if not time_ids:
                ms_on.append(MonthlySummary.time_id == -1)
            else:
                ms_on.append(MonthlySummary.time_id.in_(time_ids))
        if store_id:
            ms_on.append(MonthlySummary.store_id == store_id)

        q = (
            db.query(
                DimProduct.product_name,
                DimProduct.asin,
                DimProduct.sku,
                DimProduct.color,
                func.coalesce(func.sum(MonthlySummary.order_count), 0).label("order_count"),
                func.coalesce(func.sum(MonthlySummary.product_sales_usd), 0).label("product_sales_usd"),
                func.coalesce(func.sum(MonthlySummary.product_sales_rmb), 0).label("product_sales_rmb"),
                func.coalesce(func.sum(MonthlySummary.commission_usd), 0).label("commission_usd"),
                func.coalesce(func.sum(MonthlySummary.fba_fee_usd), 0).label("fba_fee_usd"),
                func.coalesce(func.sum(MonthlySummary.ad_spend_usd), 0).label("ad_spend_usd"),
                func.coalesce(func.sum(MonthlySummary.storage_fee_usd), 0).label("storage_fee_usd"),
                func.coalesce(func.sum(MonthlySummary.returns_fee_usd), 0).label("returns_fee_usd"),
                func.coalesce(func.sum(MonthlySummary.inbound_fee_usd), 0).label("inbound_fee_usd"),
                func.coalesce(func.sum(MonthlySummary.removal_fee_usd), 0).label("removal_fee_usd"),
                func.coalesce(func.sum(MonthlySummary.promo_rebate_usd), 0).label("promo_rebate_usd"),
                func.coalesce(func.sum(MonthlySummary.product_cost_rmb), 0).label("product_cost_rmb"),
                func.coalesce(func.sum(MonthlySummary.freight_cost_rmb), 0).label("freight_cost_rmb"),
                func.coalesce(func.sum(MonthlySummary.net_profit_rmb), 0).label("net_profit_rmb"),
            )
            .join(MonthlySummary, and_(*ms_on))
            .filter(DimProduct.asin.notlike("Amazon.%"))
            .group_by(DimProduct.id, DimProduct.product_name, DimProduct.asin, DimProduct.sku, DimProduct.color)
            .having(func.sum(MonthlySummary.order_count) > 0)
        )

        if keyword:
            kw = f"%{keyword.strip()}%"
            q = q.filter(
                (DimProduct.asin.like(kw)) | (DimProduct.product_name.like(kw))
            )

        rows = q.all()

        # 批量查产品成本
        product_ids_raw = db.query(DimProduct.id, DimProduct.asin).filter(
            DimProduct.asin.in_([r.asin for r in rows])
        ).all() if rows else []
        asin_to_id = {a: pid for pid, a in product_ids_raw}
        cost_map = {}
        if product_ids_raw:
            costs = db.query(DimProductCost).filter(DimProductCost.product_id.in_([pid for pid, _ in product_ids_raw])).all()
            for pc in costs:
                if pc.product_id not in cost_map:
                    cost_map[pc.product_id] = {"cost_rmb": float(pc.cost_rmb or 0), "freight_per_unit": float(pc.freight_per_unit or 0)}

        # 构建 Excel
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, numbers

        wb = Workbook()
        ws = wb.active
        ws.title = "产品明细"

        headers = [
            "产品名称", "ASIN", "SKU", "颜色",
            "销量", "单价(¥)", "运费/台(¥)",
            "销售额($)", "销售额(¥)",
            "佣金($)", "FBA($)", "广告($)",
            "仓储+退货+入库+移除($)", "采购成本(¥)", "头程运费(¥)",
            "净利润(¥)", "净利率"
        ]

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_align = Alignment(horizontal="center", vertical="center")

        for col_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align

        # 数据行
        total_sales_rmb = 0
        total_net = 0
        for row_idx, r in enumerate(rows, 2):
            sales_usd = float(r.product_sales_usd or 0)
            sales_rmb = float(r.product_sales_rmb or 0)
            commission = float(r.commission_usd or 0)
            fba = float(r.fba_fee_usd or 0)
            ad = float(r.ad_spend_usd or 0)
            storage = float(r.storage_fee_usd or 0) + float(r.returns_fee_usd or 0) + float(r.inbound_fee_usd or 0) + float(r.removal_fee_usd or 0)
            promo = float(r.promo_rebate_usd or 0)
            cost_rmb = float(r.product_cost_rmb or 0)
            freight_rmb = float(r.freight_cost_rmb or 0)
            net = float(r.net_profit_rmb or 0)
            rate = round(net / sales_rmb * 100, 1) if sales_rmb > 0 else 0

            pc = cost_map.get(asin_to_id.get(r.asin), {})

            values = [
                r.product_name or "-",
                r.asin,
                r.sku,
                r.color or "-",
                int(r.order_count or 0),
                pc.get("cost_rmb", 0),
                pc.get("freight_per_unit", 0),
                round(sales_usd, 2),
                round(sales_rmb, 2),
                round(commission, 2),
                round(fba, 2),
                round(ad, 2),
                round(storage, 2),
                round(cost_rmb, 2),
                round(freight_rmb, 2),
                round(net, 2),
                f"{rate}%",
            ]

            total_sales_rmb += sales_rmb
            total_net += net

            for col_idx, v in enumerate(values, 1):
                ws.cell(row=row_idx, column=col_idx, value=v)

            # 利润列颜色
            profit_cell = ws.cell(row=row_idx, column=16)
            profit_cell.font = Font(color="3F8600" if net >= 0 else "CF1322", bold=True)
            rate_cell = ws.cell(row=row_idx, column=17)
            rate_cell.font = Font(color="3F8600" if net >= 0 else "CF1322")

        # 合计行
        if rows:
            total_row = len(rows) + 2
            ws.cell(row=total_row, column=1, value="合计").font = Font(bold=True)
            total_rate = round(total_net / total_sales_rmb * 100, 1) if total_sales_rmb > 0 else 0
            ws.cell(row=total_row, column=16, value=round(total_net, 2)).font = Font(bold=True, color="3F8600" if total_net >= 0 else "CF1322")
            ws.cell(row=total_row, column=17, value=f"{total_rate}%").font = Font(bold=True)

        # 列宽
        col_widths = [30, 15, 18, 8, 8, 10, 10, 12, 12, 10, 10, 10, 16, 12, 12, 12, 8]
        for i, w in enumerate(col_widths, 1):
            ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w

        # 冻结首行
        ws.freeze_panes = "A2"

        # 生成文件
        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)

        filename = "产品明细"
        if store:
            filename += f"_{store}"
        if year:
            filename += f"_{year}"
        if month:
            filename += f"{month:02d}"
        if country:
            filename += f"_{country}"

        encoded = quote(f"{filename}.xlsx")
        return StreamingResponse(
            buf,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename*=UTF-8''{encoded}"}
        )

    except HTTPException:
        raise
    except Exception as e:
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/country-summary")
def export_country_summary(
    store: str = FastAPIQuery(None),
    year: int = FastAPIQuery(None),
    month: int = FastAPIQuery(None),
    db: Session = Depends(get_db),
):
    """导出国家汇总为Excel"""
    try:
        store_id = None
        if store:
            store_obj = db.query(DimStore).filter(DimStore.code == store).first()
            if store_obj:
                store_id = store_obj.id
            else:
                raise HTTPException(status_code=400, detail="店铺不存在")

        _, filters = _build_base_query(db, store_id=store_id, year=year, month=month)

        q = (
            db.query(
                DimCountry.code.label("country_code"),
                DimCountry.name.label("country_name"),
                func.coalesce(func.sum(MonthlySummary.order_count), 0).label("order_count"),
                func.coalesce(func.sum(MonthlySummary.product_sales_usd), 0).label("product_sales_usd"),
                func.coalesce(func.sum(MonthlySummary.product_sales_rmb), 0).label("product_sales_rmb"),
                func.coalesce(func.sum(MonthlySummary.commission_usd), 0).label("commission_usd"),
                func.coalesce(func.sum(MonthlySummary.fba_fee_usd), 0).label("fba_fee_usd"),
                func.coalesce(func.sum(MonthlySummary.ad_spend_usd), 0).label("ad_spend_usd"),
                func.coalesce(func.sum(MonthlySummary.storage_fee_usd), 0).label("storage_fee_usd"),
                func.coalesce(func.sum(MonthlySummary.returns_fee_usd), 0).label("returns_fee_usd"),
                func.coalesce(func.sum(MonthlySummary.inbound_fee_usd), 0).label("inbound_fee_usd"),
                func.coalesce(func.sum(MonthlySummary.removal_fee_usd), 0).label("removal_fee_usd"),
                func.coalesce(func.sum(MonthlySummary.product_cost_rmb), 0).label("product_cost_rmb"),
                func.coalesce(func.sum(MonthlySummary.freight_cost_rmb), 0).label("freight_cost_rmb"),
                func.coalesce(func.sum(MonthlySummary.net_profit_rmb), 0).label("net_profit_rmb"),
            )
            .join(DimCountry, MonthlySummary.country_id == DimCountry.id)
        )

        if filters:
            q = q.filter(*filters)
        q = q.group_by(DimCountry.id, DimCountry.code, DimCountry.name)
        q = q.having(func.sum(MonthlySummary.order_count) > 0)
        rows = q.all()

        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill

        wb = Workbook()
        ws = wb.active
        ws.title = "国家汇总"

        headers = [
            "国家代码", "国家名称", "销量",
            "销售额($)", "销售额(¥)",
            "佣金($)", "FBA($)", "广告($)",
            "仓储($)", "退货($)", "入库($)", "移除费($)",
            "采购成本(¥)", "头程运费(¥)",
            "净利润(¥)", "净利率", "占比"
        ]

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        for col_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        total_sales_rmb = sum(float(r.product_sales_rmb or 0) for r in rows)

        for row_idx, r in enumerate(rows, 2):
            sales_rmb = float(r.product_sales_rmb or 0)
            net = float(r.net_profit_rmb or 0)
            rate = round(net / sales_rmb * 100, 1) if sales_rmb > 0 else 0
            pct = round(sales_rmb / total_sales_rmb * 100, 1) if total_sales_rmb > 0 else 0

            values = [
                r.country_code, r.country_name, int(r.order_count or 0),
                round(float(r.product_sales_usd or 0), 2), round(sales_rmb, 2),
                round(float(r.commission_usd or 0), 2), round(float(r.fba_fee_usd or 0), 2),
                round(float(r.ad_spend_usd or 0), 2), round(float(r.storage_fee_usd or 0), 2),
                round(float(r.returns_fee_usd or 0), 2), round(float(r.inbound_fee_usd or 0), 2),
                round(float(r.removal_fee_usd or 0), 2),
                round(float(r.product_cost_rmb or 0), 2), round(float(r.freight_cost_rmb or 0), 2),
                round(net, 2), f"{rate}%", f"{pct}%",
            ]

            for col_idx, v in enumerate(values, 1):
                ws.cell(row=row_idx, column=col_idx, value=v)

            ws.cell(row=row_idx, column=14).font = Font(color="3F8600" if net >= 0 else "CF1322", bold=True)

        # 列宽
        for i, w in enumerate([10, 12, 8, 12, 12, 10, 10, 10, 10, 10, 10, 10, 12, 12, 12, 8, 8], 1):
            ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w

        ws.freeze_panes = "A2"

        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)

        filename = "国家汇总"
        if store:
            filename += f"_{store}"
        if year:
            filename += f"_{year}"
        if month:
            filename += f"{month:02d}"

        encoded = quote(f"{filename}.xlsx")
        return StreamingResponse(
            buf,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename*=UTF-8''{encoded}"}
        )

    except HTTPException:
        raise
    except Exception as e:
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))
